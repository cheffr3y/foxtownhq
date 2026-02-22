import os
import re
import uuid
from collections import defaultdict
from io import BytesIO
from datetime import date, datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from dotenv import load_dotenv
from psycopg2.extras import RealDictCursor
from werkzeug.security import check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import get_db, init_app as init_db_app
from pypdf import PdfReader

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY')
init_db_app(app)

# Admin auth (single user for MVP)
def get_admin_config():
    username = os.getenv('ADMIN_USERNAME')
    password_hash = os.getenv('ADMIN_PASSWORD_HASH')
    if not username or not password_hash:
        return None
    return {
        'username': username,
        'password_hash': password_hash
    }

# Unit conversion helpers
UNIT_DEFS = {
    # Weight (base: g)
    'g': {'type': 'weight', 'to_base': 1.0, 'display': 'g', 'system': 'metric'},
    'kg': {'type': 'weight', 'to_base': 1000.0, 'display': 'kg', 'system': 'metric'},
    'oz': {'type': 'weight', 'to_base': 28.349523125, 'display': 'oz', 'system': 'imperial'},
    'lb': {'type': 'weight', 'to_base': 453.59237, 'display': 'lb', 'system': 'imperial'},
    # Volume (base: ml)
    'ml': {'type': 'volume', 'to_base': 1.0, 'display': 'ml', 'system': 'metric'},
    'l': {'type': 'volume', 'to_base': 1000.0, 'display': 'L', 'system': 'metric'},
    'tsp': {'type': 'volume', 'to_base': 4.92892159375, 'display': 'tsp', 'system': 'imperial'},
    'tbsp': {'type': 'volume', 'to_base': 14.78676478125, 'display': 'tbsp', 'system': 'imperial'},
    'fl oz': {'type': 'volume', 'to_base': 29.5735295625, 'display': 'fl oz', 'system': 'imperial'},
    'cup': {'type': 'volume', 'to_base': 236.5882365, 'display': 'cup', 'system': 'imperial'},
    'pt': {'type': 'volume', 'to_base': 473.176473, 'display': 'pt', 'system': 'imperial'},
    'qt': {'type': 'volume', 'to_base': 946.352946, 'display': 'qt', 'system': 'imperial'},
    'gal': {'type': 'volume', 'to_base': 3785.411784, 'display': 'gal', 'system': 'imperial'},
}

UNIT_ALIASES = {
    # Weight
    'g': 'g', 'gram': 'g', 'grams': 'g',
    'kg': 'kg', 'kilogram': 'kg', 'kilograms': 'kg',
    'oz': 'oz', 'ounce': 'oz', 'ounces': 'oz',
    'lb': 'lb', 'lbs': 'lb', 'pound': 'lb', 'pounds': 'lb',
    # Volume
    'ml': 'ml', 'milliliter': 'ml', 'milliliters': 'ml', 'millilitre': 'ml', 'millilitres': 'ml',
    'l': 'l', 'liter': 'l', 'liters': 'l', 'litre': 'l', 'litres': 'l',
    'tsp': 'tsp', 'teaspoon': 'tsp', 'teaspoons': 'tsp',
    'tbsp': 'tbsp', 'tablespoon': 'tbsp', 'tablespoons': 'tbsp',
    'floz': 'fl oz', 'fluidounce': 'fl oz', 'fluidounces': 'fl oz',
    'cup': 'cup', 'cups': 'cup',
    'pt': 'pt', 'pint': 'pt', 'pints': 'pt',
    'qt': 'qt', 'quart': 'qt', 'quarts': 'qt',
    'gal': 'gal', 'gallon': 'gal', 'gallons': 'gal',
}

SYSTEM_UNITS = {
    'metric': {
        'weight': ['kg', 'g'],
        'volume': ['l', 'ml'],
    },
    'imperial': {
        'weight': ['lb', 'oz'],
        'volume': ['gal', 'qt', 'pt', 'cup', 'fl oz'],
    }
}

def normalize_unit(unit):
    if not unit:
        return None
    key = ''.join(ch for ch in unit.lower() if ch.isalnum())
    return UNIT_ALIASES.get(key)

def normalize_recipe_type(value):
    if not value:
        return None
    key = value.strip().lower()
    if key in ('menu', 'plated', 'rm', 'plated recipe'):
        return 'menu'
    if key in ('batch', 'rb', 'sub', 'subrecipe', 'sub-recipe', 'prep'):
        return 'batch'
    return None

def infer_recipe_type(name, selected_type):
    normalized = normalize_recipe_type(selected_type)
    if normalized:
        return normalized
    if name:
        trimmed = name.strip().upper()
        if trimmed.startswith('RB '):
            return 'batch'
        if trimmed.startswith('RM '):
            return 'menu'
    return None

def to_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0

def parse_float_field(value, label, errors, required=False, min_value=None):
    text = (value or '').strip()
    if not text:
        if required:
            errors.append(f"{label} is required.")
        return None
    try:
        number = float(text)
    except ValueError:
        errors.append(f"{label} must be a number.")
        return None
    if min_value is not None and number < min_value:
        errors.append(f"{label} must be at least {min_value}.")
        return None
    return number

def generate_id(prefix):
    return f"{prefix}{uuid.uuid4().hex[:12]}"

PRICE_REFRESH_DAYS = 56
RECIPE_Q_FACTOR_PERCENT = to_float(os.getenv('RECIPE_Q_FACTOR_PERCENT') or 5)

RECIPE_TYPE_CHOICES = [
    ('menu', 'Plated (RM)'),
    ('batch', 'Batch (RB)')
]

INGREDIENT_CATEGORIES = [
    'Produce',
    'Herbs',
    'Dairy',
    'Meat',
    'Poultry',
    'Seafood',
    'Dry Goods',
    'Spices',
    'Baking',
    'Oils & Vinegars',
    'Condiments & Sauces',
    'Bread',
    'Frozen',
    'Beverages',
    'Packaging',
    'Disposables',
    'Cleaning',
    'Other'
]

RECIPE_CATEGORIES = [
    'Appetizers',
    'Salads',
    'Soups',
    'Sauces',
    'Dressings',
    'Marinades',
    'Sides',
    'Proteins',
    'Sandwiches',
    'Pasta',
    'Pizza',
    'Breads',
    'Desserts',
    'Breakfast/Brunch',
    'Snacks',
    'Beverages',
    'Stocks & Bases',
    'Batch/Prep',
    'Other'
]

def get_unit_system():
    system = request.args.get('units')
    if system in ('auto', 'metric', 'imperial'):
        session['unit_system'] = system
    return session.get('unit_system', 'auto')

def format_number(value, decimals=2):
    if value is None:
        return ''
    rounded = round(value, decimals)
    if abs(rounded - round(rounded)) < (10 ** -decimals):
        return str(int(round(rounded)))
    return f"{rounded:.{decimals}f}".rstrip('0').rstrip('.')

def convert_cost_per_unit(cost_per_unit, from_unit, to_unit):
    cost = to_float(cost_per_unit)
    from_canonical = normalize_unit(from_unit)
    to_canonical = normalize_unit(to_unit)
    if not from_canonical or not to_canonical:
        return cost
    if from_canonical == to_canonical:
        return cost
    from_def = UNIT_DEFS.get(from_canonical)
    to_def = UNIT_DEFS.get(to_canonical)
    if not from_def or not to_def:
        return cost
    if from_def['type'] != to_def['type']:
        return cost
    factor = to_def['to_base'] / from_def['to_base']
    return cost * factor

def convert_quantity(quantity, unit, system='auto'):
    quantity = to_float(quantity)
    canonical = normalize_unit(unit)
    if quantity is None or canonical not in UNIT_DEFS:
        return {
            'quantity': quantity,
            'unit': unit,
            'factor': 1.0,
            'converted': False,
            'type': None
        }

    unit_def = UNIT_DEFS[canonical]
    unit_type = unit_def['type']
    base_qty = quantity * unit_def['to_base']

    if system == 'auto':
        system = unit_def['system']

    candidates = SYSTEM_UNITS.get(system, {}).get(unit_type, [])
    if not candidates:
        return {
            'quantity': quantity,
            'unit': unit_def['display'],
            'factor': 1.0,
            'converted': False,
            'type': unit_type
        }

    chosen = candidates[-1]
    for candidate in candidates:
        candidate_qty = base_qty / UNIT_DEFS[candidate]['to_base']
        if candidate_qty >= 1:
            chosen = candidate
            break

    display_qty = base_qty / UNIT_DEFS[chosen]['to_base']
    factor = unit_def['to_base'] / UNIT_DEFS[chosen]['to_base']

    return {
        'quantity': display_qty,
        'unit': UNIT_DEFS[chosen]['display'],
        'factor': factor,
        'converted': chosen != canonical,
        'type': unit_type
    }

def smart_quantity(quantity, unit, system=None):
    unit_system = system or get_unit_system()
    result = convert_quantity(quantity, unit, unit_system)
    return {
        'quantity': format_number(result['quantity']),
        'unit': result['unit'] or unit or '',
        'converted': result['converted'],
        'factor': result['factor'],
        'type': result['type']
    }

@app.context_processor
def inject_helpers():
    return {
        'smart_quantity': smart_quantity,
        'unit_system': get_unit_system(),
        'recipe_type_choices': RECIPE_TYPE_CHOICES,
        'ingredient_categories': INGREDIENT_CATEGORIES,
        'recipe_categories': RECIPE_CATEGORIES,
        'venue_defaults': VENUE_DEFAULTS
    }

# Recipe helpers
def get_recipe_by_id(cur, recipe_id):
    cur.execute(
        "SELECT id, name, category, yield_qty, yield_unit, instructions, source_venue, equipment, recipe_type, menu_descriptor FROM recipes WHERE id = %s",
        (recipe_id,)
    )
    return cur.fetchone()

def get_recipe_weighted_options(cur, recipe_id):
    cur.execute("""
        SELECT rwo.*,
               CASE
                   WHEN rwo.item_type = 'ingredient' THEN i.name
                   WHEN rwo.item_type = 'recipe' THEN r.name
               END AS item_name,
               CASE
                   WHEN rwo.item_type = 'ingredient' THEN i.unit
                   WHEN rwo.item_type = 'recipe' THEN r.yield_unit
               END AS item_default_unit,
               CASE
                   WHEN rwo.item_type = 'ingredient' THEN i.cost_per_unit
                   ELSE NULL
               END AS ingredient_cost_per_unit
        FROM recipe_weighted_options rwo
        LEFT JOIN ingredients i ON rwo.item_type = 'ingredient' AND rwo.item_id = i.id
        LEFT JOIN recipes r ON rwo.item_type = 'recipe' AND rwo.item_id = r.id
        WHERE rwo.recipe_id = %s
        ORDER BY rwo.group_name, item_name
    """, (recipe_id,))
    return cur.fetchall()

def distribute_group_weights(option_rows, errors):
    groups = {}
    for row in option_rows:
        groups.setdefault(row['group_name'], []).append(row)

    for group_name, rows in groups.items():
        explicit_total = 0.0
        has_missing = False
        for row in rows:
            if row['weight_percent'] is None:
                has_missing = True
            else:
                explicit_total += row['weight_percent']

        if has_missing:
            errors.append(f'Weighted options in "{group_name}" must include a weight percent for every row.')
            continue

        if explicit_total > 100.0001:
            errors.append(f'Weighted options in "{group_name}" exceed 100%.')
            continue

        if abs(explicit_total - 100.0) > 0.0001:
            errors.append(f'Weighted options in "{group_name}" must total 100%.')

def parse_weighted_options_from_form(request, recipe_type, cur, errors):
    if recipe_type != 'menu':
        return []

    group_names = request.form.getlist('option_group_name[]')
    item_types = request.form.getlist('option_item_type[]')
    item_ids = request.form.getlist('option_item_id[]')
    item_names = request.form.getlist('option_item_name[]')
    quantities = request.form.getlist('option_qty[]')
    units = request.form.getlist('option_unit[]')
    weights = request.form.getlist('option_weight[]')

    max_len = max(
        len(group_names),
        len(item_types),
        len(item_ids),
        len(item_names),
        len(quantities),
        len(units),
        len(weights),
        0
    )

    option_rows = []
    for idx in range(max_len):
        group_name = (group_names[idx] if idx < len(group_names) else '').strip()
        item_type = (item_types[idx] if idx < len(item_types) else '').strip().lower()
        item_id = (item_ids[idx] if idx < len(item_ids) else '').strip()
        item_name = (item_names[idx] if idx < len(item_names) else '').strip()
        qty_raw = (quantities[idx] if idx < len(quantities) else '').strip()
        unit_raw = (units[idx] if idx < len(units) else '').strip()
        weight_raw = (weights[idx] if idx < len(weights) else '').strip()

        if not any([group_name, item_id, item_name, qty_raw, unit_raw, weight_raw]):
            continue

        if not group_name:
            errors.append('Each weighted option row needs a group name.')
        if not item_type:
            item_type = 'recipe'
        if item_type != 'recipe':
            errors.append('Weighted option type must be sub-recipe.')
        if not item_id and item_name:
            errors.append(f'Weighted option "{item_name}" was not found. Select it from the list.')
        if not item_id and not item_name:
            errors.append('Each weighted option row needs an item.')

        qty_value = parse_float_field(qty_raw, 'Weighted option quantity', errors, required=True, min_value=0.0001)
        unit_value = normalize_unit(unit_raw) or unit_raw
        if not unit_value:
            errors.append('Each weighted option row needs a unit.')

        weight_value = parse_float_field(weight_raw, 'Weighted option percent', errors, required=True, min_value=0.0)
        if weight_value is not None and weight_value > 100:
            errors.append('Weighted option percent cannot exceed 100.')

        option_rows.append({
            'group_name': group_name,
            'item_type': item_type,
            'item_id': item_id,
            'quantity': qty_value,
            'unit': unit_value,
            'weight_percent': weight_value
        })

    distribute_group_weights(option_rows, errors)
    return option_rows

def get_recipe_components(cur, recipe_id):
    cur.execute("""
        SELECT ri.*, ri.type, ri.quantity, ri.unit,
               CASE 
                   WHEN ri.type = 'ingredient' THEN i.name
                   WHEN ri.type = 'recipe' THEN r.name
               END as item_name,
               CASE
                   WHEN ri.type = 'ingredient' THEN i.category
                   WHEN ri.type = 'recipe' THEN r.category
               END as category,
               CASE
                   WHEN ri.type = 'ingredient' THEN i.cost_per_unit
                   ELSE NULL
               END as cost_per_unit,
               CASE
                   WHEN ri.type = 'ingredient' THEN i.unit
                   ELSE NULL
               END as ingredient_unit
        FROM recipe_ingredients ri
        LEFT JOIN ingredients i ON ri.type = 'ingredient' AND ri.item_id = i.id
        LEFT JOIN recipes r ON ri.type = 'recipe' AND ri.item_id = r.id
        WHERE ri.recipe_id = %s
        ORDER BY ri.type, item_name
    """, (recipe_id,))
    return cur.fetchall()

def compute_scale_ratio(component, sub_recipe):
    if not sub_recipe:
        return None
    qty = to_float(component.get('quantity'))
    yield_qty = to_float(sub_recipe.get('yield_qty'))
    if qty <= 0 or yield_qty <= 0:
        return None
    unit = (component.get('unit') or '').strip()
    yield_unit = (sub_recipe.get('yield_unit') or '').strip()
    if unit and yield_unit and unit != yield_unit:
        return None
    return qty / yield_qty

def apply_recipe_q_factor(total_cost):
    if total_cost is None:
        return total_cost
    q_percent = RECIPE_Q_FACTOR_PERCENT
    if q_percent <= 0:
        return total_cost
    return total_cost * (1 + (q_percent / 100))

def build_component_tree(cur, recipe_id, scale_ratio, depth, path, unit_system, apply_q_factor=True):
    if recipe_id in path:
        return [], 0, True
    path = path | {recipe_id}
    current_recipe = get_recipe_by_id(cur, recipe_id)

    components = []
    total_cost = 0
    has_cycle = False

    for component in get_recipe_components(cur, recipe_id):
        qty = to_float(component.get('quantity'))
        scaled_qty = qty * scale_ratio
        item = dict(component)
        item['scaled_quantity'] = scaled_qty
        item['scaled_quantity_display'] = format_number(scaled_qty)
        item['children'] = []
        item['sub_total_cost'] = None
        item['sub_cost_per_unit'] = None
        item['sub_cost_unit'] = None
        item['scale_note'] = None
        item['scale_ratio'] = None
        item['cycle'] = False

        display = smart_quantity(scaled_qty, component.get('unit'), unit_system)
        item['display_quantity'] = display['quantity']
        item['display_unit'] = display['unit']
        item['display_converted'] = display['converted']
        item['display_factor'] = display['factor']
        item['display_type'] = display['type']

        if component.get('type') == 'ingredient':
            ingredient_unit = component.get('ingredient_unit')
            cost_per_unit = convert_cost_per_unit(
                component.get('cost_per_unit'),
                ingredient_unit,
                component.get('unit')
            )
            item['cost_total'] = scaled_qty * cost_per_unit
            if item['display_factor']:
                item['display_cost_per_unit'] = cost_per_unit / item['display_factor']
            else:
                item['display_cost_per_unit'] = cost_per_unit
            total_cost += item['cost_total']
        else:
            sub_recipe = get_recipe_by_id(cur, component.get('item_id'))
            item['sub_recipe'] = sub_recipe
            child_ratio = compute_scale_ratio(component, sub_recipe)
            applied_ratio = child_ratio if child_ratio is not None else 1
            if child_ratio is None:
                if sub_recipe and sub_recipe.get('yield_qty') and sub_recipe.get('yield_unit') and component.get('unit'):
                    item['scale_note'] = 'Unit mismatch; showing full batch'
                elif sub_recipe and sub_recipe.get('yield_qty'):
                    item['scale_note'] = 'No unit match; showing full batch'
                else:
                    item['scale_note'] = 'No yield info; showing full batch'
            else:
                item['scale_note'] = 'Scaled to match requested quantity'

            item['scale_ratio'] = applied_ratio
            item['display_ratio'] = applied_ratio
            item['display_ratio_pct'] = round(applied_ratio * 100, 1)

            if depth < 3 and sub_recipe:
                child_items, child_cost, child_cycle = build_component_tree(
                    cur,
                    sub_recipe['id'],
                    scale_ratio * applied_ratio,
                    depth + 1,
                    path,
                    unit_system,
                    apply_q_factor=apply_q_factor
                )
                item['children'] = child_items
                item['sub_total_cost'] = child_cost
                item['cycle'] = child_cycle
                has_cycle = has_cycle or child_cycle
                total_cost += child_cost
            else:
                item['children'] = []
                item['sub_total_cost'] = 0

            if sub_recipe:
                sub_yield_qty = to_float(sub_recipe.get('yield_qty'))
                if sub_yield_qty > 0:
                    full_cost = get_recipe_total_cost(cur, sub_recipe['id'], unit_system)
                    cost_per_yield = full_cost / sub_yield_qty
                    display_yield = smart_quantity(sub_yield_qty, sub_recipe.get('yield_unit'), unit_system)
                    display_unit = display_yield.get('unit') or sub_recipe.get('yield_unit')
                    item['sub_cost_per_unit'] = convert_cost_per_unit(
                        cost_per_yield,
                        sub_recipe.get('yield_unit'),
                        display_unit
                    )
                    item['sub_cost_unit'] = display_unit

        components.append(item)

    # Weighted option groups are only applied to plated/menu recipes.
    if current_recipe and current_recipe.get('recipe_type') == 'menu':
        weighted_rows = get_recipe_weighted_options(cur, recipe_id)
        grouped_options = {}
        for row in weighted_rows:
            group_name = (row.get('group_name') or '').strip() or 'Options'
            grouped_options.setdefault(group_name, []).append(row)

        for group_name, options in grouped_options.items():
            group_item = {
                'id': f"optgrp_{recipe_id}_{group_name}",
                'type': 'option_group',
                'item_name': group_name,
                'category': 'Weighted options',
                'scaled_quantity': '',
                'scaled_quantity_display': '',
                'display_quantity': '',
                'display_unit': '',
                'display_converted': False,
                'display_factor': 1,
                'display_type': None,
                'children': [],
                'sub_total_cost': None,
                'sub_cost_per_unit': None,
                'sub_cost_unit': None,
                'scale_note': None,
                'scale_ratio': None,
                'cycle': False,
                'group_weight_total': 0,
                'raw_cost_total': 0,
                'cost_total': 0,
                'display_cost_per_unit': None
            }

            for option in options:
                opt_weight_pct = to_float(option.get('weight_percent'))
                opt_weight_ratio = opt_weight_pct / 100.0
                qty = to_float(option.get('quantity'))
                scaled_qty = qty * scale_ratio

                opt_item = {
                    'id': option.get('id'),
                    'type': option.get('item_type'),
                    'item_name': option.get('item_name') or 'Unknown option',
                    'category': 'Weighted option',
                    'scaled_quantity': scaled_qty,
                    'scaled_quantity_display': format_number(scaled_qty),
                    'children': [],
                    'sub_total_cost': None,
                    'sub_cost_per_unit': None,
                    'sub_cost_unit': None,
                    'scale_note': None,
                    'scale_ratio': None,
                    'cycle': False,
                    'weight_percent': opt_weight_pct,
                    'weighted_cost': 0,
                    'raw_cost_total': 0
                }

                display = smart_quantity(scaled_qty, option.get('unit'), unit_system)
                opt_item['display_quantity'] = display['quantity']
                opt_item['display_unit'] = display['unit']
                opt_item['display_converted'] = display['converted']
                opt_item['display_factor'] = display['factor']
                opt_item['display_type'] = display['type']

                raw_option_cost = 0
                if option.get('item_type') == 'ingredient':
                    source_unit = option.get('item_default_unit')
                    cost_per_unit = convert_cost_per_unit(
                        option.get('ingredient_cost_per_unit'),
                        source_unit,
                        option.get('unit')
                    )
                    raw_option_cost = scaled_qty * cost_per_unit
                    if opt_item['display_factor']:
                        opt_item['display_cost_per_unit'] = cost_per_unit / opt_item['display_factor']
                    else:
                        opt_item['display_cost_per_unit'] = cost_per_unit
                else:
                    sub_recipe = get_recipe_by_id(cur, option.get('item_id'))
                    opt_item['sub_recipe'] = sub_recipe
                    pseudo_component = {
                        'quantity': option.get('quantity'),
                        'unit': option.get('unit')
                    }
                    child_ratio = compute_scale_ratio(pseudo_component, sub_recipe)
                    applied_ratio = child_ratio if child_ratio is not None else 1
                    opt_item['scale_ratio'] = applied_ratio

                    if depth < 3 and sub_recipe:
                        child_items, child_cost, child_cycle = build_component_tree(
                            cur,
                            sub_recipe['id'],
                            scale_ratio * applied_ratio,
                            depth + 1,
                            path,
                            unit_system,
                            apply_q_factor=apply_q_factor
                        )
                        opt_item['children'] = child_items
                        opt_item['sub_total_cost'] = child_cost
                        opt_item['cycle'] = child_cycle
                        has_cycle = has_cycle or child_cycle
                        raw_option_cost = child_cost
                    elif sub_recipe:
                        raw_option_cost = get_recipe_total_cost(cur, sub_recipe['id'], unit_system) * (scale_ratio * applied_ratio)

                    if sub_recipe:
                        sub_yield_qty = to_float(sub_recipe.get('yield_qty'))
                        if sub_yield_qty > 0:
                            full_cost = get_recipe_total_cost(cur, sub_recipe['id'], unit_system)
                            cost_per_yield = full_cost / sub_yield_qty
                            display_yield = smart_quantity(sub_yield_qty, sub_recipe.get('yield_unit'), unit_system)
                            display_unit = display_yield.get('unit') or sub_recipe.get('yield_unit')
                            opt_item['sub_cost_per_unit'] = convert_cost_per_unit(
                                cost_per_yield,
                                sub_recipe.get('yield_unit'),
                                display_unit
                            )
                            opt_item['sub_cost_unit'] = display_unit

                weighted_cost = raw_option_cost * opt_weight_ratio
                opt_item['raw_cost_total'] = raw_option_cost
                opt_item['cost_total'] = weighted_cost
                opt_item['weighted_cost'] = weighted_cost

                group_item['children'].append(opt_item)
                group_item['group_weight_total'] += opt_weight_pct
                group_item['raw_cost_total'] += raw_option_cost
                group_item['cost_total'] += weighted_cost

            total_cost += group_item['cost_total']
            components.append(group_item)

    if apply_q_factor:
        total_cost = apply_recipe_q_factor(total_cost)

    return components, total_cost, has_cycle

def get_recipe_total_cost(cur, recipe_id, unit_system, apply_q_factor=True):
    _, total_cost, _ = build_component_tree(cur, recipe_id, 1, 0, set(), unit_system, apply_q_factor=apply_q_factor)
    return total_cost

def parse_menu_items(
    cur,
    unit_system,
    recipe_ids,
    batch_values,
    menu_price_values=None,
    target_percent_values=None,
    popularity_values=None,
    default_target_percent=None,
    section_values=None,
    descriptor_values=None
):
    menu_items = []
    total_cost = 0
    errors = []

    for idx, recipe_id in enumerate(recipe_ids):
        recipe_id = (recipe_id or '').strip()
        if not recipe_id:
            continue
        batches = to_float(batch_values[idx]) if idx < len(batch_values) else 0
        if batches <= 0:
            errors.append('Batch counts must be greater than 0.')
            continue

        recipe = get_recipe_by_id(cur, recipe_id)
        if not recipe:
            errors.append('One or more recipes could not be found.')
            continue

        menu_price = None
        if menu_price_values and idx < len(menu_price_values):
            menu_price = to_float(menu_price_values[idx])
            if menu_price <= 0:
                menu_price = None

        target_percent = None
        if target_percent_values and idx < len(target_percent_values):
            target_percent = to_float(target_percent_values[idx])
            if target_percent <= 0:
                target_percent = None
        if target_percent is None and default_target_percent is not None:
            target_percent = to_float(default_target_percent)
            if target_percent <= 0:
                target_percent = None

        base_cost = get_recipe_total_cost(cur, recipe_id, unit_system)
        item_total = base_cost * batches
        total_cost += item_total

        section = None
        if section_values and idx < len(section_values):
            section = (section_values[idx] or '').strip() or None

        descriptor = None
        if descriptor_values and idx < len(descriptor_values):
            descriptor = (descriptor_values[idx] or '').strip() or None
        if not descriptor:
            descriptor = (recipe.get('menu_descriptor') or '').strip() or None

        menu_items.append({
            'recipe': recipe,
            'recipe_id': recipe_id,
            'batches': batches,
            'base_cost': base_cost,
            'item_total': item_total,
            'menu_price': menu_price,
            'target_food_cost_percent': target_percent,
            'popularity_score': None,
            'menu_section': section,
            'menu_descriptor': descriptor
        })

    if not menu_items:
        errors.append('Add at least one recipe to calculate a cost.')

    return menu_items, total_cost, errors

def apply_menu_pricing(menu_items, default_target_percent=20):
    priced_line_count = 0
    total_food_cost_percent = 0
    total_menu_price = 0

    for item in menu_items:
        target = to_float(item.get('target_food_cost_percent'))
        if target <= 0:
            target = to_float(default_target_percent)
        if target <= 0:
            target = 20
        item['target_food_cost_percent'] = target

        suggested_price = None
        if target > 0:
            suggested_price = item['base_cost'] / (target / 100)
        item['suggested_price'] = suggested_price

        menu_price = to_float(item.get('menu_price'))
        if menu_price <= 0:
            menu_price = suggested_price
        item['effective_menu_price'] = menu_price

        if menu_price and menu_price > 0:
            item['food_cost_percent'] = (item['base_cost'] / menu_price) * 100
            total_food_cost_percent += item['food_cost_percent']
            total_menu_price += menu_price
            priced_line_count += 1
        else:
            item['food_cost_percent'] = None

    avg_food_cost_percent = None
    avg_menu_price = None
    if priced_line_count > 0:
        avg_food_cost_percent = total_food_cost_percent / priced_line_count
        avg_menu_price = total_menu_price / priced_line_count

    return {
        'avg_food_cost_percent': avg_food_cost_percent,
        'avg_menu_price': avg_menu_price,
        'priced_line_count': priced_line_count
    }

def group_menu_items(menu_items):
    grouped = {}
    for item in menu_items:
        section = (item.get('menu_section') or '').strip() or 'Uncategorized'
        grouped.setdefault(section, []).append(item)

    ordered_sections = []
    for section in MENU_SECTION_OPTIONS:
        if section in grouped:
            ordered_sections.append(section)
    for section in sorted(grouped.keys(), key=lambda value: value.lower()):
        if section not in ordered_sections:
            ordered_sections.append(section)

    grouped_list = []
    for section in ordered_sections:
        items = grouped.get(section, [])
        items_sorted = sorted(
            items,
            key=lambda item: (
                (item.get('menu_descriptor') or '').lower(),
                (item.get('recipe', {}).get('name') or '').lower()
            )
        )
        grouped_list.append({'section': section, 'items': items_sorted})

    return grouped_list

def format_display_value(value, precision=2):
    if value is None:
        return ''
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.{precision}f}".rstrip('0').rstrip('.')

def flatten_components_for_pdf(items, level=0):
    rows = []
    for item in items:
        name = item.get('item_name') or ''
        prefix = '  ' * level
        label = ''
        if item.get('type') == 'recipe':
            label = 'Sub-recipe'
        elif item.get('category'):
            label = item.get('category')

        quantity = item.get('display_quantity')
        if quantity is None:
            quantity = item.get('scaled_quantity_display') or item.get('scaled_quantity') or ''
        qty_text = format_display_value(quantity, precision=2)

        unit = item.get('display_unit') or item.get('unit') or ''
        rows.append({
            'name': f"{prefix}{name}",
            'qty': qty_text,
            'unit': unit,
            'notes': label
        })
        if item.get('children'):
            rows.extend(flatten_components_for_pdf(item.get('children'), level + 1))
    return rows

def flatten_components_for_packet(items, ingredient_map, level=0):
    rows = []
    for item in items:
        item_type = item.get('type') or 'component'
        name = item.get('item_name') or 'Unknown'
        quantity = item.get('display_quantity')
        if quantity in (None, ''):
            quantity = item.get('scaled_quantity_display') or item.get('scaled_quantity') or ''
        unit = item.get('display_unit') or item.get('unit') or ''
        ext_cost = to_float(item.get('cost_total'))

        g_code = ''
        vendor = ''
        vendor_code = ''
        category = item.get('category') or ''
        if item_type == 'ingredient':
            ing = ingredient_map.get(item.get('item_id')) or {}
            g_code = ing.get('g_code') or ''
            vendor = ing.get('vendor') or ''
            vendor_code = ing.get('vendor_code') or ''
            category = ing.get('category') or category

        rows.append({
            'type': item_type,
            'name': f"{'  ' * level}{name}",
            'quantity': quantity,
            'unit': unit,
            'ext_cost': ext_cost,
            'g_code': g_code,
            'vendor': vendor,
            'vendor_code': vendor_code,
            'category': category,
            'notes': item.get('scale_note') or ''
        })
        if item.get('children'):
            rows.extend(flatten_components_for_packet(item['children'], ingredient_map, level + 1))
    return rows

def collect_ingredient_usage_from_components(components, menu_label, usage_map):
    for item in components:
        if item.get('type') == 'ingredient' and item.get('item_id'):
            key = item.get('item_id')
            usage_map.setdefault(key, set()).add(menu_label)
        if item.get('children'):
            collect_ingredient_usage_from_components(item['children'], menu_label, usage_map)

def compute_q_factor(total_cost, q_factor_percent):
    q_percent = to_float(q_factor_percent)
    if q_percent < 0:
        q_percent = 0
    q_amount = total_cost * (q_percent / 100)
    grand_total = total_cost + q_amount
    return q_percent, q_amount, grand_total

def build_rollout_breakdown(cur, unit_system, menu_items):
    ingredient_totals = {}
    subrecipe_ids = set()

    for item in menu_items:
        recipe_id = item.get('recipe_id')
        batches = to_float(item.get('batches'))
        if not recipe_id or batches <= 0:
            continue
        components, _, _ = build_component_tree(cur, recipe_id, batches, 0, set(), unit_system)
        collect_ingredients_from_components(components, ingredient_totals)
        collect_subrecipes_from_components(components, subrecipe_ids)

    ingredient_master = []
    ingredient_total_cost = 0
    if ingredient_totals:
        ingredient_ids = list({ing_id for ing_id, _ in ingredient_totals.keys()})
        cur.execute("""
            SELECT id, name, unit, category, cost_per_unit, vendor, vendor_code, g_code
            FROM ingredients
            WHERE id = ANY(%s)
        """, (ingredient_ids,))
        ingredient_map = {row['id']: row for row in cur.fetchall()}

        for (ing_id, unit), qty in ingredient_totals.items():
            ingredient = ingredient_map.get(ing_id, {})
            cost_per_unit_raw = ingredient.get('cost_per_unit')
            cost_per_unit = convert_cost_per_unit(
                cost_per_unit_raw,
                ingredient.get('unit'),
                unit
            )
            ext_cost = qty * cost_per_unit if cost_per_unit else 0
            ingredient_total_cost += ext_cost
            display = smart_quantity(qty, unit, unit_system)
            ingredient_master.append({
                'id': ing_id,
                'name': ingredient.get('name') or 'Unknown',
                'category': ingredient.get('category'),
                'unit': unit or ingredient.get('unit'),
                'quantity': qty,
                'display_quantity': display['quantity'],
                'display_unit': display['unit'],
                'cost_per_unit': cost_per_unit,
                'ext_cost': ext_cost,
                'vendor': ingredient.get('vendor'),
                'vendor_code': ingredient.get('vendor_code'),
                'g_code': ingredient.get('g_code')
            })

        ingredient_master.sort(key=lambda item: (item['name'] or '').lower())

    batch_recipes = []
    if subrecipe_ids:
        cur.execute("""
            SELECT id, name, category, yield_qty, yield_unit
            FROM recipes
            WHERE id = ANY(%s)
            ORDER BY name
        """, (list(subrecipe_ids),))
        for recipe in cur.fetchall():
            total_cost = get_recipe_total_cost(cur, recipe['id'], unit_system)
            yield_qty = to_float(recipe.get('yield_qty'))
            cost_per_yield = total_cost / yield_qty if yield_qty > 0 else None
            display_yield = smart_quantity(yield_qty, recipe.get('yield_unit'), unit_system) if yield_qty > 0 else None
            batch_recipes.append({
                'id': recipe['id'],
                'name': recipe['name'],
                'category': recipe.get('category'),
                'yield_qty': yield_qty,
                'yield_unit': recipe.get('yield_unit'),
                'display_yield_qty': display_yield['quantity'] if display_yield else None,
                'display_yield_unit': display_yield['unit'] if display_yield else None,
                'total_cost': total_cost,
                'cost_per_yield': cost_per_yield
            })

    return ingredient_master, ingredient_total_cost, batch_recipes

def sanitize_sheet_title(name, existing):
    title = (name or 'Vendor').strip()
    for ch in ['[', ']', ':', '*', '?', '/', '\\']:
        title = title.replace(ch, ' ')
    title = ' '.join(title.split())
    if not title:
        title = 'Vendor'
    title = title[:31]
    base = title
    counter = 1
    while title in existing:
        suffix = f" {counter}"
        max_len = 31 - len(suffix)
        title = f"{base[:max_len]}{suffix}"
        counter += 1
    existing.add(title)
    return title

def make_safe_filename(text):
    cleaned = ''.join(ch for ch in (text or '') if ch.isalnum() or ch in (' ', '-', '_')).strip()
    if not cleaned:
        return 'order_guide'
    return cleaned.replace(' ', '_').lower()

def convert_quantity_between_units(quantity, from_unit, to_unit):
    amount = to_float(quantity)
    from_canonical = normalize_unit(from_unit)
    to_canonical = normalize_unit(to_unit)
    if not from_canonical or not to_canonical:
        return amount if (from_unit or '').strip().lower() == (to_unit or '').strip().lower() else None
    if from_canonical == to_canonical:
        return amount
    from_def = UNIT_DEFS.get(from_canonical)
    to_def = UNIT_DEFS.get(to_canonical)
    if not from_def or not to_def or from_def['type'] != to_def['type']:
        return None
    base_value = amount * from_def['to_base']
    return base_value / to_def['to_base']

def get_banquet_date_window(start_raw, end_raw):
    today = date.today()
    default_start = today
    default_end = default_start + timedelta(days=9)

    start_date = default_start
    end_date = default_end
    try:
        if (start_raw or '').strip():
            start_date = datetime.strptime(start_raw.strip(), '%Y-%m-%d').date()
    except ValueError:
        start_date = default_start
    try:
        if (end_raw or '').strip():
            end_date = datetime.strptime(end_raw.strip(), '%Y-%m-%d').date()
    except ValueError:
        end_date = default_end

    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date

def clean_menu_text(value):
    text = re.sub(r'\s+', ' ', (value or '')).strip()
    text = text.replace('  ', ' ')
    return text

def normalize_match_key(value):
    text = clean_menu_text(value).lower()
    text = re.sub(r'^(rm|rb)\s+', '', text)
    text = re.sub(r'[^a-z0-9 ]+', ' ', text)
    return ' '.join(text.split())

def auto_match_menu_recipe_id(menu_name, recipes):
    needle = normalize_match_key(menu_name)
    if not needle:
        return None
    exact = [row['id'] for row in recipes if normalize_match_key(row.get('name')) == needle]
    if exact:
        return exact[0]
    contains = [row['id'] for row in recipes if needle in normalize_match_key(row.get('name')) or normalize_match_key(row.get('name')) in needle]
    if len(contains) == 1:
        return contains[0]
    return None

def get_plated_recipes_for_venue(cur, venue_id='', include_unassigned=True):
    has_recipe_venues = db_table_exists(cur, 'public.recipe_venues') and db_table_exists(cur, 'public.venues')
    if has_recipe_venues:
        if include_unassigned:
            cur.execute("""
                SELECT r.id,
                       r.name,
                       r.recipe_type,
                       r.category,
                       r.yield_qty,
                       r.yield_unit,
                       r.menu_descriptor,
                       COALESCE(string_agg(DISTINCT v.name, ', ' ORDER BY v.name), '') AS venue_names
                FROM recipes r
                LEFT JOIN recipe_venues rv ON rv.recipe_id = r.id
                LEFT JOIN venues v ON v.id = rv.venue_id AND v.active = TRUE
                WHERE r.recipe_type = 'menu'
                  AND (
                        %s = '' OR
                        EXISTS (
                            SELECT 1
                            FROM recipe_venues rvf
                            WHERE rvf.recipe_id = r.id AND rvf.venue_id = %s
                        ) OR
                        NOT EXISTS (
                            SELECT 1
                            FROM recipe_venues rvn
                            WHERE rvn.recipe_id = r.id
                        )
                  )
                GROUP BY r.id
                ORDER BY r.name
            """, (venue_id, venue_id))
        else:
            cur.execute("""
                SELECT r.id,
                       r.name,
                       r.recipe_type,
                       r.category,
                       r.yield_qty,
                       r.yield_unit,
                       r.menu_descriptor,
                       COALESCE(string_agg(DISTINCT v.name, ', ' ORDER BY v.name), '') AS venue_names
                FROM recipes r
                LEFT JOIN recipe_venues rv ON rv.recipe_id = r.id
                LEFT JOIN venues v ON v.id = rv.venue_id AND v.active = TRUE
                WHERE r.recipe_type = 'menu'
                  AND (
                        %s = '' OR
                        EXISTS (
                            SELECT 1
                            FROM recipe_venues rvf
                            WHERE rvf.recipe_id = r.id AND rvf.venue_id = %s
                        )
                  )
                GROUP BY r.id
                ORDER BY r.name
            """, (venue_id, venue_id))
    else:
        cur.execute("""
            SELECT r.id,
                   r.name,
                   r.recipe_type,
                   r.category,
                   r.yield_qty,
                   r.yield_unit,
                   r.menu_descriptor,
                   '' AS venue_names
            FROM recipes r
            WHERE r.recipe_type = 'menu'
            ORDER BY r.name
        """)
    return cur.fetchall()

def get_component_recipes_for_venue(cur, venue_id=''):
    has_recipe_venues = db_table_exists(cur, 'public.recipe_venues') and db_table_exists(cur, 'public.venues')
    if has_recipe_venues:
        cur.execute("""
            SELECT r.id,
                   r.name,
                   r.recipe_type,
                   r.category,
                   r.yield_qty,
                   r.yield_unit,
                   r.menu_descriptor,
                   COALESCE(string_agg(DISTINCT v.name, ', ' ORDER BY v.name), '') AS venue_names
            FROM recipes r
            LEFT JOIN recipe_venues rv ON rv.recipe_id = r.id
            LEFT JOIN venues v ON v.id = rv.venue_id AND v.active = TRUE
            WHERE (
                %s = '' OR
                EXISTS (
                    SELECT 1
                    FROM recipe_venues rvf
                    WHERE rvf.recipe_id = r.id AND rvf.venue_id = %s
                ) OR
                NOT EXISTS (
                    SELECT 1
                    FROM recipe_venues rvn
                    WHERE rvn.recipe_id = r.id
                )
            )
            GROUP BY r.id
            ORDER BY CASE WHEN r.recipe_type = 'menu' THEN 0 WHEN r.recipe_type = 'batch' THEN 1 ELSE 2 END, r.name
        """, (venue_id, venue_id))
    else:
        cur.execute("""
            SELECT r.id,
                   r.name,
                   r.recipe_type,
                   r.category,
                   r.yield_qty,
                   r.yield_unit,
                   r.menu_descriptor,
                   '' AS venue_names
            FROM recipes r
            ORDER BY CASE WHEN r.recipe_type = 'menu' THEN 0 WHEN r.recipe_type = 'batch' THEN 1 ELSE 2 END, r.name
        """)
    return cur.fetchall()

BANQUET_ACTIVE_STATUSES = ('planning', 'confirmed')
BANQUET_VENUE_ID = 'ven_banquets'
BANQUET_STATUS_CHOICES = [
    ('planning', 'Planning'),
    ('confirmed', 'Confirmed'),
    ('completed', 'Completed'),
    ('cancelled', 'Cancelled')
]

def resolve_banquet_venue(venues):
    for venue in venues:
        if venue.get('id') == BANQUET_VENUE_ID:
            return venue
    if venues:
        return venues[0]
    return {'id': '', 'name': 'Banquets'}

def banquet_tables_ready(cur):
    required = (
        'public.banquet_menu_items',
        'public.banquet_menu_item_recipes',
        'public.banquet_menu_item_ingredients',
        'public.banquet_events',
        'public.banquet_event_menu_items'
    )
    return all(db_table_exists(cur, table_name) for table_name in required)

def list_banquet_menu_items(cur, venue_id=''):
    if not banquet_tables_ready(cur):
        return []
    has_base_yield = db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_qty') and db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_unit')
    base_yield_qty_sql = "mi.base_yield_qty" if has_base_yield else "1::numeric"
    base_yield_unit_sql = "mi.base_yield_unit" if has_base_yield else "'each'"
    if venue_id:
        cur.execute("""
            SELECT mi.id,
                   mi.name,
                   mi.venue_id,
                   COALESCE(v.name, '') AS venue_name,
                   mi.menu_section,
                   mi.menu_descriptor,
                   mi.notes,
                   {base_yield_qty_sql} AS base_yield_qty,
                   {base_yield_unit_sql} AS base_yield_unit,
                   pr.recipe_id AS linked_recipe_id,
                   r.name AS recipe_name,
                   COALESCE(rc.recipe_count, 0) AS recipe_count,
                   COALESCE(extra.extra_count, 0) AS extra_count
            FROM banquet_menu_items mi
            LEFT JOIN venues v ON v.id = mi.venue_id
            LEFT JOIN LATERAL (
                SELECT recipe_id
                FROM banquet_menu_item_recipes
                WHERE menu_item_id = mi.id
                ORDER BY id
                LIMIT 1
            ) pr ON TRUE
            LEFT JOIN recipes r ON r.id = pr.recipe_id
            LEFT JOIN (
                SELECT menu_item_id, COUNT(*) AS recipe_count
                FROM banquet_menu_item_recipes
                GROUP BY menu_item_id
            ) rc ON rc.menu_item_id = mi.id
            LEFT JOIN (
                SELECT menu_item_id, COUNT(*) AS extra_count
                FROM banquet_menu_item_ingredients
                GROUP BY menu_item_id
            ) extra ON extra.menu_item_id = mi.id
            WHERE mi.venue_id = %s OR mi.venue_id IS NULL
            ORDER BY COALESCE(mi.menu_section, 'zzzz'), mi.name
        """.format(base_yield_qty_sql=base_yield_qty_sql, base_yield_unit_sql=base_yield_unit_sql), (venue_id,))
    else:
        cur.execute("""
            SELECT mi.id,
                   mi.name,
                   mi.venue_id,
                   COALESCE(v.name, '') AS venue_name,
                   mi.menu_section,
                   mi.menu_descriptor,
                   mi.notes,
                   {base_yield_qty_sql} AS base_yield_qty,
                   {base_yield_unit_sql} AS base_yield_unit,
                   pr.recipe_id AS linked_recipe_id,
                   r.name AS recipe_name,
                   COALESCE(rc.recipe_count, 0) AS recipe_count,
                   COALESCE(extra.extra_count, 0) AS extra_count
            FROM banquet_menu_items mi
            LEFT JOIN venues v ON v.id = mi.venue_id
            LEFT JOIN LATERAL (
                SELECT recipe_id
                FROM banquet_menu_item_recipes
                WHERE menu_item_id = mi.id
                ORDER BY id
                LIMIT 1
            ) pr ON TRUE
            LEFT JOIN recipes r ON r.id = pr.recipe_id
            LEFT JOIN (
                SELECT menu_item_id, COUNT(*) AS recipe_count
                FROM banquet_menu_item_recipes
                GROUP BY menu_item_id
            ) rc ON rc.menu_item_id = mi.id
            LEFT JOIN (
                SELECT menu_item_id, COUNT(*) AS extra_count
                FROM banquet_menu_item_ingredients
                GROUP BY menu_item_id
            ) extra ON extra.menu_item_id = mi.id
            ORDER BY COALESCE(mi.menu_section, 'zzzz'), mi.name
        """.format(base_yield_qty_sql=base_yield_qty_sql, base_yield_unit_sql=base_yield_unit_sql))
    return cur.fetchall()

def get_banquet_menu_item(cur, menu_item_id):
    if not banquet_tables_ready(cur):
        return None
    has_base_yield = db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_qty') and db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_unit')
    base_yield_qty_sql = "mi.base_yield_qty" if has_base_yield else "1::numeric"
    base_yield_unit_sql = "mi.base_yield_unit" if has_base_yield else "'each'"
    cur.execute("""
        SELECT mi.id,
               mi.name,
               mi.venue_id,
               mi.menu_section,
               mi.menu_descriptor,
               mi.notes,
               {base_yield_qty_sql} AS base_yield_qty,
               {base_yield_unit_sql} AS base_yield_unit,
               pr.recipe_id AS linked_recipe_id,
               pr.quantity AS linked_recipe_qty,
               pr.unit AS linked_recipe_unit
        FROM banquet_menu_items mi
        LEFT JOIN LATERAL (
            SELECT recipe_id, quantity, unit
            FROM banquet_menu_item_recipes
            WHERE menu_item_id = mi.id
            ORDER BY id
            LIMIT 1
        ) pr ON TRUE
        WHERE mi.id = %s
        LIMIT 1
    """.format(base_yield_qty_sql=base_yield_qty_sql, base_yield_unit_sql=base_yield_unit_sql), (menu_item_id,))
    return cur.fetchone()

def get_banquet_menu_item_recipe_components(cur, menu_item_id):
    if not banquet_tables_ready(cur):
        return []
    has_choice_columns = db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_group') and db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_weight_percent')
    choice_group_sql = "mr.choice_group AS choice_group" if has_choice_columns else "NULL::text AS choice_group"
    choice_weight_sql = "mr.choice_weight_percent AS choice_weight_percent" if has_choice_columns else "NULL::numeric AS choice_weight_percent"
    cur.execute(f"""
        SELECT mr.id,
               mr.menu_item_id,
               mr.recipe_id,
               mr.quantity,
               mr.unit,
               {choice_group_sql},
               {choice_weight_sql},
               r.name AS recipe_name,
               r.yield_qty,
               r.yield_unit,
               r.recipe_type
        FROM banquet_menu_item_recipes mr
        JOIN recipes r ON r.id = mr.recipe_id
        WHERE mr.menu_item_id = %s
        ORDER BY mr.id
    """, (menu_item_id,))
    return cur.fetchall()

def collect_banquet_recipe_components(cur, menu_item_ids):
    if not menu_item_ids:
        return {}
    has_choice_columns = db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_group') and db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_weight_percent')
    choice_group_sql = "mr.choice_group AS choice_group" if has_choice_columns else "NULL::text AS choice_group"
    choice_weight_sql = "mr.choice_weight_percent AS choice_weight_percent" if has_choice_columns else "NULL::numeric AS choice_weight_percent"
    cur.execute(f"""
        SELECT mr.menu_item_id,
               mr.recipe_id,
               mr.quantity,
               mr.unit,
               {choice_group_sql},
               {choice_weight_sql},
               r.name AS recipe_name,
               r.yield_qty,
               r.yield_unit,
               r.recipe_type
        FROM banquet_menu_item_recipes mr
        JOIN recipes r ON r.id = mr.recipe_id
        WHERE mr.menu_item_id = ANY(%s)
        ORDER BY mr.menu_item_id, mr.id
    """, (menu_item_ids,))
    grouped = defaultdict(list)
    for row in cur.fetchall():
        grouped[row['menu_item_id']].append(row)
    return grouped

def get_banquet_menu_item_component_summaries(cur, menu_item_ids):
    if not menu_item_ids:
        return {}
    summaries = {menu_item_id: {'recipes': [], 'ingredients': []} for menu_item_id in menu_item_ids}

    has_choice_columns = db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_group') and db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_weight_percent')
    choice_group_sql = "mr.choice_group AS choice_group" if has_choice_columns else "NULL::text AS choice_group"
    choice_weight_sql = "mr.choice_weight_percent AS choice_weight_percent" if has_choice_columns else "NULL::numeric AS choice_weight_percent"
    cur.execute(f"""
        SELECT mr.menu_item_id,
               mr.recipe_id,
               mr.quantity,
               mr.unit,
               {choice_group_sql},
               {choice_weight_sql},
               r.name AS recipe_name
        FROM banquet_menu_item_recipes mr
        JOIN recipes r ON r.id = mr.recipe_id
        WHERE mr.menu_item_id = ANY(%s)
        ORDER BY mr.menu_item_id, mr.id
    """, (menu_item_ids,))
    for row in cur.fetchall():
        summaries.setdefault(row['menu_item_id'], {'recipes': [], 'ingredients': []})
        summaries[row['menu_item_id']]['recipes'].append({
            'id': row.get('recipe_id'),
            'name': row.get('recipe_name'),
            'quantity': to_float(row.get('quantity')),
            'unit': row.get('unit') or '',
            'choice_group': row.get('choice_group'),
            'choice_weight_percent': to_float(row.get('choice_weight_percent'))
        })

    cur.execute("""
        SELECT ai.menu_item_id,
               ai.ingredient_id,
               ai.quantity,
               ai.unit,
               i.name AS ingredient_name
        FROM banquet_menu_item_ingredients ai
        JOIN ingredients i ON i.id = ai.ingredient_id
        WHERE ai.menu_item_id = ANY(%s)
        ORDER BY ai.menu_item_id, ai.id
    """, (menu_item_ids,))
    for row in cur.fetchall():
        summaries.setdefault(row['menu_item_id'], {'recipes': [], 'ingredients': []})
        summaries[row['menu_item_id']]['ingredients'].append({
            'id': row.get('ingredient_id'),
            'name': row.get('ingredient_name'),
            'quantity': to_float(row.get('quantity')),
            'unit': row.get('unit') or ''
        })

    return summaries

def get_banquet_menu_item_ingredients(cur, menu_item_id):
    if not banquet_tables_ready(cur):
        return []
    cur.execute("""
        SELECT ai.id,
               ai.menu_item_id,
               ai.ingredient_id,
               ai.quantity,
               ai.unit,
               i.name AS ingredient_name
        FROM banquet_menu_item_ingredients ai
        JOIN ingredients i ON i.id = ai.ingredient_id
        WHERE ai.menu_item_id = %s
        ORDER BY i.name
    """, (menu_item_id,))
    return cur.fetchall()

def get_banquet_event(cur, event_id):
    if not banquet_tables_ready(cur):
        return None
    cur.execute("""
        SELECT e.id,
               e.name,
               e.event_date,
               e.guest_count AS guests,
               e.guest_count,
               e.venue_id,
               e.building,
               e.room,
               e.service_timing,
               e.dietary_notes,
               e.notes,
               e.status,
               COALESCE(v.name, '') AS venue_name
        FROM banquet_events e
        LEFT JOIN venues v ON v.id = e.venue_id
        WHERE e.id = %s
    """, (event_id,))
    return cur.fetchone()

def get_banquet_event_lines(cur, event_id):
    if not banquet_tables_ready(cur):
        return []
    cur.execute("""
        SELECT emi.id AS line_id,
               emi.event_id,
               emi.menu_item_id,
               COALESCE(NULLIF(TRIM(emi.menu_item_name), ''), mi.name) AS menu_item_name,
               COALESCE(emi.menu_descriptor, mi.menu_descriptor) AS menu_descriptor,
               COALESCE(emi.menu_section, mi.menu_section) AS menu_section,
               COALESCE(emi.recipe_id, primary_recipe.recipe_id) AS recipe_id,
               emi.quantity,
               emi.quantity_unit,
               emi.notes AS line_notes,
               r.name AS recipe_name,
               r.yield_qty,
               r.yield_unit,
               r.menu_descriptor AS recipe_descriptor
        FROM banquet_event_menu_items emi
        JOIN banquet_menu_items mi ON mi.id = emi.menu_item_id
        LEFT JOIN LATERAL (
            SELECT recipe_id
            FROM banquet_menu_item_recipes
            WHERE menu_item_id = mi.id
            ORDER BY id
            LIMIT 1
        ) primary_recipe ON TRUE
        LEFT JOIN recipes r ON r.id = COALESCE(emi.recipe_id, primary_recipe.recipe_id)
        WHERE emi.event_id = %s
        ORDER BY COALESCE(emi.sort_order, 0), emi.id
    """, (event_id,))
    return cur.fetchall()

def resolve_or_create_banquet_menu_item(cur, line, venue_id=''):
    name = clean_menu_text(line.get('menu_item_name'))
    if not name:
        return None

    cur.execute("""
        SELECT id, menu_section, menu_descriptor
        FROM banquet_menu_items
        WHERE LOWER(name) = LOWER(%s)
          AND (venue_id = %s OR venue_id IS NULL)
        ORDER BY CASE WHEN venue_id = %s THEN 0 ELSE 1 END, created_at
        LIMIT 1
    """, (name, venue_id or None, venue_id or None))
    existing = cur.fetchone()
    if existing:
        updates = []
        params = []
        if line.get('menu_section') and not existing.get('menu_section'):
            updates.append('menu_section = %s')
            params.append(line.get('menu_section'))
        if line.get('menu_descriptor') and not existing.get('menu_descriptor'):
            updates.append('menu_descriptor = %s')
            params.append(line.get('menu_descriptor'))
        if updates:
            updates.append('updated_at = CURRENT_TIMESTAMP')
            params.append(existing['id'])
            cur.execute(f"UPDATE banquet_menu_items SET {', '.join(updates)} WHERE id = %s", params)
        return existing['id']

    menu_item_id = generate_id('bmi_')
    cur.execute("""
        INSERT INTO banquet_menu_items (id, name, venue_id, menu_section, menu_descriptor, notes)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (
        menu_item_id,
        name,
        venue_id or None,
        line.get('menu_section') or None,
        line.get('menu_descriptor') or None,
        None
    ))
    return menu_item_id

def upsert_menu_item_recipe_link(cur, menu_item_id, recipe_id):
    if not menu_item_id or not recipe_id:
        return
    cur.execute("""
        SELECT id
        FROM banquet_menu_item_recipes
        WHERE menu_item_id = %s
          AND recipe_id = %s
        LIMIT 1
    """, (menu_item_id, recipe_id))
    if cur.fetchone():
        return
    cur.execute("""
        INSERT INTO banquet_menu_item_recipes (menu_item_id, recipe_id, quantity, unit)
        VALUES (%s, %s, 1, 'serving')
    """, (menu_item_id, recipe_id))

def normalize_return_path(value):
    path = (value or '').strip()
    if not path:
        return None
    if not path.startswith('/') or path.startswith('//'):
        return None
    return path

def make_unique_banquet_menu_item_name(cur, base_name, venue_id):
    candidate = clean_menu_text(base_name) or 'Custom Menu Item'
    suffix = 2
    while True:
        cur.execute("""
            SELECT 1
            FROM banquet_menu_items
            WHERE LOWER(name) = LOWER(%s)
              AND (venue_id = %s OR venue_id IS NULL)
            LIMIT 1
        """, (candidate, venue_id or None))
        if not cur.fetchone():
            return candidate
        candidate = f"{base_name} {suffix}"
        suffix += 1

def clone_banquet_menu_item(cur, source_menu_item_id, venue_id, clone_label='Custom'):
    source_item = get_banquet_menu_item(cur, source_menu_item_id)
    if not source_item:
        return None

    base_name = f"{source_item.get('name') or 'Menu Item'} ({clone_label})"
    clone_name = make_unique_banquet_menu_item_name(cur, base_name, venue_id)
    clone_id = generate_id('bmi_')

    has_base_yield = db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_qty') and db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_unit')
    if has_base_yield:
        cur.execute("""
            INSERT INTO banquet_menu_items (
                id, name, venue_id, menu_section, menu_descriptor, base_yield_qty, base_yield_unit, notes
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            clone_id,
            clone_name,
            venue_id or source_item.get('venue_id'),
            source_item.get('menu_section'),
            source_item.get('menu_descriptor'),
            source_item.get('base_yield_qty') or 1,
            source_item.get('base_yield_unit') or 'each',
            source_item.get('notes')
        ))
    else:
        cur.execute("""
            INSERT INTO banquet_menu_items (
                id, name, venue_id, menu_section, menu_descriptor, notes
            ) VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            clone_id,
            clone_name,
            venue_id or source_item.get('venue_id'),
            source_item.get('menu_section'),
            source_item.get('menu_descriptor'),
            source_item.get('notes')
        ))

    has_choice_columns = db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_group') and db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_weight_percent')
    for row in get_banquet_menu_item_recipe_components(cur, source_menu_item_id):
        if has_choice_columns:
            cur.execute("""
                INSERT INTO banquet_menu_item_recipes (menu_item_id, recipe_id, quantity, unit, choice_group, choice_weight_percent)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                clone_id,
                row.get('recipe_id'),
                row.get('quantity') or 1,
                row.get('unit') or 'each',
                row.get('choice_group'),
                row.get('choice_weight_percent')
            ))
        else:
            cur.execute("""
                INSERT INTO banquet_menu_item_recipes (menu_item_id, recipe_id, quantity, unit)
                VALUES (%s, %s, %s, %s)
            """, (
                clone_id,
                row.get('recipe_id'),
                row.get('quantity') or 1,
                row.get('unit') or 'each'
            ))

    for row in get_banquet_menu_item_ingredients(cur, source_menu_item_id):
        cur.execute("""
            INSERT INTO banquet_menu_item_ingredients (menu_item_id, ingredient_id, quantity, unit)
            VALUES (%s, %s, %s, %s)
        """, (
            clone_id,
            row.get('ingredient_id'),
            row.get('quantity') or 0,
            row.get('unit')
        ))

    return {
        'id': clone_id,
        'name': clone_name,
        'menu_section': source_item.get('menu_section'),
        'menu_descriptor': source_item.get('menu_descriptor')
    }

def parse_event_recipe_lines(request):
    menu_item_ids = request.form.getlist('line_menu_item_id[]')
    recipe_ids = request.form.getlist('line_recipe_id[]')
    names = request.form.getlist('line_menu_item_name[]')
    descriptors = request.form.getlist('line_menu_descriptor[]')
    sections = request.form.getlist('line_menu_section[]')
    quantities = request.form.getlist('line_qty[]')
    quantity_units = request.form.getlist('line_unit[]')
    notes = request.form.getlist('line_notes[]')

    max_len = max(
        len(menu_item_ids), len(recipe_ids), len(names), len(descriptors), len(sections),
        len(quantities), len(quantity_units), len(notes), 0
    )
    rows = []
    errors = []
    for idx in range(max_len):
        menu_item_id = (menu_item_ids[idx] if idx < len(menu_item_ids) else '').strip()
        recipe_id = (recipe_ids[idx] if idx < len(recipe_ids) else '').strip()
        menu_name = clean_menu_text(names[idx] if idx < len(names) else '')
        descriptor = clean_menu_text(descriptors[idx] if idx < len(descriptors) else '')
        section = clean_menu_text(sections[idx] if idx < len(sections) else '')
        qty_raw = (quantities[idx] if idx < len(quantities) else '').strip()
        qty_unit = clean_menu_text(quantity_units[idx] if idx < len(quantity_units) else '')
        note = clean_menu_text(notes[idx] if idx < len(notes) else '')

        if not any([menu_item_id, recipe_id, menu_name, descriptor, section, qty_raw, qty_unit, note]):
            continue
        qty = parse_float_field(qty_raw, 'Menu quantity', errors, required=True, min_value=0.0001)
        if not menu_name and not menu_item_id:
            errors.append('Each banquet line needs a menu item name.')
        rows.append({
            'menu_item_id': menu_item_id or None,
            'recipe_id': recipe_id or None,
            'menu_item_name': menu_name,
            'menu_descriptor': descriptor or None,
            'menu_section': section or None,
            'quantity': qty,
            'quantity_unit': normalize_unit(qty_unit) or qty_unit or None,
            'notes': note or None
        })
    return rows, errors

def parse_menu_item_ingredient_lines(request, valid_ingredient_ids):
    ingredient_ids = request.form.getlist('ingredient_id[]')
    quantities = request.form.getlist('ingredient_qty[]')
    units = request.form.getlist('ingredient_unit[]')

    max_len = max(len(ingredient_ids), len(quantities), len(units), 0)
    rows = []
    errors = []
    for idx in range(max_len):
        ingredient_id = (ingredient_ids[idx] if idx < len(ingredient_ids) else '').strip()
        quantity_raw = (quantities[idx] if idx < len(quantities) else '').strip()
        unit = clean_menu_text(units[idx] if idx < len(units) else '')
        if not any([ingredient_id, quantity_raw, unit]):
            continue
        if ingredient_id not in valid_ingredient_ids:
            errors.append('Additional ingredient row has an invalid ingredient.')
            continue
        quantity = parse_float_field(
            quantity_raw,
            'Additional ingredient quantity',
            errors,
            required=True,
            min_value=0.0001
        )
        if quantity is None:
            continue
        rows.append({
            'ingredient_id': ingredient_id,
            'quantity': quantity,
            'unit': normalize_unit(unit) or unit
        })
    return rows, errors

def parse_menu_item_recipe_lines(request, valid_recipe_ids):
    recipe_ids = request.form.getlist('component_recipe_id[]')
    quantities = request.form.getlist('component_qty[]')
    units = request.form.getlist('component_unit[]')
    choice_groups = request.form.getlist('component_choice_group[]')
    choice_weights = request.form.getlist('component_choice_weight[]')

    max_len = max(len(recipe_ids), len(quantities), len(units), len(choice_groups), len(choice_weights), 0)
    rows = []
    errors = []
    for idx in range(max_len):
        recipe_id = (recipe_ids[idx] if idx < len(recipe_ids) else '').strip()
        quantity_raw = (quantities[idx] if idx < len(quantities) else '').strip()
        unit = clean_menu_text(units[idx] if idx < len(units) else '')
        choice_group = clean_menu_text(choice_groups[idx] if idx < len(choice_groups) else '')
        choice_weight_raw = (choice_weights[idx] if idx < len(choice_weights) else '').strip()

        if not any([recipe_id, quantity_raw, unit, choice_group, choice_weight_raw]):
            continue
        if recipe_id not in valid_recipe_ids:
            errors.append('Recipe component row has an invalid recipe.')
            continue
        quantity = parse_float_field(
            quantity_raw,
            'Recipe component quantity',
            errors,
            required=True,
            min_value=0.0001
        )
        if quantity is None:
            continue

        choice_weight = None
        if choice_weight_raw:
            choice_weight = parse_float_field(
                choice_weight_raw,
                'Choice weight percent',
                errors,
                required=False,
                min_value=0
            )
            if choice_weight is not None and choice_weight > 100:
                errors.append('Choice weight percent cannot exceed 100.')

        rows.append({
            'recipe_id': recipe_id,
            'quantity': quantity,
            'unit': normalize_unit(unit) or unit,
            'choice_group': choice_group or None,
            'choice_weight_percent': choice_weight if choice_group else None
        })

    grouped_choice_rows = defaultdict(list)
    for row in rows:
        group = (row.get('choice_group') or '').strip()
        if group:
            grouped_choice_rows[group].append(row)

    for group_name, group_rows in grouped_choice_rows.items():
        if not group_rows:
            continue
        explicit = [row for row in group_rows if row.get('choice_weight_percent') is not None]
        if not explicit:
            equal = 100.0 / len(group_rows)
            for row in group_rows:
                row['choice_weight_percent'] = equal
            continue

        explicit_total = sum(to_float(row.get('choice_weight_percent')) for row in explicit)
        if explicit_total > 100.0001:
            errors.append(f'Choice group "{group_name}" exceeds 100%.')
            continue

        missing = [row for row in group_rows if row.get('choice_weight_percent') is None]
        if missing:
            remaining = max(0.0, 100.0 - explicit_total)
            each = remaining / len(missing)
            for row in missing:
                row['choice_weight_percent'] = each
        elif 0 < explicit_total < 100:
            scale = 100.0 / explicit_total
            for row in explicit:
                row['choice_weight_percent'] = to_float(row.get('choice_weight_percent')) * scale
    return rows, errors

def ratio_from_line_quantity(line, recipe):
    qty = to_float(line.get('quantity'))
    recipe_yield = to_float(recipe.get('yield_qty'))
    if qty <= 0:
        return 0
    if recipe_yield <= 0:
        return qty

    line_unit = (line.get('quantity_unit') or '').strip()
    yield_unit = (recipe.get('yield_unit') or '').strip()
    qty_in_yield = qty
    if line_unit and yield_unit and line_unit != yield_unit:
        converted = convert_quantity_between_units(qty, line_unit, yield_unit)
        if converted is not None:
            qty_in_yield = converted
    return qty_in_yield / recipe_yield

def normalize_count_unit(unit):
    token = normalize_unit(unit)
    if token in ('', None):
        return ''
    token = token.strip().lower()
    aliases = {
        'ea': 'each',
        'each': 'each',
        'serving': 'each',
        'servings': 'each',
        'plate': 'each',
        'plates': 'each',
        'person': 'each',
        'people': 'each',
        'guest': 'each',
        'guests': 'each',
        'dozen': 'dozen',
        'doz': 'dozen'
    }
    return aliases.get(token, token)

def convert_count_units(quantity, from_unit, to_unit):
    qty = to_float(quantity)
    from_norm = normalize_count_unit(from_unit)
    to_norm = normalize_count_unit(to_unit)
    if qty <= 0:
        return 0
    if not from_norm or not to_norm:
        return None
    if from_norm == to_norm:
        return qty
    if from_norm == 'dozen' and to_norm == 'each':
        return qty * 12
    if from_norm == 'each' and to_norm == 'dozen':
        return qty / 12
    return None

def menu_line_base_multiplier(line_quantity, line_unit, base_yield_qty, base_yield_unit):
    qty = to_float(line_quantity)
    if qty <= 0:
        return 0
    base_qty = to_float(base_yield_qty)
    if base_qty <= 0:
        base_qty = 1

    converted = convert_count_units(qty, line_unit, base_yield_unit)
    base_requested = converted if converted is not None else qty
    return base_requested / base_qty

def collect_banquet_additional_ingredients(cur, menu_item_ids):
    if not menu_item_ids:
        return {}
    cur.execute("""
        SELECT ai.menu_item_id,
               ai.ingredient_id,
               ai.quantity,
               ai.unit,
               i.name AS ingredient_name,
               i.unit AS ingredient_unit,
               i.cost_per_unit,
               i.category,
               i.vendor,
               i.vendor_code,
               i.g_code
        FROM banquet_menu_item_ingredients ai
        JOIN ingredients i ON i.id = ai.ingredient_id
        WHERE ai.menu_item_id = ANY(%s)
    """, (menu_item_ids,))
    grouped = defaultdict(list)
    for row in cur.fetchall():
        grouped[row['menu_item_id']].append(row)
    return grouped

def collect_ingredients_from_components(components, totals, multiplier=1.0):
    for item in components:
        current_multiplier = multiplier
        if item.get('weight_percent') is not None:
            current_multiplier = current_multiplier * (to_float(item.get('weight_percent')) / 100.0)
        if item.get('type') == 'ingredient':
            ing_id = item.get('item_id')
            unit = (item.get('unit') or '').strip()
            qty = to_float(item.get('scaled_quantity')) * current_multiplier
            if ing_id:
                key = (ing_id, unit)
                totals[key] = totals.get(key, 0) + qty
        if item.get('children'):
            collect_ingredients_from_components(item['children'], totals, current_multiplier)

def collect_direct_ingredients_for_prep(components, totals, multiplier=1.0):
    """Collect only direct raw ingredients for a recipe prep card (exclude nested sub-recipe leaves)."""
    for item in components:
        current_multiplier = multiplier
        if item.get('weight_percent') is not None:
            current_multiplier = current_multiplier * (to_float(item.get('weight_percent')) / 100.0)

        item_type = item.get('type')
        if item_type == 'ingredient':
            ing_id = item.get('item_id')
            unit = (item.get('unit') or '').strip()
            qty = to_float(item.get('scaled_quantity')) * current_multiplier
            if ing_id:
                key = (ing_id, unit)
                totals[key] = totals.get(key, 0) + qty
            continue

        # Weighted option groups may include direct ingredient options; recurse into those.
        if item_type == 'option_group' and item.get('children'):
            collect_direct_ingredients_for_prep(item['children'], totals, current_multiplier)

def collect_direct_subrecipes_for_prep(components, totals, multiplier=1.0):
    """Collect only direct sub-recipe pulls for a recipe prep card (exclude deeper nesting)."""
    for item in components:
        current_multiplier = multiplier
        if item.get('weight_percent') is not None:
            current_multiplier = current_multiplier * (to_float(item.get('weight_percent')) / 100.0)

        item_type = item.get('type')
        if item_type == 'recipe' and item.get('sub_recipe'):
            sub_recipe = item.get('sub_recipe') or {}
            key = (sub_recipe.get('id'), item.get('unit') or sub_recipe.get('yield_unit') or '')
            totals[key] = totals.get(key, 0) + (to_float(item.get('scaled_quantity')) * current_multiplier)
            continue

        # Weighted option groups may include recipe options; recurse into those.
        if item_type == 'option_group' and item.get('children'):
            collect_direct_subrecipes_for_prep(item['children'], totals, current_multiplier)

def collect_batch_recipe_usage_from_components(
    components,
    usage_map,
    event_usage_map=None,
    menu_item_usage_map=None,
    event_id=None,
    menu_item_name=None,
    multiplier=1.0
):
    for item in components:
        current_multiplier = multiplier
        if item.get('weight_percent') is not None:
            current_multiplier = current_multiplier * (to_float(item.get('weight_percent')) / 100.0)

        if item.get('type') == 'recipe' and item.get('sub_recipe'):
            sub_recipe = item.get('sub_recipe') or {}
            key = (sub_recipe.get('id'), item.get('unit') or sub_recipe.get('yield_unit') or '')
            usage_map[key] = usage_map.get(key, 0) + (to_float(item.get('scaled_quantity')) * current_multiplier)
            if event_usage_map is not None and event_id:
                event_usage_map[sub_recipe.get('id')].add(event_id)
            if menu_item_usage_map is not None and menu_item_name:
                menu_item_usage_map[sub_recipe.get('id')].add(menu_item_name)

        if item.get('children'):
            collect_batch_recipe_usage_from_components(
                item['children'],
                usage_map,
                event_usage_map=event_usage_map,
                menu_item_usage_map=menu_item_usage_map,
                event_id=event_id,
                menu_item_name=menu_item_name,
                multiplier=current_multiplier
            )

def collect_subrecipes_from_components(components, subrecipes):
    for item in components:
        if item.get('type') == 'recipe' and item.get('item_id'):
            subrecipes.add(item['item_id'])
        if item.get('children'):
            collect_subrecipes_from_components(item['children'], subrecipes)

def fetch_banquet_event_lines(cur, start_date, end_date, venue_id=''):
    if not banquet_tables_ready(cur):
        return []
    has_base_yield = db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_qty') and db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_unit')
    base_yield_qty_sql = "mi.base_yield_qty" if has_base_yield else "1::numeric"
    base_yield_unit_sql = "mi.base_yield_unit" if has_base_yield else "'each'"
    params = [start_date, end_date, list(BANQUET_ACTIVE_STATUSES)]
    venue_filter_sql = ""
    if venue_id:
        venue_filter_sql = " AND e.venue_id = %s"
        params.append(venue_id)

    cur.execute(f"""
        SELECT e.id AS event_id,
               e.name AS event_name,
               e.event_date,
               e.guest_count AS guests,
               e.venue_id,
               e.status,
               COALESCE(v.name, e.building, 'Banquet') AS venue_name,
               e.building,
               e.room,
               e.service_timing,
               e.dietary_notes,
               e.notes AS event_notes,
               emi.id AS line_id,
               emi.menu_item_id,
               COALESCE(NULLIF(TRIM(emi.menu_item_name), ''), mi.name) AS menu_item_name,
               COALESCE(emi.menu_descriptor, mi.menu_descriptor) AS menu_descriptor,
               COALESCE(emi.menu_section, mi.menu_section) AS menu_section,
               {base_yield_qty_sql} AS base_yield_qty,
               {base_yield_unit_sql} AS base_yield_unit,
               COALESCE(emi.recipe_id, primary_recipe.recipe_id) AS recipe_id,
               emi.quantity,
               emi.quantity_unit,
               emi.notes AS line_notes,
               r.name AS recipe_name,
               r.category AS recipe_category,
               r.yield_qty,
               r.yield_unit,
               r.menu_descriptor AS recipe_descriptor
        FROM banquet_events e
        LEFT JOIN venues v ON v.id = e.venue_id
        LEFT JOIN banquet_event_menu_items emi ON emi.event_id = e.id
        LEFT JOIN banquet_menu_items mi ON mi.id = emi.menu_item_id
        LEFT JOIN LATERAL (
            SELECT recipe_id
            FROM banquet_menu_item_recipes
            WHERE menu_item_id = mi.id
            ORDER BY id
            LIMIT 1
        ) primary_recipe ON TRUE
        LEFT JOIN recipes r ON r.id = COALESCE(emi.recipe_id, primary_recipe.recipe_id)
        WHERE e.event_date BETWEEN %s AND %s
          AND e.status = ANY(%s)
          {venue_filter_sql}
        ORDER BY e.event_date, e.name, COALESCE(emi.sort_order, 0), emi.id
    """.format(base_yield_qty_sql=base_yield_qty_sql, base_yield_unit_sql=base_yield_unit_sql, venue_filter_sql=venue_filter_sql), params)
    return cur.fetchall()

def build_banquet_datasets(cur, start_date, end_date, venue_id='', unit_system='imperial'):
    rows = fetch_banquet_event_lines(cur, start_date, end_date, venue_id)
    events_map = {}
    ingredient_totals = {}
    ingredient_usage = defaultdict(set)
    batch_usage_qty = {}
    batch_usage_events = defaultdict(set)
    batch_usage_menu_items = defaultdict(set)
    event_daily_ingredients = defaultdict(lambda: defaultdict(float))

    menu_item_ids = list({row.get('menu_item_id') for row in rows if row.get('menu_item_id')})
    additional_ingredients = collect_banquet_additional_ingredients(cur, menu_item_ids)
    menu_item_recipe_components = collect_banquet_recipe_components(cur, menu_item_ids)

    for row in rows:
        event_id = row['event_id']
        if event_id not in events_map:
            events_map[event_id] = {
                'id': event_id,
                'name': row.get('event_name'),
                'event_date': row.get('event_date'),
                'guests': row.get('guests'),
                'venue_name': row.get('venue_name'),
                'building': row.get('building'),
                'room': row.get('room'),
                'service_timing': row.get('service_timing'),
                'dietary_notes': row.get('dietary_notes'),
                'notes': row.get('event_notes'),
                'status': row.get('status') or 'planning',
                'lines': [],
                'line_count': 0
            }

        if not row.get('line_id'):
            continue

        recipe = None
        if row.get('recipe_id'):
            recipe = {
                'id': row.get('recipe_id'),
                'name': row.get('recipe_name') or row.get('menu_item_name'),
                'yield_qty': row.get('yield_qty'),
                'yield_unit': row.get('yield_unit')
            }

        line_recipes = []
        for component in menu_item_recipe_components.get(row.get('menu_item_id'), []):
            line_recipes.append({
                'id': component.get('recipe_id'),
                'name': component.get('recipe_name'),
                'yield_qty': component.get('yield_qty'),
                'yield_unit': component.get('yield_unit'),
                'quantity': to_float(component.get('quantity')),
                'unit': component.get('unit') or component.get('yield_unit'),
                'recipe_type': component.get('recipe_type'),
                'choice_group': component.get('choice_group'),
                'choice_weight_percent': to_float(component.get('choice_weight_percent'))
            })
        if not line_recipes and recipe:
            line_recipes = [dict(recipe)]

        choice_group_rows = defaultdict(list)
        for line_recipe in line_recipes:
            group_name = clean_menu_text(line_recipe.get('choice_group') or '')
            if group_name:
                choice_group_rows[group_name].append(line_recipe)

        choice_multipliers = {}
        for group_name, group_rows in choice_group_rows.items():
            total_weight = sum(max(0.0, to_float(item.get('choice_weight_percent'))) for item in group_rows)
            if total_weight <= 0:
                equal_share = 1.0 / len(group_rows) if group_rows else 0
                for item in group_rows:
                    choice_multipliers[id(item)] = equal_share
            else:
                for item in group_rows:
                    choice_multipliers[id(item)] = max(0.0, to_float(item.get('choice_weight_percent'))) / total_weight

        line = {
            'id': row.get('line_id'),
            'menu_item_name': row.get('menu_item_name') or row.get('recipe_name') or 'Menu Item',
            'menu_descriptor': row.get('menu_descriptor') or row.get('recipe_descriptor'),
            'menu_section': row.get('menu_section') or 'Uncategorized',
            'quantity': to_float(row.get('quantity')),
            'quantity_unit': row.get('quantity_unit') or row.get('yield_unit') or 'each',
            'base_yield_qty': to_float(row.get('base_yield_qty')) or 1,
            'base_yield_unit': row.get('base_yield_unit') or 'each',
            'notes': row.get('line_notes'),
            'recipe': line_recipes[0] if line_recipes else recipe,
            'recipes': line_recipes
        }
        line['ratio'] = None
        line['estimated_cost_total'] = None
        line['estimated_cost_per_unit'] = None
        line['components'] = []
        line['additional_ingredients'] = []
        line['additional_ingredient_cost'] = 0

        line_cost_total = 0
        line_multiplier = menu_line_base_multiplier(
            line.get('quantity'),
            line.get('quantity_unit'),
            line.get('base_yield_qty'),
            line.get('base_yield_unit')
        )
        line['ratio'] = line_multiplier

        for line_recipe in line_recipes:
            ratio_per_menu_unit = ratio_from_line_quantity(
                {
                    'quantity': line_recipe.get('quantity', line.get('base_yield_qty')),
                    'quantity_unit': line_recipe.get('unit', line_recipe.get('yield_unit'))
                },
                line_recipe
            )
            group_multiplier = choice_multipliers.get(id(line_recipe), 1.0)
            total_ratio = ratio_per_menu_unit * line_multiplier * group_multiplier
            line_cost_total += get_recipe_total_cost(cur, line_recipe['id'], unit_system, apply_q_factor=True) * total_ratio

            recipe_yield_qty = to_float(line_recipe.get('yield_qty'))
            required_output_qty = (recipe_yield_qty * total_ratio) if recipe_yield_qty > 0 else total_ratio
            required_output_unit = line_recipe.get('yield_unit') or line_recipe.get('unit') or ''
            root_key = (line_recipe.get('id'), required_output_unit)
            batch_usage_qty[root_key] = batch_usage_qty.get(root_key, 0) + required_output_qty
            if event_id:
                batch_usage_events[line_recipe.get('id')].add(event_id)
            if line.get('menu_item_name'):
                batch_usage_menu_items[line_recipe.get('id')].add(line['menu_item_name'])

            components, _, _ = build_component_tree(cur, line_recipe['id'], total_ratio, 0, set(), unit_system, apply_q_factor=False)
            if components:
                line['components'].extend(components)
            collect_ingredients_from_components(components, ingredient_totals)
            collect_ingredients_from_components(components, event_daily_ingredients[event_id])
            collect_batch_recipe_usage_from_components(
                components,
                batch_usage_qty,
                event_usage_map=batch_usage_events,
                menu_item_usage_map=batch_usage_menu_items,
                event_id=event_id,
                menu_item_name=line['menu_item_name']
            )

            line_label = line['menu_item_name']
            collect_ingredient_usage_from_components(components, line_label, ingredient_usage)

        for extra in additional_ingredients.get(row.get('menu_item_id'), []):
            ingredient_id = extra.get('ingredient_id')
            if not ingredient_id:
                continue
            per_unit_qty = to_float(extra.get('quantity'))
            extra_qty = per_unit_qty * line_multiplier
            unit = (extra.get('unit') or extra.get('ingredient_unit') or '').strip()
            key = (ingredient_id, unit)
            ingredient_totals[key] = ingredient_totals.get(key, 0) + extra_qty
            event_daily_ingredients[event_id][key] += extra_qty
            ingredient_usage[ingredient_id].add(line.get('menu_item_name') or 'Menu Item')
            converted_cost = convert_cost_per_unit(
                extra.get('cost_per_unit'),
                extra.get('ingredient_unit'),
                unit
            )
            ext_cost = extra_qty * converted_cost if converted_cost else 0
            line_cost_total += ext_cost
            line['additional_ingredient_cost'] += ext_cost
            line['additional_ingredients'].append({
                'ingredient_id': ingredient_id,
                'name': extra.get('ingredient_name'),
                'quantity': extra_qty,
                'unit': unit,
                'ext_cost': ext_cost
            })

        line['estimated_cost_total'] = line_cost_total
        line['estimated_cost_per_unit'] = line_cost_total / line['quantity'] if line['quantity'] > 0 else None

        events_map[event_id]['lines'].append(line)
        events_map[event_id]['line_count'] += 1

    events = sorted(events_map.values(), key=lambda item: ((item.get('event_date') or date.today()), (item.get('name') or '').lower()))

    ingredient_master = []
    ingredient_total_cost = 0
    if ingredient_totals:
        ingredient_ids = list({ing_id for ing_id, _ in ingredient_totals.keys() if ing_id})
        ingredient_map = {}
        if ingredient_ids:
            cur.execute("""
                SELECT id, name, unit, category, cost_per_unit, vendor, vendor_code, g_code
                FROM ingredients
                WHERE id = ANY(%s)
            """, (ingredient_ids,))
            ingredient_map = {row['id']: row for row in cur.fetchall()}

        for (ing_id, unit), qty in ingredient_totals.items():
            ingredient = ingredient_map.get(ing_id, {})
            cost_per_unit = convert_cost_per_unit(
                ingredient.get('cost_per_unit'),
                ingredient.get('unit'),
                unit
            )
            ext_cost = qty * cost_per_unit if cost_per_unit else 0
            ingredient_total_cost += ext_cost
            display = smart_quantity(qty, unit, unit_system)
            ingredient_master.append({
                'id': ing_id,
                'name': ingredient.get('name') or 'Unknown',
                'category': ingredient.get('category') or 'Uncategorized',
                'vendor': ingredient.get('vendor'),
                'vendor_code': ingredient.get('vendor_code'),
                'g_code': ingredient.get('g_code'),
                'quantity': qty,
                'unit': unit,
                'display_quantity': display['quantity'],
                'display_unit': display['unit'],
                'cost_per_unit': cost_per_unit,
                'ext_cost': ext_cost,
                'used_in': sorted(ingredient_usage.get(ing_id, set()))
            })
    ingredient_master.sort(key=lambda item: ((item.get('vendor') or '').lower(), (item.get('category') or '').lower(), (item.get('name') or '').lower()))

    weekly_prep = []
    if batch_usage_qty:
        batch_ids = list({recipe_id for recipe_id, _ in batch_usage_qty.keys() if recipe_id})
        recipe_map = {}
        if batch_ids:
            cur.execute("""
                SELECT id, name, category, yield_qty, yield_unit, instructions, recipe_type
                FROM recipes
                WHERE id = ANY(%s)
            """, (batch_ids,))
            recipe_map = {row['id']: row for row in cur.fetchall()}

        for (recipe_id, qty_unit), required_qty in batch_usage_qty.items():
            recipe = recipe_map.get(recipe_id)
            if not recipe:
                continue
            yield_qty = to_float(recipe.get('yield_qty'))
            yield_unit = recipe.get('yield_unit')
            required_qty_in_yield = required_qty
            if qty_unit and yield_unit and qty_unit != yield_unit:
                converted = convert_quantity_between_units(required_qty, qty_unit, yield_unit)
                if converted is not None:
                    required_qty_in_yield = converted

            required_batches = (required_qty_in_yield / yield_qty) if yield_qty > 0 else required_qty_in_yield
            components, total_cost, _ = build_component_tree(cur, recipe_id, required_batches, 0, set(), unit_system, apply_q_factor=False)
            ingredient_totals_for_batch = {}
            collect_direct_ingredients_for_prep(components, ingredient_totals_for_batch)
            subrecipe_totals_for_batch = {}
            collect_direct_subrecipes_for_prep(components, subrecipe_totals_for_batch)
            ingredient_rows = []
            for (ing_id, unit), qty in ingredient_totals_for_batch.items():
                display = smart_quantity(qty, unit, unit_system)
                ingredient_rows.append({
                    'ingredient_id': ing_id,
                    'name': next((item['name'] for item in ingredient_master if item['id'] == ing_id), 'Unknown'),
                    'quantity': qty,
                    'unit': unit,
                    'display_quantity': display['quantity'],
                    'display_unit': display['unit']
                })
            ingredient_rows.sort(key=lambda item: (item.get('name') or '').lower())

            subrecipe_rows = []
            for (sub_recipe_id, sub_unit), sub_qty in subrecipe_totals_for_batch.items():
                sub_recipe = recipe_map.get(sub_recipe_id)
                if not sub_recipe:
                    sub_recipe = get_recipe_by_id(cur, sub_recipe_id)
                if not sub_recipe:
                    continue
                sub_yield_qty = to_float(sub_recipe.get('yield_qty'))
                sub_yield_unit = sub_recipe.get('yield_unit') or sub_unit
                sub_required_qty_in_yield = sub_qty
                if sub_unit and sub_yield_unit and sub_unit != sub_yield_unit:
                    converted = convert_quantity_between_units(sub_qty, sub_unit, sub_yield_unit)
                    if converted is not None:
                        sub_required_qty_in_yield = converted
                sub_required_batches = (sub_required_qty_in_yield / sub_yield_qty) if sub_yield_qty > 0 else sub_required_qty_in_yield
                display = smart_quantity(sub_qty, sub_unit, unit_system)
                subrecipe_rows.append({
                    'recipe_id': sub_recipe_id,
                    'recipe_name': sub_recipe.get('name') or 'Sub-recipe',
                    'required_qty': sub_qty,
                    'required_unit': sub_unit,
                    'display_required': display,
                    'required_batches': sub_required_batches,
                    'yield_qty': sub_yield_qty,
                    'yield_unit': sub_yield_unit
                })
            subrecipe_rows.sort(key=lambda item: (item.get('recipe_name') or '').lower())

            weekly_prep.append({
                'recipe_id': recipe_id,
                'recipe_name': recipe.get('name'),
                'category': recipe.get('category'),
                'recipe_type': normalize_recipe_type(recipe.get('recipe_type')),
                'required_batches': required_batches,
                'required_qty': required_qty_in_yield,
                'required_unit': yield_unit or qty_unit,
                'display_required': smart_quantity(required_qty_in_yield, yield_unit or qty_unit, unit_system),
                'yield_qty': yield_qty,
                'yield_unit': yield_unit,
                'ingredient_rows': ingredient_rows,
                'subrecipe_rows': subrecipe_rows,
                'instructions': recipe.get('instructions'),
                'used_in_events': sorted(batch_usage_events.get(recipe_id, set())),
                'used_in_menu_items': sorted(batch_usage_menu_items.get(recipe_id, set())),
                'estimated_cost': total_cost
            })
    weekly_prep.sort(key=lambda item: (item.get('recipe_name') or '').lower())

    daily_groups = []
    by_day = defaultdict(list)
    for event in events:
        by_day[event.get('event_date')].append(event)
    for day in sorted(by_day.keys()):
        day_events = by_day[day]
        for event in day_events:
            ingredient_rows = []
            for (ing_id, unit), qty in event_daily_ingredients.get(event['id'], {}).items():
                display = smart_quantity(qty, unit, unit_system)
                ingredient_rows.append({
                    'ingredient_id': ing_id,
                    'name': next((item['name'] for item in ingredient_master if item['id'] == ing_id), 'Unknown'),
                    'quantity': qty,
                    'unit': unit,
                    'display_quantity': display['quantity'],
                    'display_unit': display['unit']
                })
            ingredient_rows.sort(key=lambda item: (item.get('name') or '').lower())
            event['daily_ingredients'] = ingredient_rows
        daily_groups.append({'date': day, 'events': day_events})

    total_menu_lines = sum(event.get('line_count', 0) for event in events)
    total_estimated_cost = sum(sum(line.get('estimated_cost_total', 0) for line in event.get('lines', [])) for event in events)

    return {
        'events': events,
        'daily_groups': daily_groups,
        'shopping_ingredients': ingredient_master,
        'shopping_total_cost': ingredient_total_cost,
        'weekly_prep': weekly_prep,
        'event_count': len(events),
        'menu_line_count': total_menu_lines,
        'total_estimated_cost': total_estimated_cost
    }

def parse_catering_pdf_to_template_items(file_stream):
    reader = PdfReader(file_stream)
    parsed = []
    current_major_section = 'Imported Menu'
    current_subsection = 'Items'
    current_package = ''
    current_item = None

    ignored_prefixes = (
        'Pricing is subject',
        'Receptions Serving',
        'Buffets Serving',
        'Listed prices are per person',
        'All are priced',
        'Price listed',
        'Place cards',
        'ALL DINNERS INCLUDE',
        'Consumption bar'
    )
    ignored_exact = {
        'Catering Menu',
        'THE START OF SOMETHING DELICIOUS'
    }

    def flush_current():
        nonlocal current_item
        if not current_item:
            return
        current_item['item_name'] = clean_menu_text(current_item.get('item_name'))
        current_item['menu_descriptor'] = clean_menu_text(current_item.get('menu_descriptor'))
        if current_item['item_name']:
            parsed.append(current_item)
        current_item = None

    for page in reader.pages:
        text = page.extract_text() or ''
        for raw_line in text.splitlines():
            line = clean_menu_text(raw_line)
            if not line:
                continue
            if line in ignored_exact or line.startswith(ignored_prefixes):
                continue

            upper = line.upper()
            if upper in (
                'BREAKFAST ENHANCEMENTS', 'REFRESH AND RENEW: BREAKOUT PACKAGES',
                'SNACK ATTACK', 'MORNING BEVERAGES', 'STATIONS - DAYTIME',
                'RECEPTIONS DISPLAYS', 'RECEPTIONS STATIONS', 'SMALL BITES',
                "HORS D'OEUVRES", 'BUFFET PACKAGES', 'CUSTOMIZE YOUR DINNER BUFFET',
                'PLATED 3-COURSE DINNER', 'LATE NIGHT SNACKS',
                'FOXTOWN PREMIUM BEVERAGE PACKAGE', 'FOXTOWN STANDARD BEVERAGE PACKAGE',
                'BEER, WINE & SODA PACKAGES', 'PACKAGE ENHANCEMENTS'
            ):
                flush_current()
                current_major_section = line.title()
                current_subsection = 'Items'
                current_package = ''
                continue

            if upper in BANQUET_PACKAGE_SUBSECTIONS:
                flush_current()
                current_subsection = line.title()
                continue

            if upper.startswith('CHOICE OF ') or upper.startswith('CHOOSE ') or upper.endswith(' OPTIONS'):
                continue

            package_match = re.match(r'^([A-Z0-9 &\'’\-\(\)\.]+?)\s+\$\s*([0-9]+(?:\.[0-9]+)?|Market Price)$', line)
            if package_match and len(package_match.group(1).split()) >= 2:
                flush_current()
                current_package = clean_menu_text(package_match.group(1).title())
                current_subsection = 'Items'
                continue

            item_with_price = re.match(r'^(.*?)\s+\$\s*([0-9]+(?:\.[0-9]+)?|Market Price)$', line)
            if item_with_price:
                flush_current()
                section_parts = [current_major_section]
                if current_package:
                    section_parts.append(current_package)
                if current_subsection and current_subsection != 'Items':
                    section_parts.append(current_subsection)
                section = ' · '.join(section_parts)
                current_item = {
                    'menu_section': section,
                    'item_name': clean_menu_text(item_with_price.group(1)),
                    'menu_descriptor': '',
                    'price_text': item_with_price.group(2)
                }
                continue

            if '|' in line:
                flush_current()
                left, right = line.split('|', 1)
                section_parts = [current_major_section]
                if current_package:
                    section_parts.append(current_package)
                if current_subsection and current_subsection != 'Items':
                    section_parts.append(current_subsection)
                section = ' · '.join(section_parts)
                current_item = {
                    'menu_section': section,
                    'item_name': clean_menu_text(left),
                    'menu_descriptor': clean_menu_text(right),
                    'price_text': ''
                }
                continue

            is_heading_like = upper == line and len(line.split()) <= 8 and any(ch.isalpha() for ch in line)
            if is_heading_like and len(line) < 50:
                flush_current()
                current_subsection = line.title()
                continue

            if not current_item:
                section_parts = [current_major_section]
                if current_package:
                    section_parts.append(current_package)
                if current_subsection and current_subsection != 'Items':
                    section_parts.append(current_subsection)
                section = ' · '.join(section_parts)
                current_item = {
                    'menu_section': section,
                    'item_name': line,
                    'menu_descriptor': '',
                    'price_text': ''
                }
            else:
                if current_item.get('menu_descriptor'):
                    current_item['menu_descriptor'] = f"{current_item['menu_descriptor']} {line}".strip()
                else:
                    current_item['menu_descriptor'] = line

    flush_current()

    deduped = []
    seen = set()
    for item in parsed:
        key = (item.get('menu_section', '').lower(), item.get('item_name', '').lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped

def find_or_create_ingredient(cur, name, unit_value):
    cur.execute("SELECT id FROM ingredients WHERE LOWER(name) = LOWER(%s) LIMIT 1", (name,))
    existing = cur.fetchone()
    if existing:
        return existing['id'], False

    ingredient_id = generate_id('ing_')
    cur.execute("""
        INSERT INTO ingredients (id, name, unit)
        VALUES (%s, %s, %s)
    """, (
        ingredient_id,
        name,
        unit_value or None
    ))
    return ingredient_id, True

# Simple user class for authentication
class User(UserMixin):
    def __init__(self, id, username):
        self.id = id
        self.username = username

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    config = get_admin_config()
    username = config['username'] if config else 'admin'
    return User(user_id, username)  # Simplified for now

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        config = get_admin_config()
        if not config:
            flash('Admin credentials are not configured yet.', 'error')
            return render_template('login.html')

        username = request.form.get('username')
        password = request.form.get('password')
        
        is_user_match = username == config['username']
        is_pass_match = check_password_hash(config['password_hash'], password or '')

        if is_user_match and is_pass_match:
            user = User('1', username)
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT COUNT(*) AS count FROM recipes")
    recipe_count = cur.fetchone()['count']

    cur.execute("SELECT COUNT(*) AS count FROM ingredients")
    ingredient_count = cur.fetchone()['count']

    stale_cutoff = datetime.utcnow() - timedelta(days=PRICE_REFRESH_DAYS)
    cur.execute("""
        SELECT COUNT(*) AS count
        FROM ingredients
        WHERE price_updated_at IS NULL OR price_updated_at < %s
    """, (stale_cutoff,))
    stale_price_count = cur.fetchone()['count']

    recent_cutoff = datetime.utcnow() - timedelta(days=30)
    cur.execute("""
        SELECT COUNT(*) AS count
        FROM menu_rollouts
        WHERE is_one_off = FALSE
          AND created_at >= %s
    """, (recent_cutoff,))
    recent_rollout_count = cur.fetchone()['count']

    cur.close()

    return render_template(
        'dashboard.html',
        recipe_count=recipe_count,
        ingredient_count=ingredient_count,
        stale_price_count=stale_price_count,
        recent_rollout_count=recent_rollout_count,
        price_refresh_days=PRICE_REFRESH_DAYS
    )

@app.route('/banquet-planner')
@login_required
def banquet_planner():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    today = date.today()
    ten_day_end = today + timedelta(days=9)

    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    selected_venue = banquet_venue.get('id') or ''
    banquet_venue_name = banquet_venue.get('name') or 'Banquets'

    start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return render_template(
            'banquet_planner.html',
            venues=[banquet_venue] if selected_venue else venues,
            banquet_venue_name=banquet_venue_name,
            selected_venue=selected_venue,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            upcoming_events=[],
            event_count=0,
            menu_line_count=0,
            shopping_count=0,
            weekly_prep_count=0,
            menu_item_count=0,
            banquet_metrics={
                'total_events': 0,
                'completed_events': 0,
                'active_next_10_days': 0,
                'avg_guests_90d': 0
            },
            past_events=[]
        )

    datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())

    cur.execute("""
        SELECT e.id,
               e.name,
               e.event_date,
               e.guest_count AS guests,
               e.status,
               COALESCE(v.name, e.building, 'Banquet') AS venue_name,
               COUNT(emi.id) AS recipe_lines
        FROM banquet_events e
        LEFT JOIN venues v ON v.id = e.venue_id
        LEFT JOIN banquet_event_menu_items emi ON emi.event_id = e.id
        WHERE e.event_date BETWEEN %s AND %s
          AND (%s = '' OR e.venue_id = %s)
        GROUP BY e.id, v.name
        ORDER BY e.event_date, e.name
        LIMIT 60
    """, (start_date, end_date, selected_venue, selected_venue))
    upcoming_events = cur.fetchall()

    if selected_venue:
        cur.execute("""
            SELECT COUNT(*) AS count
            FROM banquet_menu_items
            WHERE venue_id = %s OR venue_id IS NULL
        """, (selected_venue,))
    else:
        cur.execute("SELECT COUNT(*) AS count FROM banquet_menu_items")
    menu_item_count = (cur.fetchone() or {}).get('count', 0)

    cur.execute("""
        SELECT
            COUNT(*) AS total_events,
            COUNT(*) FILTER (WHERE status = 'completed') AS completed_events,
            COUNT(*) FILTER (WHERE event_date BETWEEN %s AND %s AND status = ANY(%s)) AS active_next_10_days
        FROM banquet_events e
        WHERE (%s = '' OR e.venue_id = %s)
    """, (today, ten_day_end, list(BANQUET_ACTIVE_STATUSES), selected_venue, selected_venue))
    metrics_row = cur.fetchone() or {}

    cur.execute("""
        SELECT ROUND(AVG(guest_count)::numeric, 1) AS avg_guests_90d
        FROM banquet_events e
        WHERE guest_count IS NOT NULL
          AND guest_count > 0
          AND e.event_date >= %s
          AND (%s = '' OR e.venue_id = %s)
    """, (today - timedelta(days=90), selected_venue, selected_venue))
    avg_row = cur.fetchone() or {}
    avg_guests_90d = avg_row.get('avg_guests_90d')

    banquet_metrics = {
        'total_events': int(metrics_row.get('total_events') or 0),
        'completed_events': int(metrics_row.get('completed_events') or 0),
        'active_next_10_days': int(metrics_row.get('active_next_10_days') or 0),
        'avg_guests_90d': float(avg_guests_90d) if avg_guests_90d is not None else 0
    }

    cur.execute("""
        SELECT e.id,
               e.name,
               e.event_date,
               e.guest_count AS guests,
               e.status,
               COALESCE(v.name, e.building, 'Banquet') AS venue_name,
               COUNT(emi.id) AS recipe_lines
        FROM banquet_events e
        LEFT JOIN venues v ON v.id = e.venue_id
        LEFT JOIN banquet_event_menu_items emi ON emi.event_id = e.id
        WHERE e.event_date < %s
          AND (%s = '' OR e.venue_id = %s)
        GROUP BY e.id, v.name
        ORDER BY e.event_date DESC, e.name
        LIMIT 20
    """, (today, selected_venue, selected_venue))
    past_events = cur.fetchall()

    cur.close()

    return render_template(
        'banquet_planner.html',
        venues=[banquet_venue] if selected_venue else venues,
        banquet_venue_name=banquet_venue_name,
        selected_venue=selected_venue,
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        upcoming_events=upcoming_events,
        event_count=datasets['event_count'],
        menu_line_count=datasets['menu_line_count'],
        shopping_count=len(datasets['shopping_ingredients']),
        weekly_prep_count=len(datasets['weekly_prep']),
        menu_item_count=menu_item_count,
        banquet_metrics=banquet_metrics,
        past_events=past_events
    )

@app.route('/banquet-planner/events/new', methods=['GET', 'POST'])
@login_required
def banquet_event_new():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))

    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    default_venue_id = banquet_venue.get('id') or ''
    banquet_venue_name = banquet_venue.get('name') or 'Banquets'

    event = {
        'name': '',
        'event_date': '',
        'guests': '',
        'venue_id': default_venue_id,
        'building': '',
        'room': '',
        'service_timing': '',
        'dietary_notes': '',
        'notes': '',
        'status': 'planning'
    }

    if request.method == 'POST':
        errors = []
        event_name = clean_menu_text(request.form.get('name'))
        event_date_raw = (request.form.get('event_date') or '').strip()
        venue_id = default_venue_id

        status = (request.form.get('status') or 'planning').strip().lower()
        if status not in {choice[0] for choice in BANQUET_STATUS_CHOICES}:
            status = 'planning'

        guests = parse_float_field(request.form.get('guests'), 'Guests', errors, required=False, min_value=1)
        event_date = None
        try:
            event_date = datetime.strptime(event_date_raw, '%Y-%m-%d').date()
        except ValueError:
            errors.append('Event date is required.')

        lines, line_errors = parse_event_recipe_lines(request)
        errors.extend(line_errors)
        if not lines:
            errors.append('Add at least one menu line for the event.')
        if not event_name:
            errors.append('Event name is required.')

        event = {
            'name': event_name,
            'event_date': event_date_raw,
            'guests': request.form.get('guests') or '',
            'venue_id': venue_id,
            'building': clean_menu_text(request.form.get('building')),
            'room': clean_menu_text(request.form.get('room')),
            'service_timing': clean_menu_text(request.form.get('service_timing')),
            'dietary_notes': clean_menu_text(request.form.get('dietary_notes')),
            'notes': clean_menu_text(request.form.get('notes')),
            'status': status
        }

        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
        else:
            event_id = generate_id('bev_')
            try:
                cur.execute("""
                    INSERT INTO banquet_events (
                        id, name, event_date, guest_count, venue_id, building, room, service_timing, dietary_notes, notes, status
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    event_id,
                    event_name,
                    event_date,
                    int(guests) if guests is not None else None,
                    venue_id or None,
                    event['building'] or None,
                    event['room'] or None,
                    event['service_timing'] or None,
                    event['dietary_notes'] or None,
                    event['notes'] or None,
                    status
                ))

                for idx, line in enumerate(lines):
                    menu_item_id = line.get('menu_item_id')
                    line_name = line.get('menu_item_name') or ''
                    line_section = line.get('menu_section')
                    line_descriptor = line.get('menu_descriptor')
                    recipe_id = line.get('recipe_id')

                    if menu_item_id:
                        cur.execute("""
                            SELECT id, name, menu_section, menu_descriptor
                            FROM banquet_menu_items
                            WHERE id = %s
                        """, (menu_item_id,))
                        menu_item = cur.fetchone()
                        if not menu_item:
                            continue
                        line_name = line_name or menu_item.get('name') or line_name
                        line_section = line_section or menu_item.get('menu_section')
                        line_descriptor = line_descriptor or menu_item.get('menu_descriptor')
                        if not recipe_id:
                            cur.execute("""
                                SELECT recipe_id
                                FROM banquet_menu_item_recipes
                                WHERE menu_item_id = %s
                                ORDER BY id
                                LIMIT 1
                            """, (menu_item_id,))
                            match = cur.fetchone()
                            recipe_id = match.get('recipe_id') if match else None
                    else:
                        menu_item_id = resolve_or_create_banquet_menu_item(cur, line, venue_id)
                        if not menu_item_id:
                            continue
                        upsert_menu_item_recipe_link(cur, menu_item_id, recipe_id)

                    cur.execute("""
                        INSERT INTO banquet_event_menu_items (
                            event_id, menu_item_id, menu_item_name, recipe_id, quantity, quantity_unit,
                            menu_section, menu_descriptor, notes, sort_order
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        event_id,
                        menu_item_id,
                        line_name,
                        recipe_id,
                        line['quantity'],
                        line['quantity_unit'] or 'each',
                        line_section or None,
                        line_descriptor or None,
                        line['notes'] or None,
                        idx
                    ))

                conn.commit()
                flash('Banquet event created.', 'success')
                cur.close()
                return redirect(url_for('banquet_event_edit', event_id=event_id))
            except Exception:
                conn.rollback()
                flash('Error creating event.', 'error')

        lines = lines or []
    else:
        lines = []

    menu_items = list_banquet_menu_items(cur, event.get('venue_id') or '')
    cur.close()
    return render_template(
        'banquet_event_form.html',
        page_title='New Banquet Event',
        event=event,
        lines=lines,
        menu_items=menu_items,
        venues=venues,
        banquet_venue_name=banquet_venue_name,
        section_options=BANQUET_MENU_SECTION_OPTIONS,
        status_options=BANQUET_STATUS_CHOICES
    )

@app.route('/banquet-planner/events/<event_id>/edit', methods=['GET', 'POST'])
@login_required
def banquet_event_edit(event_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))

    event = get_banquet_event(cur, event_id)
    if not event:
        cur.close()
        flash('Event not found.', 'error')
        return redirect(url_for('banquet_planner'))

    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    banquet_venue_id = banquet_venue.get('id') or ''
    banquet_venue_name = banquet_venue.get('name') or 'Banquets'
    if banquet_venue_id:
        event['venue_id'] = banquet_venue_id

    if request.method == 'POST':
        errors = []
        event_name = clean_menu_text(request.form.get('name'))
        event_date_raw = (request.form.get('event_date') or '').strip()
        venue_id = banquet_venue_id

        status = (request.form.get('status') or 'planning').strip().lower()
        if status not in {choice[0] for choice in BANQUET_STATUS_CHOICES}:
            status = 'planning'

        guests = parse_float_field(request.form.get('guests'), 'Guests', errors, required=False, min_value=1)
        event_date = None
        try:
            event_date = datetime.strptime(event_date_raw, '%Y-%m-%d').date()
        except ValueError:
            errors.append('Event date is required.')

        lines, line_errors = parse_event_recipe_lines(request)
        errors.extend(line_errors)
        if not lines:
            errors.append('Add at least one menu line for the event.')
        if not event_name:
            errors.append('Event name is required.')

        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
            event = {
                **event,
                'name': event_name,
                'event_date': event_date_raw,
                'guests': request.form.get('guests') or '',
                'venue_id': venue_id,
                'building': clean_menu_text(request.form.get('building')),
                'room': clean_menu_text(request.form.get('room')),
                'service_timing': clean_menu_text(request.form.get('service_timing')),
                'dietary_notes': clean_menu_text(request.form.get('dietary_notes')),
                'notes': clean_menu_text(request.form.get('notes')),
                'status': status
            }
        else:
            try:
                cur.execute("""
                    UPDATE banquet_events
                    SET name = %s,
                        event_date = %s,
                        guest_count = %s,
                        venue_id = %s,
                        building = %s,
                        room = %s,
                        service_timing = %s,
                        dietary_notes = %s,
                        notes = %s,
                        status = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, (
                    event_name,
                    event_date,
                    int(guests) if guests is not None else None,
                    venue_id or None,
                    clean_menu_text(request.form.get('building')) or None,
                    clean_menu_text(request.form.get('room')) or None,
                    clean_menu_text(request.form.get('service_timing')) or None,
                    clean_menu_text(request.form.get('dietary_notes')) or None,
                    clean_menu_text(request.form.get('notes')) or None,
                    status,
                    event_id
                ))

                cur.execute("DELETE FROM banquet_event_menu_items WHERE event_id = %s", (event_id,))
                for idx, line in enumerate(lines):
                    menu_item_id = line.get('menu_item_id')
                    line_name = line.get('menu_item_name') or ''
                    line_section = line.get('menu_section')
                    line_descriptor = line.get('menu_descriptor')
                    recipe_id = line.get('recipe_id')

                    if menu_item_id:
                        cur.execute("""
                            SELECT id, name, menu_section, menu_descriptor
                            FROM banquet_menu_items
                            WHERE id = %s
                        """, (menu_item_id,))
                        menu_item = cur.fetchone()
                        if not menu_item:
                            continue
                        line_name = line_name or menu_item.get('name') or line_name
                        line_section = line_section or menu_item.get('menu_section')
                        line_descriptor = line_descriptor or menu_item.get('menu_descriptor')
                        if not recipe_id:
                            cur.execute("""
                                SELECT recipe_id
                                FROM banquet_menu_item_recipes
                                WHERE menu_item_id = %s
                                ORDER BY id
                                LIMIT 1
                            """, (menu_item_id,))
                            match = cur.fetchone()
                            recipe_id = match.get('recipe_id') if match else None
                    else:
                        menu_item_id = resolve_or_create_banquet_menu_item(cur, line, venue_id)
                        if not menu_item_id:
                            continue
                        upsert_menu_item_recipe_link(cur, menu_item_id, recipe_id)

                    cur.execute("""
                        INSERT INTO banquet_event_menu_items (
                            event_id, menu_item_id, menu_item_name, recipe_id, quantity, quantity_unit,
                            menu_section, menu_descriptor, notes, sort_order
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        event_id,
                        menu_item_id,
                        line_name,
                        recipe_id,
                        line['quantity'],
                        line['quantity_unit'] or 'each',
                        line_section or None,
                        line_descriptor or None,
                        line['notes'] or None,
                        idx
                    ))
                conn.commit()
                flash('Banquet event updated.', 'success')
                cur.close()
                return redirect(url_for('banquet_event_edit', event_id=event_id))
            except Exception:
                conn.rollback()
                flash('Error updating event.', 'error')

    if request.method == 'GET':
        lines = get_banquet_event_lines(cur, event_id)
    else:
        lines = parse_event_recipe_lines(request)[0]

    for line in lines:
        if 'notes' not in line:
            line['notes'] = line.get('line_notes')

    menu_items = list_banquet_menu_items(cur, event.get('venue_id') or '')
    cur.close()
    return render_template(
        'banquet_event_form.html',
        page_title='Edit Banquet Event',
        event=event,
        lines=lines,
        menu_items=menu_items,
        venues=venues,
        banquet_venue_name=banquet_venue_name,
        section_options=BANQUET_MENU_SECTION_OPTIONS,
        status_options=BANQUET_STATUS_CHOICES
    )

@app.route('/banquet-planner/events/<event_id>/delete', methods=['POST'])
@login_required
def banquet_event_delete(event_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))

    cur.execute("SELECT id, name FROM banquet_events WHERE id = %s", (event_id,))
    event = cur.fetchone()
    if not event:
        cur.close()
        flash('Event not found.', 'error')
        return redirect(url_for('banquet_planner'))
    try:
        cur.execute("DELETE FROM banquet_events WHERE id = %s", (event_id,))
        conn.commit()
        flash(f"Deleted event: {event['name']}", 'success')
    except Exception:
        conn.rollback()
        flash('Error deleting event.', 'error')
    cur.close()
    return redirect(url_for('banquet_planner'))

@app.route('/banquet-planner/events/<event_id>/apply-template', methods=['POST'])
@login_required
def banquet_event_apply_template(event_id):
    # Backwards-compatible endpoint: now applies the full menu catalog for the event venue.
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_event_edit', event_id=event_id))

    cur.execute("SELECT id, venue_id, guest_count FROM banquet_events WHERE id = %s", (event_id,))
    event = cur.fetchone()
    if not event:
        cur.close()
        flash('Event not found.', 'error')
        return redirect(url_for('banquet_planner'))

    rows = list_banquet_menu_items(cur, event.get('venue_id') or '')
    if not rows:
        cur.close()
        flash('No menu catalog items found for this venue.', 'error')
        return redirect(url_for('banquet_event_edit', event_id=event_id))

    guests = int(event.get('guest_count') or 0)
    try:
        existing_lines = get_banquet_event_lines(cur, event_id)
        sort_base = len(existing_lines)
        for idx, item in enumerate(rows):
            qty = 1
            if guests > 0 and qty == 1:
                qty = guests
            line = {
                'menu_item_name': item.get('name') or '',
                'menu_section': item.get('menu_section'),
                'menu_descriptor': item.get('menu_descriptor'),
                'recipe_id': item.get('linked_recipe_id')
            }
            menu_item_id = resolve_or_create_banquet_menu_item(cur, line, event.get('venue_id') or '')
            upsert_menu_item_recipe_link(cur, menu_item_id, line.get('recipe_id'))
            cur.execute("""
                INSERT INTO banquet_event_menu_items (
                    event_id, menu_item_id, menu_item_name, recipe_id, quantity, quantity_unit,
                    menu_section, menu_descriptor, notes, sort_order
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                event_id,
                menu_item_id,
                line.get('menu_item_name'),
                line.get('recipe_id'),
                qty,
                'each',
                line.get('menu_section'),
                line.get('menu_descriptor'),
                'Loaded from menu catalog',
                sort_base + idx
            ))
        conn.commit()
        flash('Loaded menu catalog lines to event.', 'success')
    except Exception:
        conn.rollback()
        flash('Error loading menu catalog.', 'error')
    cur.close()
    return redirect(url_for('banquet_event_edit', event_id=event_id))

@app.route('/banquet-planner/events/<event_id>/lines/<int:line_id>/clone-menu-item', methods=['POST'])
@login_required
def banquet_event_line_clone_menu_item(event_id, line_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))

    cur.execute("""
        SELECT e.id,
               e.name,
               e.venue_id,
               emi.id AS line_id,
               emi.menu_item_id
        FROM banquet_events e
        JOIN banquet_event_menu_items emi ON emi.event_id = e.id
        WHERE e.id = %s AND emi.id = %s
    """, (event_id, line_id))
    row = cur.fetchone()
    if not row or not row.get('menu_item_id'):
        cur.close()
        flash('Could not find the selected menu line to clone.', 'error')
        return redirect(url_for('banquet_event_edit', event_id=event_id))

    try:
        clone_label = clean_menu_text(request.form.get('clone_label')) or f"Custom {row.get('name') or 'Event'}"
        clone_item = clone_banquet_menu_item(
            cur,
            row.get('menu_item_id'),
            row.get('venue_id'),
            clone_label=clone_label
        )
        if not clone_item:
            conn.rollback()
            cur.close()
            flash('Unable to clone menu item.', 'error')
            return redirect(url_for('banquet_event_edit', event_id=event_id))

        cur.execute("""
            SELECT recipe_id
            FROM banquet_menu_item_recipes
            WHERE menu_item_id = %s
            ORDER BY id
            LIMIT 1
        """, (clone_item['id'],))
        primary = cur.fetchone()
        recipe_id = primary.get('recipe_id') if primary else None

        cur.execute("""
            UPDATE banquet_event_menu_items
            SET menu_item_id = %s,
                menu_item_name = %s,
                menu_section = %s,
                menu_descriptor = %s,
                recipe_id = %s
            WHERE id = %s AND event_id = %s
        """, (
            clone_item['id'],
            clone_item.get('name'),
            clone_item.get('menu_section'),
            clone_item.get('menu_descriptor'),
            recipe_id,
            line_id,
            event_id
        ))
        conn.commit()

        return_to = url_for('banquet_event_edit', event_id=event_id)
        flash('Created custom copy. Update components, then return to event.', 'success')
        cur.close()
        return redirect(url_for('banquet_menu_item_edit', menu_item_id=clone_item['id'], return_to=return_to))
    except Exception:
        conn.rollback()
        flash('Error creating custom menu copy.', 'error')
        cur.close()
        return redirect(url_for('banquet_event_edit', event_id=event_id))

@app.route('/banquet-planner/templates/import', methods=['GET', 'POST'])
@login_required
def banquet_template_import():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    banquet_venue_id = banquet_venue.get('id') or ''
    banquet_venue_name = banquet_venue.get('name') or 'Banquets'

    if request.method == 'POST':
        venue_id = banquet_venue_id
        source_pdf = request.files.get('menu_pdf')
        if not source_pdf or not source_pdf.filename.lower().endswith('.pdf'):
            flash('Upload a PDF menu file.', 'error')
        else:
            try:
                parsed_items = parse_catering_pdf_to_template_items(source_pdf.stream)
            except Exception:
                parsed_items = []
            if not parsed_items:
                flash('Could not parse menu items from the PDF.', 'error')
            else:
                recipes = get_plated_recipes_for_venue(cur, venue_id)
                try:
                    created_count = 0
                    updated_count = 0
                    for item in parsed_items:
                        menu_name = clean_menu_text(item.get('item_name'))
                        if not menu_name:
                            continue
                        cur.execute("""
                            SELECT id
                            FROM banquet_menu_items
                            WHERE LOWER(name) = LOWER(%s)
                              AND (venue_id = %s OR venue_id IS NULL)
                            ORDER BY CASE WHEN venue_id = %s THEN 0 ELSE 1 END, created_at
                            LIMIT 1
                        """, (menu_name, venue_id or None, venue_id or None))
                        existing = cur.fetchone()
                        matched_recipe_id = auto_match_menu_recipe_id(menu_name, recipes)
                        if existing:
                            cur.execute("""
                                UPDATE banquet_menu_items
                                SET menu_section = COALESCE(NULLIF(%s, ''), menu_section),
                                    menu_descriptor = COALESCE(NULLIF(%s, ''), menu_descriptor),
                                    venue_id = COALESCE(venue_id, %s),
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE id = %s
                            """, (
                                item.get('menu_section') or '',
                                item.get('menu_descriptor') or '',
                                venue_id or None,
                                existing['id']
                            ))
                            menu_item_id = existing['id']
                            updated_count += 1
                        else:
                            menu_item_id = generate_id('bmi_')
                            cur.execute("""
                                INSERT INTO banquet_menu_items (id, name, venue_id, menu_section, menu_descriptor, notes)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, (
                                menu_item_id,
                                menu_name,
                                venue_id or None,
                                item.get('menu_section') or None,
                                item.get('menu_descriptor') or None,
                                f"Imported from {source_pdf.filename}"
                            ))
                            created_count += 1

                        if matched_recipe_id:
                            upsert_menu_item_recipe_link(cur, menu_item_id, matched_recipe_id)

                    conn.commit()
                    flash(f'Imported menu catalog: {created_count} created, {updated_count} updated.', 'success')
                    cur.close()
                    return redirect(url_for('banquet_planner'))
                except Exception:
                    conn.rollback()
                    flash('Error saving imported menu catalog.', 'error')

    menu_items = list_banquet_menu_items(cur)
    menu_item_ids = [row.get('id') for row in menu_items if row.get('id')]
    summaries = get_banquet_menu_item_component_summaries(cur, menu_item_ids)
    for item in menu_items:
        summary = summaries.get(item.get('id'), {'recipes': [], 'ingredients': []})
        item['recipe_components'] = summary.get('recipes', [])
        item['ingredient_components'] = summary.get('ingredients', [])
        item['yield_display'] = f"{format_number(to_float(item.get('base_yield_qty')) or 1)} {item.get('base_yield_unit') or 'each'}"
    cur.close()
    return render_template(
        'banquet_template_import.html',
        venues=venues,
        menu_items=menu_items,
        banquet_venue_name=banquet_venue_name
    )

def save_banquet_menu_item(cur, menu_item_id, form_state, ingredient_rows):
    has_base_yield = db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_qty') and db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_unit')
    has_choice_columns = db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_group') and db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_weight_percent')
    if menu_item_id:
        if has_base_yield:
            cur.execute("""
                UPDATE banquet_menu_items
                SET name = %s,
                    venue_id = %s,
                    menu_section = %s,
                    menu_descriptor = %s,
                    base_yield_qty = %s,
                    base_yield_unit = %s,
                    notes = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['base_yield_qty'],
                form_state['base_yield_unit'],
                form_state['notes'],
                menu_item_id
            ))
        else:
            cur.execute("""
                UPDATE banquet_menu_items
                SET name = %s,
                    venue_id = %s,
                    menu_section = %s,
                    menu_descriptor = %s,
                    notes = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['notes'],
                menu_item_id
            ))
    else:
        menu_item_id = generate_id('bmi_')
        if has_base_yield:
            cur.execute("""
                INSERT INTO banquet_menu_items (
                    id, name, venue_id, menu_section, menu_descriptor, base_yield_qty, base_yield_unit, notes
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                menu_item_id,
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['base_yield_qty'],
                form_state['base_yield_unit'],
                form_state['notes']
            ))
        else:
            cur.execute("""
                INSERT INTO banquet_menu_items (
                    id, name, venue_id, menu_section, menu_descriptor, notes
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                menu_item_id,
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['notes']
            ))

    cur.execute("DELETE FROM banquet_menu_item_recipes WHERE menu_item_id = %s", (menu_item_id,))
    cur.execute("DELETE FROM banquet_menu_item_ingredients WHERE menu_item_id = %s", (menu_item_id,))

    for component in form_state.get('recipe_components', []):
        if has_choice_columns:
            cur.execute("""
                INSERT INTO banquet_menu_item_recipes (menu_item_id, recipe_id, quantity, unit, choice_group, choice_weight_percent)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                menu_item_id,
                component['recipe_id'],
                component['quantity'],
                component['unit'],
                component.get('choice_group'),
                component.get('choice_weight_percent')
            ))
        else:
            cur.execute("""
                INSERT INTO banquet_menu_item_recipes (menu_item_id, recipe_id, quantity, unit)
                VALUES (%s, %s, %s, %s)
            """, (
                menu_item_id,
                component['recipe_id'],
                component['quantity'],
                component['unit']
            ))

    for row in ingredient_rows:
        cur.execute("""
            INSERT INTO banquet_menu_item_ingredients (menu_item_id, ingredient_id, quantity, unit)
            VALUES (%s, %s, %s, %s)
        """, (
            menu_item_id,
            row['ingredient_id'],
            row['quantity'],
            row['unit'] or None
        ))
    return menu_item_id

def render_banquet_menu_item_form(cur, mode='new', menu_item_id=None, form_state=None, ingredient_rows=None, errors=None, return_to=None):
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    banquet_venue_id = banquet_venue.get('id') or ''
    banquet_venue_name = banquet_venue.get('name') or 'Banquets'

    if form_state is None:
        form_state = {
            'name': '',
            'venue_id': banquet_venue_id or None,
            'menu_section': '',
            'menu_descriptor': '',
            'base_yield_qty': 1,
            'base_yield_unit': 'each',
            'notes': '',
            'recipe_components': []
        }
        if menu_item_id:
            existing = get_banquet_menu_item(cur, menu_item_id)
            if existing:
                form_state.update({
                    'name': existing.get('name') or '',
                    'venue_id': existing.get('venue_id') or banquet_venue_id or None,
                    'menu_section': existing.get('menu_section') or '',
                    'menu_descriptor': existing.get('menu_descriptor') or '',
                    'base_yield_qty': to_float(existing.get('base_yield_qty')) or 1,
                    'base_yield_unit': existing.get('base_yield_unit') or 'each',
                    'notes': existing.get('notes') or '',
                    'recipe_components': get_banquet_menu_item_recipe_components(cur, menu_item_id)
                })
    if ingredient_rows is None:
        ingredient_rows = get_banquet_menu_item_ingredients(cur, menu_item_id) if menu_item_id else []

    component_recipes = get_component_recipes_for_venue(cur, banquet_venue_id)
    for recipe in component_recipes:
        recipe_type = (recipe.get('recipe_type') or '').upper()
        recipe['display_label'] = f"{recipe.get('name')}{f' ({recipe_type})' if recipe_type else ''}"
    cur.execute("SELECT id, name, unit, category FROM ingredients ORDER BY name")
    ingredients = cur.fetchall()
    for ingredient in ingredients:
        unit_label = ingredient.get('unit')
        ingredient['display_label'] = f"{ingredient.get('name')}{f' ({unit_label})' if unit_label else ''}"

    recipe_label_map = {row['id']: row.get('display_label') or row.get('name') for row in component_recipes}
    ingredient_label_map = {row['id']: row.get('display_label') or row.get('name') for row in ingredients}

    for row in form_state.get('recipe_components', []):
        if isinstance(row, dict):
            row['display_label'] = recipe_label_map.get(row.get('recipe_id'), row.get('recipe_name') or '')
    for row in ingredient_rows:
        if isinstance(row, dict):
            row['display_label'] = ingredient_label_map.get(row.get('ingredient_id'), row.get('ingredient_name') or '')

    return render_template(
        'banquet_menu_item_form.html',
        mode=mode,
        menu_item_id=menu_item_id,
        form_state=form_state,
        ingredient_rows=ingredient_rows,
        component_recipes=component_recipes,
        ingredients=ingredients,
        section_options=BANQUET_MENU_SECTION_OPTIONS,
        banquet_venue_name=banquet_venue_name,
        errors=errors or [],
        return_to=return_to
    )

@app.route('/banquet-planner/menu-items/new', methods=['GET', 'POST'])
@login_required
def banquet_menu_item_new():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))
    return_to = normalize_return_path(request.values.get('return_to'))

    venues = get_active_venues(cur)
    banquet_venue_id = (resolve_banquet_venue(venues) or {}).get('id') or ''
    component_recipes = get_component_recipes_for_venue(cur, banquet_venue_id)
    valid_recipe_ids = {row['id'] for row in component_recipes}
    cur.execute("SELECT id FROM ingredients")
    valid_ingredient_ids = {row['id'] for row in cur.fetchall()}

    if request.method == 'POST':
        errors = []
        base_yield_qty = parse_float_field(
            request.form.get('base_yield_qty'),
            'Menu yield quantity',
            errors,
            required=True,
            min_value=0.0001
        )
        base_yield_unit = normalize_unit(clean_menu_text(request.form.get('base_yield_unit'))) or 'each'
        form_state = {
            'name': clean_menu_text(request.form.get('name')),
            'venue_id': banquet_venue_id or None,
            'menu_section': clean_menu_text(request.form.get('menu_section')) or None,
            'menu_descriptor': clean_menu_text(request.form.get('menu_descriptor')) or None,
            'base_yield_qty': base_yield_qty if base_yield_qty is not None else 1,
            'base_yield_unit': base_yield_unit,
            'notes': clean_menu_text(request.form.get('notes')) or None,
            'recipe_components': []
        }
        if not form_state['name']:
            errors.append('Menu item name is required.')

        recipe_components, component_errors = parse_menu_item_recipe_lines(request, valid_recipe_ids)
        form_state['recipe_components'] = recipe_components
        errors.extend(component_errors)

        ingredient_rows, ingredient_errors = parse_menu_item_ingredient_lines(request, valid_ingredient_ids)
        errors.extend(ingredient_errors)

        if errors:
            response = render_banquet_menu_item_form(
                cur,
                mode='new',
                menu_item_id=None,
                form_state=form_state,
                ingredient_rows=ingredient_rows,
                errors=sorted(set(errors)),
                return_to=return_to
            )
            cur.close()
            return response

        try:
            save_banquet_menu_item(cur, None, form_state, ingredient_rows)
            conn.commit()
            flash('Banquet menu item created.', 'success')
            cur.close()
            return redirect(return_to or url_for('banquet_template_import'))
        except Exception:
            conn.rollback()
            response = render_banquet_menu_item_form(
                cur,
                mode='new',
                menu_item_id=None,
                form_state=form_state,
                ingredient_rows=ingredient_rows,
                errors=['Error saving menu item.'],
                return_to=return_to
            )
            cur.close()
            return response

    response = render_banquet_menu_item_form(cur, mode='new', menu_item_id=None, return_to=return_to)
    cur.close()
    return response

@app.route('/banquet-planner/menu-items/<menu_item_id>/edit', methods=['GET', 'POST'])
@login_required
def banquet_menu_item_edit(menu_item_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))
    return_to = normalize_return_path(request.values.get('return_to'))

    existing = get_banquet_menu_item(cur, menu_item_id)
    if not existing:
        cur.close()
        flash('Menu item not found.', 'error')
        return redirect(url_for('banquet_template_import'))

    venues = get_active_venues(cur)
    banquet_venue_id = (resolve_banquet_venue(venues) or {}).get('id') or existing.get('venue_id') or ''
    component_recipes = get_component_recipes_for_venue(cur, banquet_venue_id)
    valid_recipe_ids = {row['id'] for row in component_recipes}
    cur.execute("SELECT id FROM ingredients")
    valid_ingredient_ids = {row['id'] for row in cur.fetchall()}

    if request.method == 'POST':
        errors = []
        base_yield_qty = parse_float_field(
            request.form.get('base_yield_qty'),
            'Menu yield quantity',
            errors,
            required=True,
            min_value=0.0001
        )
        base_yield_unit = normalize_unit(clean_menu_text(request.form.get('base_yield_unit'))) or 'each'
        form_state = {
            'name': clean_menu_text(request.form.get('name')),
            'venue_id': banquet_venue_id or None,
            'menu_section': clean_menu_text(request.form.get('menu_section')) or None,
            'menu_descriptor': clean_menu_text(request.form.get('menu_descriptor')) or None,
            'base_yield_qty': base_yield_qty if base_yield_qty is not None else 1,
            'base_yield_unit': base_yield_unit,
            'notes': clean_menu_text(request.form.get('notes')) or None,
            'recipe_components': []
        }
        if not form_state['name']:
            errors.append('Menu item name is required.')

        recipe_components, component_errors = parse_menu_item_recipe_lines(request, valid_recipe_ids)
        form_state['recipe_components'] = recipe_components
        errors.extend(component_errors)

        ingredient_rows, ingredient_errors = parse_menu_item_ingredient_lines(request, valid_ingredient_ids)
        errors.extend(ingredient_errors)

        if errors:
            response = render_banquet_menu_item_form(
                cur,
                mode='edit',
                menu_item_id=menu_item_id,
                form_state=form_state,
                ingredient_rows=ingredient_rows,
                errors=sorted(set(errors)),
                return_to=return_to
            )
            cur.close()
            return response

        try:
            save_banquet_menu_item(cur, menu_item_id, form_state, ingredient_rows)
            conn.commit()
            flash('Banquet menu item updated.', 'success')
            cur.close()
            return redirect(return_to or url_for('banquet_template_import'))
        except Exception:
            conn.rollback()
            response = render_banquet_menu_item_form(
                cur,
                mode='edit',
                menu_item_id=menu_item_id,
                form_state=form_state,
                ingredient_rows=ingredient_rows,
                errors=['Error saving menu item.'],
                return_to=return_to
            )
            cur.close()
            return response

    response = render_banquet_menu_item_form(cur, mode='edit', menu_item_id=menu_item_id, return_to=return_to)
    cur.close()
    return response

@app.route('/banquet-planner/menu-items/<menu_item_id>/delete', methods=['POST'])
@login_required
def banquet_menu_item_delete(menu_item_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if not banquet_tables_ready(cur):
        cur.close()
        flash('Banquet tables are missing. Run migrations first.', 'error')
        return redirect(url_for('banquet_planner'))

    cur.execute("SELECT id, name FROM banquet_menu_items WHERE id = %s", (menu_item_id,))
    menu_item = cur.fetchone()
    if not menu_item:
        cur.close()
        flash('Menu item not found.', 'error')
        return redirect(url_for('banquet_template_import'))
    try:
        cur.execute("DELETE FROM banquet_menu_items WHERE id = %s", (menu_item_id,))
        conn.commit()
        flash(f"Deleted menu item: {menu_item['name']}", 'success')
    except Exception:
        conn.rollback()
        flash('Error deleting menu item.', 'error')
    cur.close()
    return redirect(url_for('banquet_template_import'))

@app.route('/banquet-planner/templates/<template_id>/delete', methods=['POST'])
@login_required
def banquet_template_delete(template_id):
    # Backwards-compatible alias while old UI links still exist.
    return banquet_menu_item_delete(template_id)

@app.route('/banquet-planner/shopping')
@login_required
def banquet_shopping():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    selected_venue = banquet_venue.get('id') or ''
    start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
    datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
    grouped = defaultdict(lambda: defaultdict(list))
    for row in datasets['shopping_ingredients']:
        vendor = row.get('vendor') or 'Unassigned Vendor'
        category = row.get('category') or 'Uncategorized'
        grouped[vendor][category].append(row)
    cur.close()
    return render_template(
        'banquet_shopping.html',
        venues=[banquet_venue] if selected_venue else venues,
        selected_venue=selected_venue,
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        grouped=grouped,
        ingredient_total=len(datasets['shopping_ingredients']),
        total_cost=datasets['shopping_total_cost']
    )

@app.route('/banquet-planner/prep-weekly')
@login_required
def banquet_prep_weekly():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    selected_venue = banquet_venue.get('id') or ''
    start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
    datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
    event_names = {event['id']: event['name'] for event in datasets['events']}
    for row in datasets['weekly_prep']:
        row['used_in_event_names'] = [event_names[event_id] for event_id in row['used_in_events'] if event_id in event_names]
    cur.close()
    return render_template(
        'banquet_prep_weekly.html',
        venues=[banquet_venue] if selected_venue else venues,
        selected_venue=selected_venue,
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        weekly_prep=datasets['weekly_prep']
    )

@app.route('/banquet-planner/prep-daily')
@login_required
def banquet_prep_daily():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    selected_venue = banquet_venue.get('id') or ''
    start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
    datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
    cur.close()
    return render_template(
        'banquet_prep_daily.html',
        venues=[banquet_venue] if selected_venue else venues,
        selected_venue=selected_venue,
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        daily_groups=datasets['daily_groups']
    )

@app.route('/banquet-planner/export/excel')
@login_required
def banquet_export_excel():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    selected_venue = banquet_venue.get('id') or ''
    start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
    datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Event Summary'
    headers = ['Date', 'Event', 'Venue', 'Guests', 'Menu Item', 'Section', 'Recipe', 'Qty', 'Unit', 'Descriptor', 'Line Mods', 'Estimated Cost']
    ws_summary.append(headers)
    for col in ws_summary[1]:
        col.font = Font(bold=True, color='FFFFFF')
        col.fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    for event in datasets['events']:
        for line in event['lines']:
            ws_summary.append([
                event.get('event_date').isoformat() if event.get('event_date') else '',
                event.get('name'),
                event.get('venue_name'),
                event.get('guests'),
                line.get('menu_item_name'),
                line.get('menu_section'),
                ', '.join(r.get('name') for r in line.get('recipes', []) if r.get('name')) or (line.get('recipe') or {}).get('name'),
                line.get('quantity'),
                line.get('quantity_unit'),
                line.get('menu_descriptor'),
                line.get('notes'),
                line.get('estimated_cost_total')
            ])

    ws_shopping = wb.create_sheet('Shopping List')
    ws_shopping.append(['Vendor', 'Category', 'Ingredient', 'Qty', 'Unit', 'Price/Unit', 'Ext Cost', 'Vendor Code', 'G-Code', 'Used In'])
    for col in ws_shopping[1]:
        col.font = Font(bold=True, color='FFFFFF')
        col.fill = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
    for ing in datasets['shopping_ingredients']:
        ws_shopping.append([
            ing.get('vendor') or 'Unassigned',
            ing.get('category'),
            ing.get('name'),
            ing.get('display_quantity'),
            ing.get('display_unit'),
            ing.get('cost_per_unit'),
            ing.get('ext_cost'),
            ing.get('vendor_code'),
            ing.get('g_code'),
            ', '.join(ing.get('used_in') or [])
        ])

    ws_weekly = wb.create_sheet('Weekly Prep')
    ws_weekly.append(['Prep Recipe', 'Type', 'Required Output', 'Yield', 'Required Batches', 'Estimated Cost', 'Used In Events', 'Used In Menu Items'])
    for col in ws_weekly[1]:
        col.font = Font(bold=True, color='FFFFFF')
        col.fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
    event_name_map = {event['id']: event['name'] for event in datasets['events']}
    for row in datasets['weekly_prep']:
        ws_weekly.append([
            row.get('recipe_name'),
            (row.get('recipe_type') or 'other'),
            f"{row['display_required']['quantity']} {row['display_required']['unit']}",
            f"{format_number(row.get('yield_qty'))} {row.get('yield_unit') or ''}",
            round(row.get('required_batches') or 0, 4),
            row.get('estimated_cost'),
            ', '.join(event_name_map.get(event_id, event_id) for event_id in row.get('used_in_events', [])),
            ', '.join(row.get('used_in_menu_items', []))
        ])

    ws_daily = wb.create_sheet('Daily Prep')
    ws_daily.append(['Date', 'Event', 'Menu Item', 'Descriptor', 'Qty', 'Unit', 'Recipe', 'Line Mods'])
    for col in ws_daily[1]:
        col.font = Font(bold=True, color='FFFFFF')
        col.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    for day in datasets['daily_groups']:
        for event in day['events']:
            for line in event.get('lines', []):
                ws_daily.append([
                    day['date'].isoformat() if day.get('date') else '',
                    event.get('name'),
                    line.get('menu_item_name'),
                    line.get('menu_descriptor'),
                    line.get('quantity'),
                    line.get('quantity_unit'),
                    ', '.join(r.get('name') for r in line.get('recipes', []) if r.get('name')) or (line.get('recipe') or {}).get('name'),
                    line.get('notes')
                ])

    for ws in [ws_summary, ws_shopping, ws_weekly, ws_daily]:
        for column_cells in ws.columns:
            max_length = 0
            letter = column_cells[0].column_letter
            for cell in column_cells:
                value = '' if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 56)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    venue_name = banquet_venue.get('name') if banquet_venue else 'banquets'
    filename = f"{make_safe_filename(venue_name)}_{start_date.isoformat()}_{end_date.isoformat()}_banquet_planner.xlsx"

    cur.close()
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/banquet-planner/packet/print')
@login_required
def banquet_packet_print():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    selected_venue = banquet_venue.get('id') or ''
    start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
    datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
    cur.close()
    venue_name = banquet_venue.get('name') if banquet_venue else 'Banquets'
    return render_template(
        'banquet_packet_print.html',
        venue_name=venue_name,
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        generated_at=datetime.now().strftime('%b %d, %Y %I:%M %p'),
        datasets=datasets
    )

@app.route('/recipes')
@login_required
def recipes():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    venues = get_active_venues(cur)
    search_query = (request.args.get('q') or '').strip()
    selected_recipe_type = normalize_recipe_type(request.args.get('recipe_type'))
    selected_category = (request.args.get('category') or '').strip()
    if selected_category.lower() == 'all':
        selected_category = ''

    selected_venue = (request.args.get('venue') or '').strip()
    if selected_venue and venues:
        selected_venue_ids = {row['id'] for row in venues}
        if selected_venue not in selected_venue_ids:
            selected_venue = ''

    cur.execute("""
        SELECT COALESCE(NULLIF(TRIM(category), ''), 'Uncategorized') AS category
        FROM recipes
        GROUP BY 1
        ORDER BY 1
    """)
    category_options = [row['category'] for row in cur.fetchall()]
    if selected_category and selected_category not in category_options:
        selected_category = ''

    has_recipe_venues = db_table_exists(cur, 'public.recipe_venues') and db_table_exists(cur, 'public.venues')
    if has_recipe_venues:
        cur.execute("""
            SELECT r.*,
                   COALESCE(ic.ingredient_count, 0) AS ingredient_count,
                   COALESCE(string_agg(DISTINCT v.name, ', ' ORDER BY v.name), '') AS venue_names,
                   COALESCE(string_agg(DISTINCT rv.venue_id, ',' ORDER BY rv.venue_id), '') AS venue_ids
            FROM recipes r
            LEFT JOIN (
                SELECT recipe_id, COUNT(*) AS ingredient_count
                FROM recipe_ingredients
                GROUP BY recipe_id
            ) ic ON ic.recipe_id = r.id
            LEFT JOIN recipe_venues rv ON rv.recipe_id = r.id
            LEFT JOIN venues v ON v.id = rv.venue_id AND v.active = TRUE
            WHERE (%s = '' OR r.name ILIKE %s OR COALESCE(r.menu_descriptor, '') ILIKE %s)
              AND (%s IS NULL OR r.recipe_type = %s)
              AND (
                    %s = '' OR
                    COALESCE(NULLIF(TRIM(r.category), ''), 'Uncategorized') = %s
              )
              AND (
                    %s = '' OR
                    EXISTS (
                        SELECT 1
                        FROM recipe_venues rvf
                        WHERE rvf.recipe_id = r.id AND rvf.venue_id = %s
                    ) OR
                    NOT EXISTS (
                        SELECT 1
                        FROM recipe_venues rvn
                        WHERE rvn.recipe_id = r.id
                    )
              )
            GROUP BY r.id, ic.ingredient_count
            ORDER BY r.name
        """, (
            search_query,
            f"%{search_query}%",
            f"%{search_query}%",
            selected_recipe_type,
            selected_recipe_type,
            selected_category,
            selected_category,
            selected_venue,
            selected_venue
        ))
    else:
        cur.execute("""
            SELECT r.*, COUNT(ri.id) as ingredient_count
            FROM recipes r
            LEFT JOIN recipe_ingredients ri ON r.id = ri.recipe_id
            WHERE (%s = '' OR r.name ILIKE %s OR COALESCE(r.menu_descriptor, '') ILIKE %s)
              AND (%s IS NULL OR r.recipe_type = %s)
              AND (
                    %s = '' OR
                    COALESCE(NULLIF(TRIM(r.category), ''), 'Uncategorized') = %s
              )
            GROUP BY r.id
            ORDER BY r.name
        """, (
            search_query,
            f"%{search_query}%",
            f"%{search_query}%",
            selected_recipe_type,
            selected_recipe_type,
            selected_category,
            selected_category
        ))

    recipes_list = cur.fetchall()
    menu_count = sum(1 for row in recipes_list if row.get('recipe_type') == 'menu')
    batch_count = sum(1 for row in recipes_list if row.get('recipe_type') == 'batch')
    active_filter_count = sum(1 for value in [
        search_query,
        selected_recipe_type,
        selected_category,
        selected_venue
    ] if value)

    def recipe_list_url(unit_value=None):
        params = {}
        if search_query:
            params['q'] = search_query
        if selected_recipe_type:
            params['recipe_type'] = selected_recipe_type
        if selected_category:
            params['category'] = selected_category
        if selected_venue:
            params['venue'] = selected_venue
        if unit_value:
            params['units'] = unit_value
        return url_for('recipes', **params)

    unit_urls = {
        'auto': recipe_list_url('auto'),
        'imperial': recipe_list_url('imperial'),
        'metric': recipe_list_url('metric')
    }

    cur.close()
    conn.close()

    return render_template(
        'recipes.html',
        recipes=recipes_list,
        venues=venues,
        selected_venue=selected_venue,
        search_query=search_query,
        selected_recipe_type=selected_recipe_type or '',
        selected_category=selected_category,
        category_options=category_options,
        menu_count=menu_count,
        batch_count=batch_count,
        active_filter_count=active_filter_count,
        clear_filters_url=url_for('recipes'),
        unit_urls=unit_urls
    )

@app.route('/menu-costing', methods=['GET', 'POST'])
@login_required
def menu_costing():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("""
        SELECT id, name, yield_qty, yield_unit, recipe_type
        FROM recipes
        ORDER BY name
    """)
    recipes_list = cur.fetchall()

    menu_items = []
    total_cost = 0
    q_factor_percent = 5
    q_amount = 0
    grand_total = 0

    if request.method == 'POST':
        q_factor_raw = (request.form.get('q_factor_percent') or '').strip()
        if q_factor_raw:
            q_factor_percent = q_factor_raw
        recipe_ids = request.form.getlist('menu_recipe_id[]')
        batch_values = request.form.getlist('menu_batches[]')
        menu_items, total_cost, errors = parse_menu_items(cur, unit_system, recipe_ids, batch_values)
        q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
    else:
        q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)

    cur.close()
    conn.close()

    return render_template(
        'menu_costing.html',
        recipes=recipes_list,
        menu_items=menu_items,
        total_cost=total_cost,
        q_factor_percent=q_factor_percent,
        q_amount=q_amount,
        grand_total=grand_total
    )

@app.route('/menu-rollouts')
@login_required
def menu_rollouts():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT mr.*, COUNT(mri.id) as item_count
        FROM menu_rollouts mr
        LEFT JOIN menu_rollout_items mri ON mr.id = mri.rollout_id
        WHERE mr.is_one_off = FALSE
        GROUP BY mr.id
        ORDER BY mr.year DESC NULLS LAST, mr.quarter DESC NULLS LAST, mr.venue, mr.name
    """)
    rollouts = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('menu_rollouts.html', rollouts=rollouts)

MENU_SECTION_OPTIONS = [
    'Appetizers',
    'Salads',
    'Wraps',
    'BBQ Dinner',
    'Handhelds',
    'Fish Fry Friday',
    'Entrees',
    'Sides',
    'Desserts'
]

VENUE_DEFAULTS = [
    'Foxtown Brewing',
    "Renard's",
    'Interurban',
    'Heritage Meats',
    "11's Lounge",
    'Banquets',
    'Foxtown Landing'
]

BANQUET_MENU_SECTION_OPTIONS = [
    'Breakfast',
    'Breakouts',
    'Snacks',
    'Daytime Stations',
    'Reception Displays',
    'Reception Stations',
    'Small Bites',
    "Hors D'Oeuvres",
    'Buffet Packages',
    'Custom Buffet',
    'Plated Dinner',
    'Late Night Snacks',
    'Beverages',
    'Enhancements',
    'Other'
]

BANQUET_PACKAGE_SUBSECTIONS = {
    'STARTERS', 'ENTREES', 'SIDES', 'SWEETS', 'DESSERT', 'DESSERTS', 'COLD', 'WARM',
    'COLD APPETIZERS', 'WARM APPETIZERS'
}

def db_table_exists(cur, table_name):
    cur.execute("SELECT to_regclass(%s) AS table_ref", (table_name,))
    row = cur.fetchone() or {}
    return bool(row.get('table_ref'))

def db_column_exists(cur, table_name, column_name):
    if '.' in table_name:
        schema_name, plain_table = table_name.split('.', 1)
    else:
        schema_name, plain_table = 'public', table_name
    cur.execute("""
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema = %s
          AND table_name = %s
          AND column_name = %s
        LIMIT 1
    """, (schema_name, plain_table, column_name))
    return cur.fetchone() is not None

def get_active_venues(cur):
    if not db_table_exists(cur, 'public.venues'):
        return []
    cur.execute("""
        SELECT id, name, active, sort_order
        FROM venues
        WHERE active = TRUE
        ORDER BY sort_order NULLS LAST, name
    """)
    return cur.fetchall()

def get_recipe_venue_ids(cur, recipe_id):
    if not db_table_exists(cur, 'public.recipe_venues'):
        return []
    cur.execute("""
        SELECT venue_id
        FROM recipe_venues
        WHERE recipe_id = %s
        ORDER BY venue_id
    """, (recipe_id,))
    return [row['venue_id'] for row in cur.fetchall()]

def parse_recipe_venue_ids(request, cur, errors):
    raw_ids = request.form.getlist('recipe_venue_ids[]')
    candidate_ids = sorted({(value or '').strip() for value in raw_ids if (value or '').strip()})
    if not candidate_ids:
        return []
    if not db_table_exists(cur, 'public.venues'):
        return []
    cur.execute("SELECT id FROM venues WHERE id = ANY(%s)", (candidate_ids,))
    valid_ids = {row['id'] for row in cur.fetchall()}
    invalid_ids = [value for value in candidate_ids if value not in valid_ids]
    if invalid_ids:
        errors.append('One or more selected venues were invalid.')
    return [value for value in candidate_ids if value in valid_ids]

def normalize_quarter(value):
    if not value:
        return None
    value = value.strip().upper()
    if value in ('1', 'Q1'):
        return 'Q1'
    if value in ('2', 'Q2'):
        return 'Q2'
    if value in ('3', 'Q3'):
        return 'Q3'
    if value in ('4', 'Q4'):
        return 'Q4'
    return value

@app.route('/menu-rollouts/new', methods=['GET', 'POST'])
@login_required
def menu_rollout_new():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("""
        SELECT id, name, yield_qty, yield_unit, recipe_type, menu_descriptor
        FROM recipes
        WHERE recipe_type = 'menu' OR recipe_type IS NULL
        ORDER BY name
    """)
    recipes_list = cur.fetchall()
    venues = get_active_venues(cur)
    venue_options = venues if venues else [{'id': '', 'name': name} for name in VENUE_DEFAULTS]

    menu_items = []
    menu_groups = []
    total_cost = 0
    ingredient_master = []
    ingredient_total_cost = 0
    batch_recipes = []
    q_factor_percent = 5
    default_target_percent = 20
    q_amount = 0
    grand_total = 0
    rollout_data = {
        'name': '',
        'venue': '',
        'year': '',
        'quarter': '',
        'notes': '',
        'q_factor_percent': q_factor_percent,
        'target_food_cost_percent': default_target_percent
    }
    pricing_summary = {
        'avg_food_cost_percent': None,
        'avg_menu_price': None,
        'priced_line_count': 0
    }

    if request.method == 'POST':
        rollout_name = (request.form.get('rollout_name') or request.form.get('name') or '').strip()
        venue = (request.form.get('venue') or '').strip()
        year_value = (request.form.get('year') or '').strip()
        quarter = normalize_quarter(request.form.get('quarter'))
        notes = (request.form.get('notes') or '').strip()
        q_factor_raw = (request.form.get('q_factor_percent') or '').strip()
        q_factor_percent = q_factor_raw if q_factor_raw else q_factor_percent
        target_raw = (request.form.get('target_food_cost_percent') or '').strip()
        default_target_percent = target_raw if target_raw else default_target_percent

        rollout_data = {
            'name': rollout_name,
            'venue': venue,
            'year': year_value,
            'quarter': quarter,
            'notes': notes,
            'q_factor_percent': q_factor_percent,
            'target_food_cost_percent': default_target_percent
        }

        year = int(year_value) if year_value.isdigit() else None
        if not rollout_name and venue and year and quarter:
            rollout_name = f"{venue} {quarter} {year}"
            rollout_data['name'] = rollout_name

        recipe_ids = request.form.getlist('menu_recipe_id[]')
        recipe_names = request.form.getlist('menu_recipe_name[]')
        batch_values = request.form.getlist('menu_batches[]')
        menu_prices = request.form.getlist('menu_price[]')
        target_values = request.form.getlist('menu_target_percent[]')
        section_values = request.form.getlist('menu_section[]')
        descriptor_values = request.form.getlist('menu_descriptor[]')
        menu_items, total_cost, errors = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target_percent,
            section_values,
            descriptor_values
        )
        menu_groups = group_menu_items(menu_items) if menu_items else []
        for idx, recipe_id in enumerate(recipe_ids):
            recipe_name = (recipe_names[idx] if idx < len(recipe_names) else '').strip()
            if recipe_name and not recipe_id:
                errors.append(f'Recipe "{recipe_name}" was not found. Select it from the list.')
        q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
        pricing_summary = apply_menu_pricing(menu_items, default_target_percent)
        if menu_items:
            ingredient_master, ingredient_total_cost, batch_recipes = build_rollout_breakdown(
                cur,
                unit_system,
                menu_items
            )

        if not rollout_name:
            errors.append('Menu name is required.')

        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
        else:
            rollout_id = generate_id('menu_')
            try:
                cur.execute("""
                    INSERT INTO menu_rollouts (id, name, venue, year, quarter, notes, q_factor_percent, target_food_cost_percent, is_one_off)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, FALSE)
                """, (
                    rollout_id,
                    rollout_name,
                    venue or None,
                    year,
                    quarter or None,
                    notes or None,
                    to_float(q_factor_percent),
                    to_float(default_target_percent)
                ))

                for item in menu_items:
                    cur.execute("""
                        INSERT INTO menu_rollout_items (id, rollout_id, recipe_id, batches, menu_price, target_food_cost_percent, popularity_score, menu_section, menu_descriptor)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        generate_id('mri_'),
                        rollout_id,
                        item['recipe_id'],
                        item['batches'],
                        item.get('menu_price'),
                        item.get('target_food_cost_percent'),
                        item.get('popularity_score'),
                        item.get('menu_section'),
                        item.get('menu_descriptor')
                    ))

                conn.commit()
                cur.close()
                conn.close()
                flash('Menu rollout created', 'success')
                return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))
            except Exception:
                conn.rollback()
                flash('Error saving menu rollout', 'error')

    cur.close()
    conn.close()

    return render_template(
        'menu_rollout_form.html',
        mode='new',
        rollout=rollout_data,
        recipes=recipes_list,
        menu_items=menu_items,
        menu_groups=menu_groups,
        total_cost=total_cost,
        q_factor_percent=q_factor_percent,
        q_amount=q_amount,
        grand_total=grand_total,
        pricing_summary=pricing_summary,
        default_target_percent=to_float(default_target_percent) or 20,
        ingredient_master=ingredient_master,
        ingredient_total_cost=ingredient_total_cost,
        batch_recipes=batch_recipes,
        section_options=MENU_SECTION_OPTIONS,
        venue_options=venue_options
    )

@app.route('/menu-rollouts/<rollout_id>', methods=['GET', 'POST'])
@login_required
def menu_rollout_edit(rollout_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
    rollout = cur.fetchone()
    if not rollout:
        cur.close()
        conn.close()
        flash('Menu rollout not found', 'error')
        return redirect(url_for('menu_rollouts'))

    cur.execute("SELECT id, name, yield_qty, yield_unit, recipe_type, menu_descriptor FROM recipes ORDER BY name")
    recipes_list = cur.fetchall()
    venues = get_active_venues(cur)
    venue_options = venues if venues else [{'id': '', 'name': name} for name in VENUE_DEFAULTS]

    menu_items = []
    menu_groups = []
    total_cost = 0
    ingredient_master = []
    ingredient_total_cost = 0
    batch_recipes = []
    q_factor_percent = rollout.get('q_factor_percent') if rollout and rollout.get('q_factor_percent') is not None else 5
    default_target_percent = rollout.get('target_food_cost_percent') if rollout and rollout.get('target_food_cost_percent') is not None else 20
    q_amount = 0
    grand_total = 0
    pricing_summary = {
        'avg_food_cost_percent': None,
        'avg_menu_price': None,
        'priced_line_count': 0
    }

    if request.method == 'POST':
        rollout_name = (request.form.get('rollout_name') or request.form.get('name') or '').strip()
        venue = (request.form.get('venue') or '').strip()
        year_value = (request.form.get('year') or '').strip()
        quarter = normalize_quarter(request.form.get('quarter'))
        notes = (request.form.get('notes') or '').strip()
        q_factor_raw = (request.form.get('q_factor_percent') or '').strip()
        q_factor_percent = q_factor_raw if q_factor_raw else q_factor_percent
        target_raw = (request.form.get('target_food_cost_percent') or '').strip()
        default_target_percent = target_raw if target_raw else default_target_percent

        rollout = dict(rollout)
        rollout.update({
            'name': rollout_name,
            'venue': venue,
            'year': year_value,
            'quarter': quarter,
            'notes': notes,
            'q_factor_percent': q_factor_percent,
            'target_food_cost_percent': default_target_percent
        })

        year = int(year_value) if year_value.isdigit() else None
        if not rollout_name and venue and year and quarter:
            rollout_name = f"{venue} {quarter} {year}"
            rollout['name'] = rollout_name

        recipe_ids = request.form.getlist('menu_recipe_id[]')
        recipe_names = request.form.getlist('menu_recipe_name[]')
        batch_values = request.form.getlist('menu_batches[]')
        menu_prices = request.form.getlist('menu_price[]')
        target_values = request.form.getlist('menu_target_percent[]')
        section_values = request.form.getlist('menu_section[]')
        descriptor_values = request.form.getlist('menu_descriptor[]')
        menu_items, total_cost, errors = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target_percent,
            section_values,
            descriptor_values
        )
        menu_groups = group_menu_items(menu_items) if menu_items else []
        for idx, recipe_id in enumerate(recipe_ids):
            recipe_name = (recipe_names[idx] if idx < len(recipe_names) else '').strip()
            if recipe_name and not recipe_id:
                errors.append(f'Recipe "{recipe_name}" was not found. Select it from the list.')
        q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
        pricing_summary = apply_menu_pricing(menu_items, default_target_percent)
        if menu_items:
            ingredient_master, ingredient_total_cost, batch_recipes = build_rollout_breakdown(
                cur,
                unit_system,
                menu_items
            )

        if not rollout_name:
            errors.append('Menu name is required.')

        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
        else:
            try:
                cur.execute("""
                    UPDATE menu_rollouts
                    SET name = %s,
                        venue = %s,
                        year = %s,
                        quarter = %s,
                        notes = %s,
                        q_factor_percent = %s,
                        target_food_cost_percent = %s
                    WHERE id = %s
                """, (
                    rollout_name,
                    venue or None,
                    year,
                    quarter or None,
                    notes or None,
                    to_float(q_factor_percent),
                    to_float(default_target_percent),
                    rollout_id
                ))

                cur.execute("DELETE FROM menu_rollout_items WHERE rollout_id = %s", (rollout_id,))
                for item in menu_items:
                    cur.execute("""
                        INSERT INTO menu_rollout_items (id, rollout_id, recipe_id, batches, menu_price, target_food_cost_percent, popularity_score, menu_section, menu_descriptor)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        generate_id('mri_'),
                        rollout_id,
                        item['recipe_id'],
                        item['batches'],
                        item.get('menu_price'),
                        item.get('target_food_cost_percent'),
                        item.get('popularity_score'),
                        item.get('menu_section'),
                        item.get('menu_descriptor')
                    ))

                conn.commit()
                cur.close()
                conn.close()
                flash('Menu rollout updated', 'success')
                return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))
            except Exception:
                conn.rollback()
                flash('Error updating menu rollout', 'error')

    if request.method == 'GET':
        cur.execute("""
            SELECT mri.recipe_id,
                   mri.batches,
                   mri.menu_price,
                   mri.target_food_cost_percent,
                   mri.menu_section,
                   mri.menu_descriptor,
                   r.name
            FROM menu_rollout_items mri
            JOIN recipes r ON r.id = mri.recipe_id
            WHERE mri.rollout_id = %s
            ORDER BY r.name
        """, (rollout_id,))
        saved_items = cur.fetchall()

        recipe_ids = [row['recipe_id'] for row in saved_items]
        batch_values = [row['batches'] for row in saved_items]
        menu_prices = [row.get('menu_price') for row in saved_items]
        target_values = [row.get('target_food_cost_percent') for row in saved_items]
        section_values = [row.get('menu_section') for row in saved_items]
        descriptor_values = [row.get('menu_descriptor') for row in saved_items]
        if recipe_ids:
            menu_items, total_cost, _ = parse_menu_items(
                cur,
                unit_system,
                recipe_ids,
                batch_values,
                menu_prices,
                target_values,
                None,
                default_target_percent,
                section_values,
                descriptor_values
            )
            q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
            pricing_summary = apply_menu_pricing(menu_items, default_target_percent)
            ingredient_master, ingredient_total_cost, batch_recipes = build_rollout_breakdown(
                cur,
                unit_system,
                menu_items
            )
            menu_groups = group_menu_items(menu_items)

    cur.close()
    conn.close()

    return render_template(
        'menu_rollout_form.html',
        mode='edit',
        rollout=rollout,
        recipes=recipes_list,
        menu_items=menu_items,
        menu_groups=menu_groups,
        total_cost=total_cost,
        q_factor_percent=q_factor_percent,
        q_amount=q_amount,
        grand_total=grand_total,
        pricing_summary=pricing_summary,
        default_target_percent=to_float(default_target_percent) or 20,
        ingredient_master=ingredient_master,
        ingredient_total_cost=ingredient_total_cost,
        batch_recipes=batch_recipes,
        section_options=MENU_SECTION_OPTIONS,
        venue_options=venue_options
    )

@app.route('/menu-rollouts/<rollout_id>/order-guide')
@login_required
def menu_rollout_order_guide(rollout_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
    rollout = cur.fetchone()
    if not rollout:
        cur.close()
        conn.close()
        flash('Menu rollout not found', 'error')
        return redirect(url_for('menu_rollouts'))

    cur.execute("SELECT recipe_id, batches FROM menu_rollout_items WHERE rollout_id = %s", (rollout_id,))
    items = cur.fetchall()
    if not items:
        cur.close()
        conn.close()
        flash('Add recipes to this rollout before exporting an order guide.', 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    recipe_ids = [row['recipe_id'] for row in items]
    batch_values = [row['batches'] for row in items]
    menu_items, _, errors = parse_menu_items(cur, unit_system, recipe_ids, batch_values)
    if errors:
        cur.close()
        conn.close()
        flash(' '.join(sorted(set(errors))), 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    ingredient_master, _, _ = build_rollout_breakdown(cur, unit_system, menu_items)
    if not ingredient_master:
        cur.close()
        conn.close()
        flash('No ingredients found for this rollout.', 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    grouped = {}
    for ing in ingredient_master:
        vendor = ing.get('vendor') or 'Unassigned Vendor'
        category = ing.get('category') or 'Uncategorized'
        grouped.setdefault(vendor, {}).setdefault(category, []).append(ing)

    wb = Workbook()
    existing_titles = set()
    header_fill = PatternFill('solid', fgColor='E2E8F0')
    category_fill = PatternFill('solid', fgColor='DCFCE7')
    stripe_fill = PatternFill('solid', fgColor='F8FAFC')
    header_font = Font(bold=True, color='1F2937')
    category_font = Font(bold=True, color='14532D')
    align = Alignment(vertical='center')
    thin = Side(border_style='thin', color='E5E7EB')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers = ["Ingredient", "Unit", "Vendor Code", "G-Code", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    column_widths = [36, 10, 16, 16, 10, 10, 10, 10, 10, 10]

    first_sheet = True
    for vendor, categories in sorted(grouped.items(), key=lambda item: item[0].lower()):
        ws = wb.active if first_sheet else wb.create_sheet()
        first_sheet = False
        ws.title = sanitize_sheet_title(vendor, existing_titles)

        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = column_widths[col_idx - 1]

        ws.freeze_panes = "A2"

        row_idx = 2
        for category, items_list in sorted(categories.items(), key=lambda item: item[0].lower()):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            cell = ws.cell(row=row_idx, column=1, value=category)
            cell.fill = category_fill
            cell.font = category_font
            cell.alignment = align
            cell.border = border
            row_idx += 1

            zebra = False
            for ing in sorted(items_list, key=lambda item: (item.get('name') or '').lower()):
                row_values = [
                    ing.get('name') or '',
                    ing.get('unit') or ing.get('display_unit') or '',
                    ing.get('vendor_code') or '',
                    ing.get('g_code') or '',
                    '', '', '', '', '', ''
                ]
                for col_idx, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = align
                    cell.border = border
                    if zebra:
                        cell.fill = stripe_fill
                zebra = not zebra
                row_idx += 1

            row_idx += 1

    cur.close()
    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{make_safe_filename(rollout.get('name') or 'order_guide')}_order_guide.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/menu-rollouts/<rollout_id>/pricing-export')
@login_required
def menu_rollout_pricing_export(rollout_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
    rollout = cur.fetchone()
    if not rollout:
        cur.close()
        conn.close()
        flash('Menu rollout not found', 'error')
        return redirect(url_for('menu_rollouts'))

    cur.execute("""
        SELECT mri.recipe_id,
               mri.batches,
               mri.menu_price,
               mri.target_food_cost_percent,
               mri.menu_section,
               mri.menu_descriptor,
               r.name
        FROM menu_rollout_items mri
        JOIN recipes r ON r.id = mri.recipe_id
        WHERE mri.rollout_id = %s
        ORDER BY r.name
    """, (rollout_id,))
    items = cur.fetchall()
    if not items:
        cur.close()
        conn.close()
        flash('Add recipes to this rollout before exporting.', 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    recipe_ids = [row['recipe_id'] for row in items]
    batch_values = [row['batches'] for row in items]
    menu_prices = [row.get('menu_price') for row in items]
    target_values = [row.get('target_food_cost_percent') for row in items]
    section_values = [row.get('menu_section') for row in items]
    descriptor_values = [row.get('menu_descriptor') for row in items]
    default_target = rollout.get('target_food_cost_percent') or 20

    menu_items, _, errors = parse_menu_items(
        cur,
        unit_system,
        recipe_ids,
        batch_values,
        menu_prices,
        target_values,
        None,
        default_target,
        section_values,
        descriptor_values
    )
    if errors:
        cur.close()
        conn.close()
        flash(' '.join(sorted(set(errors))), 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    apply_menu_pricing(menu_items, default_target)
    grouped = group_menu_items(menu_items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Menu Pricing"

    header_fill = PatternFill('solid', fgColor='E2E8F0')
    header_font = Font(bold=True, color='1F2937')
    section_fill = PatternFill('solid', fgColor='FFF7ED')
    section_font = Font(bold=True, color='9A3412')
    align = Alignment(vertical='center')
    thin = Side(border_style='thin', color='E5E7EB')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    headers = [
        "Section",
        "Recipe",
        "Menu Descriptor",
        "Batches",
        "Cost / Batch",
        "Target FC%",
        "Target Food Cost Price",
        "Menu Price",
        "Food Cost %"
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = border

    row_idx = 2
    for group in grouped:
        section = group['section']
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        section_cell = ws.cell(row=row_idx, column=1, value=section)
        section_cell.fill = section_fill
        section_cell.font = section_font
        section_cell.alignment = align
        section_cell.border = border
        row_idx += 1

        for item in group['items']:
            ws.cell(row=row_idx, column=1, value=section).border = border
            ws.cell(row=row_idx, column=2, value=item['recipe']['name']).border = border
            ws.cell(row=row_idx, column=3, value=item.get('menu_descriptor') or '').border = border
            ws.cell(row=row_idx, column=4, value=item['batches']).border = border
            ws.cell(row=row_idx, column=5, value=round(to_float(item['base_cost']), 2)).border = border
            ws.cell(row=row_idx, column=6, value=round(to_float(item.get('target_food_cost_percent')), 1)).border = border
            ws.cell(row=row_idx, column=7, value=round(to_float(item.get('suggested_price')), 2)).border = border
            ws.cell(row=row_idx, column=8, value=item.get('menu_price')).border = border
            fc_val = item.get('food_cost_percent')
            ws.cell(row=row_idx, column=9, value=round(fc_val, 1) if fc_val else None).border = border
            row_idx += 1

        row_idx += 1

    column_widths = [18, 24, 34, 10, 14, 18, 12, 12, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    cur.close()
    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{make_safe_filename(rollout.get('name') or 'menu_rollout')}_menu_pricing.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/menu-rollouts/<rollout_id>/packet')
@login_required
def menu_rollout_packet_export(rollout_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
    rollout = cur.fetchone()
    if not rollout:
        cur.close()
        flash('Menu rollout not found', 'error')
        return redirect(url_for('menu_rollouts'))

    cur.execute("""
        SELECT mri.recipe_id,
               mri.batches,
               mri.menu_price,
               mri.target_food_cost_percent,
               mri.menu_section,
               mri.menu_descriptor,
               r.name
        FROM menu_rollout_items mri
        JOIN recipes r ON r.id = mri.recipe_id
        WHERE mri.rollout_id = %s
        ORDER BY r.name
    """, (rollout_id,))
    items = cur.fetchall()
    if not items:
        cur.close()
        flash('Add recipes to this rollout before exporting a packet.', 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    recipe_ids = [row['recipe_id'] for row in items]
    batch_values = [row['batches'] for row in items]
    menu_prices = [row.get('menu_price') for row in items]
    target_values = [row.get('target_food_cost_percent') for row in items]
    section_values = [row.get('menu_section') for row in items]
    descriptor_values = [row.get('menu_descriptor') for row in items]
    default_target = rollout.get('target_food_cost_percent') or 20

    menu_items, _, errors = parse_menu_items(
        cur,
        unit_system,
        recipe_ids,
        batch_values,
        menu_prices,
        target_values,
        None,
        default_target,
        section_values,
        descriptor_values
    )
    if errors:
        cur.close()
        flash(' '.join(sorted(set(errors))), 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    apply_menu_pricing(menu_items, default_target)
    menu_groups = group_menu_items(menu_items)

    cur.execute("""
        SELECT id, name, category, unit, cost_per_unit, vendor, vendor_code, g_code
        FROM ingredients
    """)
    ingredient_map = {row['id']: row for row in cur.fetchall()}

    ingredient_usage = {}
    rm_component_sets = []
    subrecipe_ids = set()
    for item in menu_items:
        components, total_cost, _ = build_component_tree(cur, item['recipe_id'], item.get('batches') or 1, 0, set(), unit_system)
        menu_descriptor = item.get('menu_descriptor') or ''
        usage_label = menu_descriptor or item['recipe']['name']
        rm_component_sets.append({
            'menu_item': menu_descriptor,
            'recipe_name': item['recipe']['name'],
            'components': components,
            'total_cost': total_cost
        })
        collect_subrecipes_from_components(components, subrecipe_ids)
        collect_ingredient_usage_from_components(
            components,
            usage_label,
            ingredient_usage
        )

    rb_recipes = []
    if subrecipe_ids:
        cur.execute("""
            SELECT id, name, category, yield_qty, yield_unit, recipe_type
            FROM recipes
            WHERE id = ANY(%s)
            ORDER BY name
        """, (list(subrecipe_ids),))
        for recipe in cur.fetchall():
            if normalize_recipe_type(recipe.get('recipe_type')) != 'batch':
                continue
            components, total_cost, _ = build_component_tree(cur, recipe['id'], 1, 0, set(), unit_system)
            yield_qty = to_float(recipe.get('yield_qty'))
            yield_unit = recipe.get('yield_unit') or ''
            cost_per_yield = (total_cost / yield_qty) if yield_qty > 0 else None
            rb_recipes.append({
                'recipe': recipe,
                'components': components,
                'total_cost': total_cost,
                'yield_qty': yield_qty,
                'yield_unit': yield_unit,
                'cost_per_yield': cost_per_yield
            })

    ingredient_master, _, _ = build_rollout_breakdown(cur, unit_system, menu_items)
    cur.close()

    wb = Workbook()
    header_fill = PatternFill('solid', fgColor='E2E8F0')
    header_font = Font(bold=True, color='1F2937')
    section_fill = PatternFill('solid', fgColor='FFF7ED')
    section_font = Font(bold=True, color='9A3412')
    thin = Side(border_style='thin', color='E5E7EB')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def style_headers(sheet, row=1):
        for cell in sheet[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

    # 1) Menu lines
    ws_menu = wb.active
    ws_menu.title = "01 Menu Lines"
    ws_menu.append([
        "Section", "Recipe", "Menu Descriptor", "Cost/Serving", "Target FC%", "Target Food Cost Price", "Menu Price", "Food Cost %"
    ])
    style_headers(ws_menu)
    for group in menu_groups:
        ws_menu.append([group['section'], '', '', '', '', '', '', ''])
        for cell in ws_menu[ws_menu.max_row]:
            cell.fill = section_fill
            cell.font = section_font
        for item in group['items']:
            ws_menu.append([
                group['section'],
                item['recipe']['name'],
                item.get('menu_descriptor') or '',
                round(to_float(item.get('base_cost')), 4),
                round(to_float(item.get('target_food_cost_percent')), 1),
                round(to_float(item.get('suggested_price')), 2),
                item.get('menu_price'),
                round(to_float(item.get('food_cost_percent')), 1) if item.get('food_cost_percent') else None
            ])
    for width, col in [(20, 'A'), (24, 'B'), (36, 'C'), (14, 'D'), (12, 'E'), (18, 'F'), (12, 'G'), (12, 'H')]:
        ws_menu.column_dimensions[col].width = width

    # 2) RM builds
    ws_rm = wb.create_sheet("02 RM Builds")
    ws_rm.append([
        "Menu Descriptor", "RM Recipe", "Type", "Component", "Qty", "Unit", "Ext Cost", "G-Code", "Vendor", "Vendor SKU", "Notes"
    ])
    style_headers(ws_rm)
    for rm in rm_component_sets:
        ws_rm.append([rm['menu_item'], rm['recipe_name'], "RM", rm['recipe_name'], '', '', round(to_float(rm['total_cost']), 4), '', '', '', 'Total plated cost'])
        for row in flatten_components_for_packet(rm['components'], ingredient_map):
            ws_rm.append([
                rm['menu_item'],
                rm['recipe_name'],
                row.get('type'),
                row.get('name'),
                row.get('quantity'),
                row.get('unit'),
                round(to_float(row.get('ext_cost')), 4),
                row.get('g_code'),
                row.get('vendor'),
                row.get('vendor_code'),
                row.get('notes')
            ])
        ws_rm.append([''] * 11)
    for width, col in [(28, 'A'), (24, 'B'), (14, 'C'), (40, 'D'), (10, 'E'), (10, 'F'), (12, 'G'), (14, 'H'), (18, 'I'), (16, 'J'), (30, 'K')]:
        ws_rm.column_dimensions[col].width = width

    # 3) RB batches
    ws_rb = wb.create_sheet("03 RB Batches")
    ws_rb.append([
        "RB Recipe", "Yield Qty", "Yield Unit", "Cost/Batch", "Cost/Yield Unit", "Type", "Component", "Qty", "Unit", "Ext Cost", "G-Code", "Vendor", "Vendor SKU"
    ])
    style_headers(ws_rb)
    for rb in rb_recipes:
        recipe = rb['recipe']
        ws_rb.append([
            recipe['name'],
            rb['yield_qty'] or '',
            rb['yield_unit'] or '',
            round(to_float(rb['total_cost']), 4),
            round(to_float(rb['cost_per_yield']), 4) if rb.get('cost_per_yield') else None,
            "RB",
            recipe['name'],
            '',
            '',
            '',
            '',
            '',
            ''
        ])
        for row in flatten_components_for_packet(rb['components'], ingredient_map):
            ws_rb.append([
                recipe['name'],
                '',
                '',
                '',
                '',
                row.get('type'),
                row.get('name'),
                row.get('quantity'),
                row.get('unit'),
                round(to_float(row.get('ext_cost')), 4),
                row.get('g_code'),
                row.get('vendor'),
                row.get('vendor_code')
            ])
        ws_rb.append([''] * 13)
    for width, col in [(28, 'A'), (10, 'B'), (12, 'C'), (12, 'D'), (14, 'E'), (12, 'F'), (40, 'G'), (10, 'H'), (10, 'I'), (12, 'J'), (14, 'K'), (18, 'L'), (16, 'M')]:
        ws_rb.column_dimensions[col].width = width

    # 4) Ingredient crosswalk
    ws_ing = wb.create_sheet("04 Ingredient Xwalk")
    ws_ing.append(["Ingredient", "Category", "Unit", "Cost/Unit", "G-Code", "Vendor", "Vendor SKU", "Used In RM Items"])
    style_headers(ws_ing)
    for ing in ingredient_master:
        used_in = sorted(ingredient_usage.get(ing['id'], set()))
        ws_ing.append([
            ing.get('name'),
            ing.get('category'),
            ing.get('unit'),
            round(to_float(ing.get('cost_per_unit')), 5) if ing.get('cost_per_unit') is not None else None,
            ing.get('g_code'),
            ing.get('vendor'),
            ing.get('vendor_code'),
            ', '.join(used_in)
        ])
    for width, col in [(32, 'A'), (18, 'B'), (10, 'C'), (12, 'D'), (14, 'E'), (18, 'F'), (16, 'G'), (48, 'H')]:
        ws_ing.column_dimensions[col].width = width

    # 5) Entry order guide
    ws_steps = wb.create_sheet("05 Entry Order")
    ws_steps.append(["Step", "Action"])
    style_headers(ws_steps)
    ws_steps.append([1, "Update ingredient records in Acumatica first (cost/uom/vendor/G-code)."])
    ws_steps.append([2, f"Enter RB batch recipes ({len(rb_recipes)} total), including yields and components."])
    ws_steps.append([3, f"Enter RM plated recipes ({len(menu_items)} total), linking RB components where used."])
    ws_steps.append([4, "Apply menu pricing and verify target food cost % against target food cost price."])
    ws_steps.append([5, "Final review: run where-used checks for every ingredient with G-code mapping."])
    ws_steps.column_dimensions['A'].width = 8
    ws_steps.column_dimensions['B'].width = 110

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{make_safe_filename(rollout.get('name') or 'menu_rollout')}_acumatica_packet.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/menu-rollouts/<rollout_id>/packet-print')
@login_required
def menu_rollout_packet_print(rollout_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
    rollout = cur.fetchone()
    if not rollout:
        cur.close()
        flash('Menu rollout not found', 'error')
        return redirect(url_for('menu_rollouts'))

    cur.execute("""
        SELECT mri.recipe_id,
               mri.batches,
               mri.menu_price,
               mri.target_food_cost_percent,
               mri.menu_section,
               mri.menu_descriptor,
               r.name
        FROM menu_rollout_items mri
        JOIN recipes r ON r.id = mri.recipe_id
        WHERE mri.rollout_id = %s
        ORDER BY r.name
    """, (rollout_id,))
    items = cur.fetchall()
    if not items:
        cur.close()
        flash('Add recipes to this rollout before printing a packet.', 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    recipe_ids = [row['recipe_id'] for row in items]
    batch_values = [row['batches'] for row in items]
    menu_prices = [row.get('menu_price') for row in items]
    target_values = [row.get('target_food_cost_percent') for row in items]
    section_values = [row.get('menu_section') for row in items]
    descriptor_values = [row.get('menu_descriptor') for row in items]
    default_target = rollout.get('target_food_cost_percent') or 20

    menu_items, _, errors = parse_menu_items(
        cur,
        unit_system,
        recipe_ids,
        batch_values,
        menu_prices,
        target_values,
        None,
        default_target,
        section_values,
        descriptor_values
    )
    if errors:
        cur.close()
        flash(' '.join(sorted(set(errors))), 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    apply_menu_pricing(menu_items, default_target)
    menu_groups = group_menu_items(menu_items)

    cur.execute("""
        SELECT id, name, category, unit, cost_per_unit, vendor, vendor_code, g_code
        FROM ingredients
    """)
    ingredient_map = {row['id']: row for row in cur.fetchall()}

    ingredient_usage = {}
    rm_component_sets = []
    subrecipe_ids = set()
    for item in menu_items:
        components, total_cost, _ = build_component_tree(cur, item['recipe_id'], item.get('batches') or 1, 0, set(), unit_system)
        menu_descriptor = item.get('menu_descriptor') or ''
        usage_label = menu_descriptor or item['recipe']['name']
        rm_component_sets.append({
            'menu_item': menu_descriptor,
            'recipe_name': item['recipe']['name'],
            'total_cost': total_cost,
            'rows': flatten_components_for_packet(components, ingredient_map)
        })
        collect_subrecipes_from_components(components, subrecipe_ids)
        collect_ingredient_usage_from_components(components, usage_label, ingredient_usage)

    rb_recipes = []
    if subrecipe_ids:
        cur.execute("""
            SELECT id, name, category, yield_qty, yield_unit, recipe_type
            FROM recipes
            WHERE id = ANY(%s)
            ORDER BY name
        """, (list(subrecipe_ids),))
        for recipe in cur.fetchall():
            if normalize_recipe_type(recipe.get('recipe_type')) != 'batch':
                continue
            components, total_cost, _ = build_component_tree(cur, recipe['id'], 1, 0, set(), unit_system)
            yield_qty = to_float(recipe.get('yield_qty'))
            rb_recipes.append({
                'recipe': recipe,
                'total_cost': total_cost,
                'cost_per_yield': (total_cost / yield_qty) if yield_qty > 0 else None,
                'rows': flatten_components_for_packet(components, ingredient_map)
            })

    ingredient_master, _, _ = build_rollout_breakdown(cur, unit_system, menu_items)
    for ing in ingredient_master:
        ing['used_in'] = ', '.join(sorted(ingredient_usage.get(ing['id'], set())))

    cur.close()

    return render_template(
        'menu_rollout_packet_print.html',
        rollout=rollout,
        menu_groups=menu_groups,
        rm_component_sets=rm_component_sets,
        rb_recipes=rb_recipes,
        ingredient_master=ingredient_master,
        generated_at=datetime.utcnow(),
        default_target_percent=default_target
    )

@app.route('/menu-rollouts/<rollout_id>/print')
@login_required
def menu_rollout_print(rollout_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()

    cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
    rollout = cur.fetchone()
    if not rollout:
        cur.close()
        conn.close()
        flash('Menu rollout not found', 'error')
        return redirect(url_for('menu_rollouts'))

    cur.execute("""
        SELECT mri.recipe_id,
               mri.batches,
               mri.menu_price,
               mri.target_food_cost_percent,
               mri.menu_section,
               mri.menu_descriptor,
               r.name
        FROM menu_rollout_items mri
        JOIN recipes r ON r.id = mri.recipe_id
        WHERE mri.rollout_id = %s
        ORDER BY r.name
    """, (rollout_id,))
    items = cur.fetchall()
    if not items:
        cur.close()
        conn.close()
        flash('Add recipes to this rollout before printing.', 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    recipe_ids = [row['recipe_id'] for row in items]
    batch_values = [row['batches'] for row in items]
    menu_prices = [row.get('menu_price') for row in items]
    target_values = [row.get('target_food_cost_percent') for row in items]
    section_values = [row.get('menu_section') for row in items]
    descriptor_values = [row.get('menu_descriptor') for row in items]
    default_target = rollout.get('target_food_cost_percent') or 20

    menu_items, _, _ = parse_menu_items(
        cur,
        unit_system,
        recipe_ids,
        batch_values,
        menu_prices,
        target_values,
        None,
        default_target,
        section_values,
        descriptor_values
    )
    apply_menu_pricing(menu_items, default_target)
    menu_groups = group_menu_items(menu_items)

    cur.close()
    conn.close()

    return render_template(
        'menu_rollout_print.html',
        rollout=rollout,
        menu_groups=menu_groups,
        default_target_percent=default_target,
        generated_at=datetime.utcnow()
    )

@app.route('/menu-rollouts/<rollout_id>/delete', methods=['POST'])
@login_required
def menu_rollout_delete(rollout_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT id, name FROM menu_rollouts WHERE id = %s", (rollout_id,))
    rollout = cur.fetchone()
    if not rollout:
        cur.close()
        conn.close()
        flash('Menu rollout not found', 'error')
        return redirect(url_for('menu_rollouts'))

    try:
        cur.execute("DELETE FROM menu_rollout_items WHERE rollout_id = %s", (rollout_id,))
        cur.execute("DELETE FROM menu_rollouts WHERE id = %s", (rollout_id,))
        conn.commit()
        cur.close()
        conn.close()
        flash('Menu rollout deleted', 'success')
        return redirect(url_for('menu_rollouts'))
    except Exception:
        conn.rollback()
        cur.close()
        conn.close()
        flash('Error deleting menu rollout', 'error')
        return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

@app.route('/recipes/new', methods=['GET', 'POST'])
@login_required
def recipe_new():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    option_items_input = []
    selected_venue_ids = []

    if request.method == 'POST':
        errors = []
        name = (request.form.get('name') or '').strip()
        category = (request.form.get('category') or '').strip()
        yield_qty = (request.form.get('yield_qty') or '').strip()
        yield_unit = (request.form.get('yield_unit') or '').strip()
        yield_unit = normalize_unit(yield_unit) or yield_unit
        instructions = (request.form.get('instructions') or '').strip()
        menu_descriptor = (request.form.get('menu_descriptor') or '').strip()
        recipe_type = infer_recipe_type(name, request.form.get('recipe_type'))
        if recipe_type == 'menu':
            yield_qty = '1'
            yield_unit = 'serving'

        if not name:
            errors.append('Recipe name is required.')

        selected_venue_ids = parse_recipe_venue_ids(request, cur, errors)

        ingredient_ids = request.form.getlist('ingredient_id[]')
        ingredient_names = request.form.getlist('ingredient_name[]')
        ingredient_qtys = request.form.getlist('ingredient_qty[]')
        ingredient_units = request.form.getlist('ingredient_unit[]')

        for idx, ing_id in enumerate(ingredient_ids):
            ing_id = (ing_id or '').strip()
            ing_name = (ingredient_names[idx] if idx < len(ingredient_names) else '').strip()
            if not ing_id and ing_name:
                errors.append(f'Ingredient \"{ing_name}\" was not found. Select it from the list or create it first.')

        sub_ids = request.form.getlist('sub_recipe_id[]')
        sub_names = request.form.getlist('sub_recipe_name[]')
        for idx, sub_id in enumerate(sub_ids):
            sub_id = (sub_id or '').strip()
            sub_name = (sub_names[idx] if idx < len(sub_names) else '').strip()
            if not sub_id and sub_name:
                errors.append(f'Sub-recipe \"{sub_name}\" was not found. Select it from the list.')

        weighted_option_rows = parse_weighted_options_from_form(request, recipe_type, cur, errors)
        option_group_names = request.form.getlist('option_group_name[]')
        option_item_types = request.form.getlist('option_item_type[]')
        option_item_ids = request.form.getlist('option_item_id[]')
        option_item_names = request.form.getlist('option_item_name[]')
        option_qtys = request.form.getlist('option_qty[]')
        option_units = request.form.getlist('option_unit[]')
        option_weights = request.form.getlist('option_weight[]')
        max_options = max(
            len(option_group_names),
            len(option_item_types),
            len(option_item_ids),
            len(option_item_names),
            len(option_qtys),
            len(option_units),
            len(option_weights),
            0
        )
        for idx in range(max_options):
            group_name = (option_group_names[idx] if idx < len(option_group_names) else '').strip()
            item_type = (option_item_types[idx] if idx < len(option_item_types) else '').strip()
            item_id = (option_item_ids[idx] if idx < len(option_item_ids) else '').strip()
            item_name = (option_item_names[idx] if idx < len(option_item_names) else '').strip()
            qty = (option_qtys[idx] if idx < len(option_qtys) else '').strip()
            unit = (option_units[idx] if idx < len(option_units) else '').strip()
            weight = (option_weights[idx] if idx < len(option_weights) else '').strip()
            if not any([group_name, item_id, item_name, qty, unit, weight]):
                continue
            option_items_input.append({
                'group_name': group_name,
                'item_type': item_type or 'recipe',
                'item_id': item_id,
                'item_name': item_name,
                'quantity': qty,
                'unit': unit,
                'weight_percent': weight
            })

        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
        else:
            recipe_id = generate_id('rec_')
            try:
                cur.execute("""
                    INSERT INTO recipes (id, name, category, yield_qty, yield_unit, instructions, recipe_type, menu_descriptor)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    recipe_id,
                    name,
                    category or None,
                    yield_qty or None,
                    yield_unit or None,
                    instructions or None,
                    recipe_type,
                    menu_descriptor or None
                ))

                for venue_id in selected_venue_ids:
                    cur.execute("""
                        INSERT INTO recipe_venues (recipe_id, venue_id)
                        VALUES (%s, %s)
                        ON CONFLICT (recipe_id, venue_id) DO NOTHING
                    """, (recipe_id, venue_id))

                # Ingredients
                for idx, ing_id in enumerate(ingredient_ids):
                    ing_id = (ing_id or '').strip()
                    if not ing_id:
                        continue
                    qty = (ingredient_qtys[idx] if idx < len(ingredient_qtys) else '').strip()
                    unit = (ingredient_units[idx] if idx < len(ingredient_units) else '').strip()
                    unit = normalize_unit(unit) or unit
                    if qty == '':
                        continue
                    cur.execute("""
                        INSERT INTO recipe_ingredients (id, recipe_id, type, item_id, quantity, unit)
                        VALUES (%s, %s, 'ingredient', %s, %s, %s)
                    """, (
                        generate_id('ri_'),
                        recipe_id,
                        ing_id,
                        qty,
                        unit or None
                    ))

                # Sub-recipes
                sub_qtys = request.form.getlist('sub_recipe_qty[]')
                sub_units = request.form.getlist('sub_recipe_unit[]')

                for idx, sub_id in enumerate(sub_ids):
                    sub_id = (sub_id or '').strip()
                    if not sub_id:
                        continue
                    qty = (sub_qtys[idx] if idx < len(sub_qtys) else '').strip()
                    unit = (sub_units[idx] if idx < len(sub_units) else '').strip()
                    unit = normalize_unit(unit) or unit
                    if qty == '':
                        continue
                    cur.execute("""
                        INSERT INTO recipe_ingredients (id, recipe_id, type, item_id, quantity, unit)
                        VALUES (%s, %s, 'recipe', %s, %s, %s)
                    """, (
                        generate_id('ri_'),
                        recipe_id,
                        sub_id,
                        qty,
                        unit or None
                    ))

                for option in weighted_option_rows:
                    cur.execute("""
                        INSERT INTO recipe_weighted_options (id, recipe_id, group_name, item_type, item_id, quantity, unit, weight_percent)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        generate_id('rwo_'),
                        recipe_id,
                        option['group_name'],
                        option['item_type'],
                        option['item_id'],
                        option['quantity'],
                        option['unit'] or None,
                        option['weight_percent']
                    ))

                conn.commit()
                cur.close()
                conn.close()

                flash('Recipe created', 'success')
                return redirect(url_for('recipe_detail', recipe_id=recipe_id))
            except Exception:
                conn.rollback()
                flash('Error creating recipe', 'error')

    # GET or validation error fallback
    cur.execute("SELECT id, name, unit, category FROM ingredients ORDER BY name")
    ingredients_list = cur.fetchall()
    venues = get_active_venues(cur)

    cur.execute("""
        SELECT id, name, yield_qty, yield_unit, recipe_type
        FROM recipes
        WHERE recipe_type = 'batch' OR recipe_type IS NULL
        ORDER BY name
    """)
    recipes_list = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        'recipe_form.html',
        mode='new',
        recipe={
            'name': (request.form.get('name') if request.method == 'POST' else '') or '',
            'category': (request.form.get('category') if request.method == 'POST' else '') or '',
            'yield_qty': (request.form.get('yield_qty') if request.method == 'POST' else '') or '',
            'yield_unit': (request.form.get('yield_unit') if request.method == 'POST' else '') or '',
            'instructions': (request.form.get('instructions') if request.method == 'POST' else '') or '',
            'recipe_type': (request.form.get('recipe_type') if request.method == 'POST' else '') or '',
            'menu_descriptor': (request.form.get('menu_descriptor') if request.method == 'POST' else '') or ''
        },
        ingredients=ingredients_list,
        recipes=recipes_list,
        venues=venues,
        selected_venue_ids=selected_venue_ids,
        ingredient_items=[],
        subrecipe_items=[],
        option_items=option_items_input
    )

@app.route('/recipes/<recipe_id>/edit', methods=['GET', 'POST'])
@login_required
def recipe_edit(recipe_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    recipe = get_recipe_by_id(cur, recipe_id)
    if not recipe:
        cur.close()
        conn.close()
        flash('Recipe not found', 'error')
        return redirect(url_for('recipes'))

    selected_venue_ids = get_recipe_venue_ids(cur, recipe_id)

    if request.method == 'POST':
        errors = []
        name = (request.form.get('name') or '').strip()
        category = (request.form.get('category') or '').strip()
        yield_qty = (request.form.get('yield_qty') or '').strip()
        yield_unit = (request.form.get('yield_unit') or '').strip()
        yield_unit = normalize_unit(yield_unit) or yield_unit
        instructions = (request.form.get('instructions') or '').strip()
        menu_descriptor = (request.form.get('menu_descriptor') or '').strip()
        recipe_type = infer_recipe_type(name, request.form.get('recipe_type'))
        if recipe_type == 'menu':
            yield_qty = '1'
            yield_unit = 'serving'

        recipe = dict(recipe)
        recipe.update({
            'name': name,
            'category': category,
            'yield_qty': yield_qty,
            'yield_unit': yield_unit,
            'instructions': instructions,
            'recipe_type': recipe_type,
            'menu_descriptor': menu_descriptor
        })

        if not name:
            errors.append('Recipe name is required.')

        selected_venue_ids = parse_recipe_venue_ids(request, cur, errors)

        ingredient_ids = request.form.getlist('ingredient_id[]')
        ingredient_names = request.form.getlist('ingredient_name[]')
        ingredient_qtys = request.form.getlist('ingredient_qty[]')
        ingredient_units = request.form.getlist('ingredient_unit[]')

        for idx, ing_id in enumerate(ingredient_ids):
            ing_id = (ing_id or '').strip()
            ing_name = (ingredient_names[idx] if idx < len(ingredient_names) else '').strip()
            if not ing_id and ing_name:
                errors.append(f'Ingredient \"{ing_name}\" was not found. Select it from the list or create it first.')

        sub_ids = request.form.getlist('sub_recipe_id[]')
        sub_names = request.form.getlist('sub_recipe_name[]')
        for idx, sub_id in enumerate(sub_ids):
            sub_id = (sub_id or '').strip()
            sub_name = (sub_names[idx] if idx < len(sub_names) else '').strip()
            if not sub_id and sub_name:
                errors.append(f'Sub-recipe \"{sub_name}\" was not found. Select it from the list.')

        weighted_option_rows = parse_weighted_options_from_form(request, recipe_type, cur, errors)
        option_group_names = request.form.getlist('option_group_name[]')
        option_item_types = request.form.getlist('option_item_type[]')
        option_item_ids = request.form.getlist('option_item_id[]')
        option_item_names = request.form.getlist('option_item_name[]')
        option_qtys = request.form.getlist('option_qty[]')
        option_units = request.form.getlist('option_unit[]')
        option_weights = request.form.getlist('option_weight[]')

        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
        else:
            try:
                cur.execute("""
                    UPDATE recipes
                    SET name = %s,
                        category = %s,
                        yield_qty = %s,
                        yield_unit = %s,
                        instructions = %s,
                        recipe_type = %s,
                        menu_descriptor = %s
                    WHERE id = %s
                """, (
                    name,
                    category or None,
                    yield_qty or None,
                    yield_unit or None,
                    instructions or None,
                    recipe_type,
                    menu_descriptor or None,
                    recipe_id
                ))

                if db_table_exists(cur, 'public.recipe_venues'):
                    cur.execute("DELETE FROM recipe_venues WHERE recipe_id = %s", (recipe_id,))
                    for venue_id in selected_venue_ids:
                        cur.execute("""
                            INSERT INTO recipe_venues (recipe_id, venue_id)
                            VALUES (%s, %s)
                            ON CONFLICT (recipe_id, venue_id) DO NOTHING
                        """, (recipe_id, venue_id))

                cur.execute("DELETE FROM recipe_ingredients WHERE recipe_id = %s", (recipe_id,))

                for idx, ing_id in enumerate(ingredient_ids):
                    ing_id = (ing_id or '').strip()
                    if not ing_id:
                        continue
                    qty = (ingredient_qtys[idx] if idx < len(ingredient_qtys) else '').strip()
                    unit = (ingredient_units[idx] if idx < len(ingredient_units) else '').strip()
                    unit = normalize_unit(unit) or unit
                    if qty == '':
                        continue
                    cur.execute("""
                        INSERT INTO recipe_ingredients (id, recipe_id, type, item_id, quantity, unit)
                        VALUES (%s, %s, 'ingredient', %s, %s, %s)
                    """, (
                        generate_id('ri_'),
                        recipe_id,
                        ing_id,
                        qty,
                        unit or None
                    ))

                sub_qtys = request.form.getlist('sub_recipe_qty[]')
                sub_units = request.form.getlist('sub_recipe_unit[]')

                for idx, sub_id in enumerate(sub_ids):
                    sub_id = (sub_id or '').strip()
                    if not sub_id:
                        continue
                    qty = (sub_qtys[idx] if idx < len(sub_qtys) else '').strip()
                    unit = (sub_units[idx] if idx < len(sub_units) else '').strip()
                    unit = normalize_unit(unit) or unit
                    if qty == '':
                        continue
                    cur.execute("""
                        INSERT INTO recipe_ingredients (id, recipe_id, type, item_id, quantity, unit)
                        VALUES (%s, %s, 'recipe', %s, %s, %s)
                    """, (
                        generate_id('ri_'),
                        recipe_id,
                        sub_id,
                        qty,
                        unit or None
                    ))

                cur.execute("DELETE FROM recipe_weighted_options WHERE recipe_id = %s", (recipe_id,))
                for option in weighted_option_rows:
                    cur.execute("""
                        INSERT INTO recipe_weighted_options (id, recipe_id, group_name, item_type, item_id, quantity, unit, weight_percent)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        generate_id('rwo_'),
                        recipe_id,
                        option['group_name'],
                        option['item_type'],
                        option['item_id'],
                        option['quantity'],
                        option['unit'] or None,
                        option['weight_percent']
                    ))

                conn.commit()
                cur.close()
                conn.close()

                flash('Recipe updated', 'success')
                return redirect(url_for('recipe_detail', recipe_id=recipe_id))
            except Exception:
                conn.rollback()
                flash('Error updating recipe', 'error')

    # GET or validation error fallback
    components = get_recipe_components(cur, recipe_id)
    ingredient_items = [c for c in components if c.get('type') == 'ingredient']
    subrecipe_items = [c for c in components if c.get('type') == 'recipe']
    if request.method == 'POST':
        option_items = []
        max_options = max(
            len(option_group_names),
            len(option_item_types),
            len(option_item_ids),
            len(option_item_names),
            len(option_qtys),
            len(option_units),
            len(option_weights),
            0
        )
        for idx in range(max_options):
            group_name = (option_group_names[idx] if idx < len(option_group_names) else '').strip()
            item_type = (option_item_types[idx] if idx < len(option_item_types) else '').strip()
            item_id = (option_item_ids[idx] if idx < len(option_item_ids) else '').strip()
            item_name = (option_item_names[idx] if idx < len(option_item_names) else '').strip()
            qty = (option_qtys[idx] if idx < len(option_qtys) else '').strip()
            unit = (option_units[idx] if idx < len(option_units) else '').strip()
            weight = (option_weights[idx] if idx < len(option_weights) else '').strip()
            if not any([group_name, item_id, item_name, qty, unit, weight]):
                continue
            option_items.append({
                'group_name': group_name,
                'item_type': item_type or 'recipe',
                'item_id': item_id,
                'item_name': item_name,
                'quantity': qty,
                'unit': unit,
                'weight_percent': weight
            })
    else:
        option_items = get_recipe_weighted_options(cur, recipe_id)

    cur.execute("SELECT id, name, unit, category FROM ingredients ORDER BY name")
    ingredients_list = cur.fetchall()
    venues = get_active_venues(cur)

    cur.execute("""
        SELECT id, name, yield_qty, yield_unit, recipe_type
        FROM recipes
        WHERE id != %s AND (recipe_type = 'batch' OR recipe_type IS NULL)
        ORDER BY name
    """, (recipe_id,))
    recipes_list = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        'recipe_form.html',
        mode='edit',
        recipe=recipe,
        ingredients=ingredients_list,
        recipes=recipes_list,
        venues=venues,
        selected_venue_ids=selected_venue_ids,
        ingredient_items=ingredient_items,
        subrecipe_items=subrecipe_items,
        option_items=option_items
    )

@app.route('/recipe-generator', methods=['GET', 'POST'])
@login_required
def recipe_generator():
    flash('Recipe Generator has been merged into New Recipe.', 'info')
    return redirect(url_for('recipe_new'))

@app.route('/recipes/<recipe_id>/delete', methods=['POST'])
@login_required
def recipe_delete(recipe_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT id, name FROM recipes WHERE id = %s", (recipe_id,))
    recipe = cur.fetchone()
    if not recipe:
        flash('Recipe not found', 'error')
        return redirect(url_for('recipes'))

    cur.execute("""
        SELECT COUNT(*) AS count
        FROM recipe_ingredients
        WHERE type = 'recipe' AND item_id = %s
    """, (recipe_id,))
    used_in_recipes = int((cur.fetchone() or {}).get('count') or 0)

    cur.execute("""
        SELECT COUNT(*) AS count
        FROM menu_rollout_items
        WHERE recipe_id = %s
    """, (recipe_id,))
    used_in_rollouts = int((cur.fetchone() or {}).get('count') or 0)

    cur.execute("""
        SELECT COUNT(*) AS count
        FROM recipe_weighted_options
        WHERE item_type = 'recipe' AND item_id = %s
    """, (recipe_id,))
    used_in_weighted = int((cur.fetchone() or {}).get('count') or 0)

    blockers = []
    if used_in_recipes:
        blockers.append(f"{used_in_recipes} recipe(s)")
    if used_in_rollouts:
        blockers.append(f"{used_in_rollouts} rollout item(s)")
    if used_in_weighted:
        blockers.append(f"{used_in_weighted} weighted option(s)")

    if blockers:
        flash(f"Can't delete {recipe['name']} — used in {', '.join(blockers)}.", 'error')
        return redirect(url_for('recipe_detail', recipe_id=recipe_id))

    try:
        cur.execute("DELETE FROM recipe_weighted_options WHERE recipe_id = %s", (recipe_id,))
        cur.execute("DELETE FROM recipe_ingredients WHERE recipe_id = %s", (recipe_id,))
        cur.execute("DELETE FROM recipes WHERE id = %s", (recipe_id,))
        conn.commit()
        flash('Recipe deleted', 'success')
        return redirect(url_for('recipes'))
    except Exception:
        conn.rollback()
        flash('Error deleting recipe', 'error')
        return redirect(url_for('recipe_detail', recipe_id=recipe_id))

@app.route('/recipes/<recipe_id>')
@login_required
def recipe_detail(recipe_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    unit_system = get_unit_system()
    
    # Get recipe details
    recipe = get_recipe_by_id(cur, recipe_id)
    
    if not recipe:
        flash('Recipe not found', 'error')
        return redirect(url_for('recipes'))

    recipe_venues = []
    if db_table_exists(cur, 'public.recipe_venues') and db_table_exists(cur, 'public.venues'):
        cur.execute("""
            SELECT v.id, v.name
            FROM recipe_venues rv
            JOIN venues v ON v.id = rv.venue_id
            WHERE rv.recipe_id = %s
            ORDER BY v.name
        """, (recipe_id,))
        recipe_venues = cur.fetchall()
    
    components, total_cost, _ = build_component_tree(cur, recipe_id, 1, 0, set(), unit_system, apply_q_factor=True)
    base_total_cost = None
    q_factor_amount = None
    q_factor_percent = RECIPE_Q_FACTOR_PERCENT
    if total_cost is not None and q_factor_percent and q_factor_percent > 0:
        divisor = 1 + (q_factor_percent / 100)
        base_total_cost = total_cost / divisor
        q_factor_amount = total_cost - base_total_cost
    elif total_cost is not None:
        base_total_cost = total_cost
    yield_qty_float = to_float(recipe.get('yield_qty'))
    cost_per_yield = None
    if total_cost and yield_qty_float > 0:
        cost_per_yield = total_cost / yield_qty_float

    yield_display = smart_quantity(recipe.get('yield_qty'), recipe.get('yield_unit'), unit_system)
    pdf_data = {
        'name': recipe.get('name') or '',
        'venue': recipe.get('source_venue') or '',
        'equipment': recipe.get('equipment') or '',
        'yield': f"{yield_display.get('quantity')} {yield_display.get('unit')}".strip(),
        'ingredients': flatten_components_for_pdf(components),
        'instructions': recipe.get('instructions') or ''
    }
    
    cur.close()
    conn.close()
    
    return render_template(
        'recipe_detail.html',
        recipe=recipe,
        recipe_venues=recipe_venues,
        components=components,
        component_count=len(components),
        total_cost=total_cost,
        cost_per_yield=cost_per_yield,
        base_total_cost=base_total_cost,
        q_factor_percent=RECIPE_Q_FACTOR_PERCENT,
        q_factor_amount=q_factor_amount,
        recipe_pdf=pdf_data
    )

@app.route('/ingredients')
@login_required
def ingredients():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Get all ingredients
    cur.execute("""
        SELECT * FROM ingredients
        ORDER BY category, name
    """)
    
    ingredients_list = cur.fetchall()
    print(f"DEBUG: Found {len(ingredients_list)} ingredients")

    now = datetime.utcnow()
    needs_update_count = 0
    for ingredient in ingredients_list:
        updated_at = ingredient.get('price_updated_at')
        age_days = None
        if updated_at:
            if getattr(updated_at, 'tzinfo', None):
                updated_at = updated_at.replace(tzinfo=None)
            age_days = (now - updated_at).days
        ingredient['price_age_days'] = age_days
        ingredient['needs_price_update'] = (age_days is None) or (age_days >= PRICE_REFRESH_DAYS)
        if ingredient['needs_price_update']:
            needs_update_count += 1
    cur.close()
    conn.close()
    
    return render_template(
        'ingredients.html',
        ingredients=ingredients_list,
        needs_update_count=needs_update_count,
        price_refresh_days=PRICE_REFRESH_DAYS
    )

@app.route('/ingredients/new', methods=['GET', 'POST'])
@login_required
def new_ingredient():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            flash('Ingredient name is required', 'error')
        else:
            unit_value = (request.form.get('unit') or '').strip()
            unit_value = normalize_unit(unit_value) or unit_value
            cost_raw = request.form.get('cost_per_unit')
            cost_value = to_float(cost_raw) if cost_raw not in (None, '') else None
            price_updated = cost_value is not None

            try:
                cur.execute("""
                    INSERT INTO ingredients (id, name, category, unit, vendor, vendor_code, g_code, cost_per_unit, price_updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CASE WHEN %s THEN NOW() ELSE NULL END)
                """, (
                    generate_id('ing_'),
                    name,
                    request.form.get('category') or None,
                    unit_value or None,
                    request.form.get('vendor') or None,
                    request.form.get('vendor_code') or None,
                    request.form.get('g_code') or None,
                    cost_value,
                    price_updated
                ))
                conn.commit()
                cur.close()
                conn.close()

                flash('Ingredient created', 'success')
                return redirect(url_for('ingredients'))
            except Exception as e:
                conn.rollback()
                flash('Error saving ingredient', 'error')

    cur.close()
    conn.close()

    return render_template(
        'new_ingredient.html',
        ingredient={
            'name': '',
            'category': '',
            'unit': '',
            'vendor': '',
            'vendor_code': '',
            'g_code': '',
            'cost_per_unit': ''
        }
    )

@app.route('/ingredients/<ingredient_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_ingredient(ingredient_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            cur.close()
            conn.close()
            flash('Ingredient name is required', 'error')
            return redirect(url_for('edit_ingredient', ingredient_id=ingredient_id))

        cur.execute("SELECT cost_per_unit FROM ingredients WHERE id = %s", (ingredient_id,))
        existing = cur.fetchone() or {}
        existing_cost = to_float(existing.get('cost_per_unit'))
        new_cost_raw = request.form.get('cost_per_unit')
        new_cost = to_float(new_cost_raw) if new_cost_raw not in (None, '') else None
        cost_changed = (new_cost is not None) and (new_cost != existing_cost)

        # Update ingredient
        unit_value = (request.form.get('unit') or '').strip()
        unit_value = normalize_unit(unit_value) or unit_value

        try:
            cur.execute("""
                UPDATE ingredients
                SET name = %s,
                    category = %s,
                    unit = %s,
                    vendor = %s,
                    vendor_code = %s,
                    g_code = %s,
                    cost_per_unit = %s,
                    price_updated_at = CASE WHEN %s THEN NOW() ELSE price_updated_at END
                WHERE id = %s
            """, (
                request.form.get('name'),
                request.form.get('category'),
                unit_value or None,
                request.form.get('vendor'),
                request.form.get('vendor_code') or None,
                request.form.get('g_code') or None,
                new_cost,
                cost_changed,
                ingredient_id
            ))
            conn.commit()
            cur.close()
            conn.close()
            
            flash('Ingredient updated successfully', 'success')
            return redirect(url_for('ingredients'))
        except Exception:
            conn.rollback()
            cur.close()
            conn.close()
            flash('Error updating ingredient', 'error')
            return redirect(url_for('edit_ingredient', ingredient_id=ingredient_id))
    
    # GET - show edit form
    cur.execute("SELECT * FROM ingredients WHERE id = %s", (ingredient_id,))
    ingredient = cur.fetchone()
    
    cur.close()
    conn.close()
    
    if not ingredient:
        flash('Ingredient not found', 'error')
        return redirect(url_for('ingredients'))
    
    return render_template('edit_ingredient.html', ingredient=ingredient)

@app.route('/ingredients/<ingredient_id>/delete', methods=['POST'])
@login_required
def delete_ingredient(ingredient_id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT id, name FROM ingredients WHERE id = %s", (ingredient_id,))
    ingredient = cur.fetchone()
    if not ingredient:
        cur.close()
        conn.close()
        flash('Ingredient not found', 'error')
        return redirect(url_for('ingredients'))

    cur.execute("""
        SELECT COUNT(*) AS usage_count
        FROM recipe_ingredients
        WHERE type = 'ingredient' AND item_id = %s
    """, (ingredient_id,))
    usage = cur.fetchone() or {}
    usage_count = int(usage.get('usage_count') or 0)
    if usage_count > 0:
        cur.close()
        conn.close()
        flash(f"Can't delete {ingredient['name']} — used in {usage_count} recipe(s).", 'error')
        return redirect(url_for('edit_ingredient', ingredient_id=ingredient_id))

    try:
        cur.execute("DELETE FROM ingredients WHERE id = %s", (ingredient_id,))
        conn.commit()
        cur.close()
        conn.close()
        flash('Ingredient deleted', 'success')
        return redirect(url_for('ingredients'))
    except Exception:
        conn.rollback()
        cur.close()
        conn.close()
        flash('Error deleting ingredient', 'error')
        return redirect(url_for('edit_ingredient', ingredient_id=ingredient_id))

@app.route('/api/recipes/search')
@login_required
def api_recipes_search():
    query = (request.args.get('q') or '').strip()
    recipe_type = normalize_recipe_type(request.args.get('type'))
    try:
        limit = int(request.args.get('limit', 20))
    except (TypeError, ValueError):
        limit = 20
    limit = min(max(limit, 1), 100)

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    sql = """
        SELECT id, name, recipe_type, yield_unit
        FROM recipes
        WHERE (%s = '' OR name ILIKE %s)
    """
    params = [query, f"%{query}%"]
    if recipe_type:
        sql += " AND recipe_type = %s"
        params.append(recipe_type)
    sql += " ORDER BY name LIMIT %s"
    params.append(limit)

    cur.execute(sql, tuple(params))
    results = cur.fetchall()
    cur.close()

    return jsonify({'results': results})

@app.route('/api/ingredients/search')
@login_required
def api_ingredients_search():
    query = (request.args.get('q') or '').strip()
    try:
        limit = int(request.args.get('limit', 20))
    except (TypeError, ValueError):
        limit = 20
    limit = min(max(limit, 1), 100)

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, name, unit, category, vendor
        FROM ingredients
        WHERE (%s = '' OR name ILIKE %s)
        ORDER BY name
        LIMIT %s
    """, (query, f"%{query}%", limit))
    results = cur.fetchall()
    cur.close()

    return jsonify({'results': results})

if __name__ == '__main__':
    debug_flag = os.getenv('FLASK_DEBUG', '').lower() in ('1', 'true', 'yes')
    debug_env = os.getenv('FLASK_ENV', '').lower() == 'development'
    debug = debug_flag or debug_env
    port = int(os.getenv('PORT', 5000))
    app.run(debug=debug, host='0.0.0.0', port=port)
