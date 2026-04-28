from datetime import datetime
from io import BytesIO

from config import DEFAULT_Q_FACTOR_PERCENT, DEFAULT_TARGET_FOOD_COST_PERCENT
from db import get_cursor, get_db
from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from helpers.formatting import (
    collect_ingredient_usage_from_components,
    flatten_components_for_packet,
    make_safe_filename,
    sanitize_sheet_title,
)
from helpers.menu import (
    apply_menu_pricing,
    build_rollout_breakdown,
    build_rollout_ops_dataset,
    group_menu_items,
    parse_menu_items,
)
from helpers.recipes import build_component_tree, collect_subrecipes_from_components, compute_q_factor, normalize_recipe_type
from helpers.shared import generate_id, handle_route_error, to_float
from helpers.units import get_unit_system, summarize_yield_pricing

bp = Blueprint('menu', __name__)

@bp.errorhandler(Exception)
def handle_menu_error(error):
    return handle_route_error(error, 'menu')

@bp.route('/menu-costing', methods=['GET', 'POST'])
@login_required
def menu_costing():
    unit_system = get_unit_system()
    with get_cursor() as cur:
        cur.execute("""
            SELECT id, name, yield_qty, yield_unit, recipe_type
            FROM recipes
            ORDER BY name
        """)
        recipes_list = cur.fetchall()

        menu_items = []
        total_cost = 0
        q_factor_percent = DEFAULT_Q_FACTOR_PERCENT
        q_amount = 0
        grand_total = 0

        if request.method == 'POST':
            q_factor_raw = (request.form.get('q_factor_percent') or '').strip()
            if q_factor_raw:
                q_factor_percent = q_factor_raw
            recipe_ids = request.form.getlist('menu_recipe_id[]')
            batch_values = request.form.getlist('menu_batches[]')
            menu_items, total_cost, errors = parse_menu_items(cur, unit_system, recipe_ids, batch_values)
            q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
            if errors:
                flash(' '.join(sorted(set(errors))), 'error')
        else:
            q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)

    return render_template(
        'menu_costing.html',
        recipes=recipes_list,
        menu_items=menu_items,
        total_cost=total_cost,
        q_factor_percent=q_factor_percent,
        q_amount=q_amount,
        grand_total=grand_total
    )

@bp.route('/menu-rollouts')
@login_required
def menu_rollouts():
    with get_cursor() as cur:
        cur.execute("""
            SELECT
                mr.id,
                mr.name,
                mr.venue,
                mr.year,
                mr.quarter,
                (
                    SELECT COUNT(*)
                    FROM menu_rollout_items mri
                    WHERE mri.rollout_id = mr.id
                ) AS item_count
            FROM menu_rollouts mr
            WHERE mr.is_one_off = FALSE
            ORDER BY mr.year DESC NULLS LAST, mr.quarter DESC NULLS LAST, mr.venue, mr.name
        """)
        rollouts = cur.fetchall()

    return render_template('menu_rollouts.html', rollouts=rollouts)

MENU_SECTION_OPTIONS = [
    'Appetizers',
    'Salads',
    'Wraps',
    'BBQ Dinner',
    'Handhelds',
    'Fish Fry Friday',
    'Entrees',
    'Sides',
    'Modifier',
    'Desserts',
    'Other'
]

VENUE_DEFAULTS = [
    'Foxtown Brewing',
    "Renard's",
    'Interurban',
    'Heritage Meats',
    "11's Lounge",
    'Banquets',
    'Foxtown Landing'
]

BANQUET_MENU_SECTION_OPTIONS = [
    'Breakfast',
    'Breakouts',
    'Snacks',
    'Daytime Stations',
    'Reception Displays',
    'Reception Stations',
    'Small Bites',
    "Hors D'Oeuvres",
    'Buffet Packages',
    'Custom Buffet',
    'Plated Dinner',
    'Late Night Snacks',
    'Beverages',
    'Enhancements',
    'Other'
]

BANQUET_PACKAGE_SUBSECTIONS = {
    'STARTERS', 'ENTREES', 'SIDES', 'SWEETS', 'DESSERT', 'DESSERTS', 'COLD', 'WARM',
    'COLD APPETIZERS', 'WARM APPETIZERS'
}

def db_table_exists(cur, table_name):
    cur.execute("SELECT to_regclass(%s) AS table_ref", (table_name,))
    row = cur.fetchone() or {}
    return bool(row.get('table_ref'))

def db_column_exists(cur, table_name, column_name):
    if '.' in table_name:
        schema_name, plain_table = table_name.split('.', 1)
    else:
        schema_name, plain_table = 'public', table_name
    cur.execute("""
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema = %s
          AND table_name = %s
          AND column_name = %s
        LIMIT 1
    """, (schema_name, plain_table, column_name))
    return cur.fetchone() is not None

def ensure_banquet_menu_item_base_yield_columns(cur):
    """Backfill banquet menu item sold-as fields on older databases before save."""
    if not db_table_exists(cur, 'public.banquet_menu_items'):
        return False

    cur.execute("""
        ALTER TABLE banquet_menu_items
        ADD COLUMN IF NOT EXISTS base_yield_qty NUMERIC DEFAULT 1
    """)
    cur.execute("""
        ALTER TABLE banquet_menu_items
        ADD COLUMN IF NOT EXISTS base_yield_unit TEXT DEFAULT 'each'
    """)
    cur.execute("""
        UPDATE banquet_menu_items
        SET base_yield_qty = 1
        WHERE base_yield_qty IS NULL
    """)
    cur.execute("""
        UPDATE banquet_menu_items
        SET base_yield_unit = 'each'
        WHERE base_yield_unit IS NULL OR TRIM(base_yield_unit) = ''
    """)
    return True

def get_active_venues(cur):
    if not db_table_exists(cur, 'public.venues'):
        return []
    cur.execute("""
        SELECT id, name, active, sort_order
        FROM venues
        WHERE active = TRUE
        ORDER BY sort_order NULLS LAST, name
    """)
    return cur.fetchall()

def get_recipe_venue_ids(cur, recipe_id):
    if not db_table_exists(cur, 'public.recipe_venues'):
        return []
    cur.execute("""
        SELECT venue_id
        FROM recipe_venues
        WHERE recipe_id = %s
        ORDER BY venue_id
    """, (recipe_id,))
    return [row['venue_id'] for row in cur.fetchall()]

def parse_recipe_venue_ids(request, cur, errors):
    raw_ids = request.form.getlist('recipe_venue_ids[]')
    candidate_ids = sorted({(value or '').strip() for value in raw_ids if (value or '').strip()})
    if not candidate_ids:
        return []
    if not db_table_exists(cur, 'public.venues'):
        return []
    cur.execute("SELECT id FROM venues WHERE id = ANY(%s)", (candidate_ids,))
    valid_ids = {row['id'] for row in cur.fetchall()}
    invalid_ids = [value for value in candidate_ids if value not in valid_ids]
    if invalid_ids:
        errors.append('One or more selected venues were invalid.')
    return [value for value in candidate_ids if value in valid_ids]

def normalize_quarter(value):
    if not value:
        return None
    value = value.strip().upper()
    if value in ('1', 'Q1'):
        return 'Q1'
    if value in ('2', 'Q2'):
        return 'Q2'
    if value in ('3', 'Q3'):
        return 'Q3'
    if value in ('4', 'Q4'):
        return 'Q4'
    return value

@bp.route('/menu-rollouts/new', methods=['GET', 'POST'])
@login_required
def menu_rollout_new():
    conn = get_db()
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("""
            SELECT id, name, yield_qty, yield_unit, recipe_type, menu_descriptor
            FROM recipes
            WHERE recipe_type = 'menu' OR recipe_type IS NULL
            ORDER BY name
        """)
        recipes_list = cur.fetchall()
        venues = get_active_venues(cur)
        venue_options = venues if venues else [{'id': '', 'name': name} for name in VENUE_DEFAULTS]

        menu_items = []
        menu_groups = []
        total_cost = 0
        ingredient_master = []
        ingredient_total_cost = 0
        batch_recipes = []
        q_factor_percent = DEFAULT_Q_FACTOR_PERCENT
        default_target_percent = DEFAULT_TARGET_FOOD_COST_PERCENT
        q_amount = 0
        grand_total = 0
        rollout_data = {
            'name': '',
            'venue': '',
            'year': '',
            'quarter': '',
            'notes': '',
            'q_factor_percent': q_factor_percent,
            'target_food_cost_percent': default_target_percent
        }
        pricing_summary = {
            'avg_food_cost_percent': None,
            'avg_menu_price': None,
            'priced_line_count': 0
        }

        if request.method == 'POST':
            rollout_name = (request.form.get('rollout_name') or request.form.get('name') or '').strip()
            venue = (request.form.get('venue') or '').strip()
            year_value = (request.form.get('year') or '').strip()
            quarter = normalize_quarter(request.form.get('quarter'))
            notes = (request.form.get('notes') or '').strip()
            q_factor_raw = (request.form.get('q_factor_percent') or '').strip()
            q_factor_percent = q_factor_raw if q_factor_raw else q_factor_percent
            target_raw = (request.form.get('target_food_cost_percent') or '').strip()
            default_target_percent = target_raw if target_raw else default_target_percent

            rollout_data = {
                'name': rollout_name,
                'venue': venue,
                'year': year_value,
                'quarter': quarter,
                'notes': notes,
                'q_factor_percent': q_factor_percent,
                'target_food_cost_percent': default_target_percent
            }

            year = int(year_value) if year_value.isdigit() else None
            if not rollout_name and venue and year and quarter:
                rollout_name = f"{venue} {quarter} {year}"
                rollout_data['name'] = rollout_name

            recipe_ids = request.form.getlist('menu_recipe_id[]')
            recipe_names = request.form.getlist('menu_recipe_name[]')
            batch_values = request.form.getlist('menu_batches[]')
            menu_prices = request.form.getlist('menu_price[]')
            target_values = request.form.getlist('menu_target_percent[]')
            section_values = request.form.getlist('menu_section[]')
            descriptor_values = request.form.getlist('menu_descriptor[]')
            menu_items, total_cost, errors = parse_menu_items(
                cur,
                unit_system,
                recipe_ids,
                batch_values,
                menu_prices,
                target_values,
                None,
                default_target_percent,
                section_values,
                descriptor_values
            )
            menu_groups = group_menu_items(menu_items) if menu_items else []
            for idx, recipe_id in enumerate(recipe_ids):
                recipe_name = (recipe_names[idx] if idx < len(recipe_names) else '').strip()
                if recipe_name and not recipe_id:
                    errors.append(f'Recipe "{recipe_name}" was not found. Select it from the list.')
            q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
            pricing_summary = apply_menu_pricing(menu_items, default_target_percent)
            if menu_items:
                ingredient_master, ingredient_total_cost, batch_recipes = build_rollout_breakdown(
                    cur,
                    unit_system,
                    menu_items
                )

            if not rollout_name:
                errors.append('Menu name is required.')

            if errors:
                flash(' '.join(sorted(set(errors))), 'error')
            else:
                rollout_id = generate_id('menu_')
                try:
                    cur.execute("""
                        INSERT INTO menu_rollouts (id, name, venue, year, quarter, notes, q_factor_percent, target_food_cost_percent, is_one_off)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, FALSE)
                    """, (
                        rollout_id,
                        rollout_name,
                        venue or None,
                        year,
                        quarter or None,
                        notes or None,
                        to_float(q_factor_percent),
                        to_float(default_target_percent)
                    ))

                    for item in menu_items:
                        cur.execute("""
                            INSERT INTO menu_rollout_items (id, rollout_id, recipe_id, batches, menu_price, target_food_cost_percent, popularity_score, menu_section, menu_descriptor)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            generate_id('mri_'),
                            rollout_id,
                            item['recipe_id'],
                            item['batches'],
                            item.get('menu_price'),
                            item.get('target_food_cost_percent'),
                            item.get('popularity_score'),
                            item.get('menu_section'),
                            item.get('menu_descriptor')
                        ))

                    conn.commit()
                    flash('Menu rollout created', 'success')
                    return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))
                except Exception:
                    conn.rollback()
                    flash('Error saving menu rollout', 'error')


        return render_template(
            'menu_rollout_form.html',
            mode='new',
            rollout=rollout_data,
            recipes=recipes_list,
            menu_items=menu_items,
            menu_groups=menu_groups,
            total_cost=total_cost,
            q_factor_percent=q_factor_percent,
            q_amount=q_amount,
            grand_total=grand_total,
            pricing_summary=pricing_summary,
            default_target_percent=to_float(default_target_percent) or 20,
            ingredient_master=ingredient_master,
            ingredient_total_cost=ingredient_total_cost,
            batch_recipes=batch_recipes,
            section_options=MENU_SECTION_OPTIONS,
            venue_options=venue_options
        )

@bp.route('/menu-rollouts/<rollout_id>', methods=['GET', 'POST'])
@login_required
def menu_rollout_edit(rollout_id):
    conn = get_db()
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("SELECT id, name, yield_qty, yield_unit, recipe_type, menu_descriptor FROM recipes ORDER BY name")
        recipes_list = cur.fetchall()
        venues = get_active_venues(cur)
        venue_options = venues if venues else [{'id': '', 'name': name} for name in VENUE_DEFAULTS]

        menu_items = []
        menu_groups = []
        total_cost = 0
        ingredient_master = []
        ingredient_total_cost = 0
        batch_recipes = []
        q_factor_percent = rollout.get('q_factor_percent') if rollout and rollout.get('q_factor_percent') is not None else 5
        default_target_percent = rollout.get('target_food_cost_percent') if rollout and rollout.get('target_food_cost_percent') is not None else 20
        q_amount = 0
        grand_total = 0
        pricing_summary = {
            'avg_food_cost_percent': None,
            'avg_menu_price': None,
            'priced_line_count': 0
        }

        if request.method == 'POST':
            rollout_name = (request.form.get('rollout_name') or request.form.get('name') or '').strip()
            venue = (request.form.get('venue') or '').strip()
            year_value = (request.form.get('year') or '').strip()
            quarter = normalize_quarter(request.form.get('quarter'))
            notes = (request.form.get('notes') or '').strip()
            q_factor_raw = (request.form.get('q_factor_percent') or '').strip()
            q_factor_percent = q_factor_raw if q_factor_raw else q_factor_percent
            target_raw = (request.form.get('target_food_cost_percent') or '').strip()
            default_target_percent = target_raw if target_raw else default_target_percent

            rollout = dict(rollout)
            rollout.update({
                'name': rollout_name,
                'venue': venue,
                'year': year_value,
                'quarter': quarter,
                'notes': notes,
                'q_factor_percent': q_factor_percent,
                'target_food_cost_percent': default_target_percent
            })

            year = int(year_value) if year_value.isdigit() else None
            if not rollout_name and venue and year and quarter:
                rollout_name = f"{venue} {quarter} {year}"
                rollout['name'] = rollout_name

            recipe_ids = request.form.getlist('menu_recipe_id[]')
            recipe_names = request.form.getlist('menu_recipe_name[]')
            batch_values = request.form.getlist('menu_batches[]')
            menu_prices = request.form.getlist('menu_price[]')
            target_values = request.form.getlist('menu_target_percent[]')
            section_values = request.form.getlist('menu_section[]')
            descriptor_values = request.form.getlist('menu_descriptor[]')
            menu_items, total_cost, errors = parse_menu_items(
                cur,
                unit_system,
                recipe_ids,
                batch_values,
                menu_prices,
                target_values,
                None,
                default_target_percent,
                section_values,
                descriptor_values
            )
            menu_groups = group_menu_items(menu_items) if menu_items else []
            for idx, recipe_id in enumerate(recipe_ids):
                recipe_name = (recipe_names[idx] if idx < len(recipe_names) else '').strip()
                if recipe_name and not recipe_id:
                    errors.append(f'Recipe "{recipe_name}" was not found. Select it from the list.')
            q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
            pricing_summary = apply_menu_pricing(menu_items, default_target_percent)
            if menu_items:
                ingredient_master, ingredient_total_cost, batch_recipes = build_rollout_breakdown(
                    cur,
                    unit_system,
                    menu_items
                )

            if not rollout_name:
                errors.append('Menu name is required.')

            if errors:
                flash(' '.join(sorted(set(errors))), 'error')
            else:
                try:
                    cur.execute("""
                        UPDATE menu_rollouts
                        SET name = %s,
                            venue = %s,
                            year = %s,
                            quarter = %s,
                            notes = %s,
                            q_factor_percent = %s,
                            target_food_cost_percent = %s
                        WHERE id = %s
                    """, (
                        rollout_name,
                        venue or None,
                        year,
                        quarter or None,
                        notes or None,
                        to_float(q_factor_percent),
                        to_float(default_target_percent),
                        rollout_id
                    ))

                    cur.execute("DELETE FROM menu_rollout_items WHERE rollout_id = %s", (rollout_id,))
                    for item in menu_items:
                        cur.execute("""
                            INSERT INTO menu_rollout_items (id, rollout_id, recipe_id, batches, menu_price, target_food_cost_percent, popularity_score, menu_section, menu_descriptor)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            generate_id('mri_'),
                            rollout_id,
                            item['recipe_id'],
                            item['batches'],
                            item.get('menu_price'),
                            item.get('target_food_cost_percent'),
                            item.get('popularity_score'),
                            item.get('menu_section'),
                            item.get('menu_descriptor')
                        ))

                    conn.commit()
                    flash('Menu rollout updated', 'success')
                    return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))
                except Exception:
                    conn.rollback()
                    flash('Error updating menu rollout', 'error')

        if request.method == 'GET':
            cur.execute("""
                SELECT mri.recipe_id,
                       mri.batches,
                       mri.menu_price,
                       mri.target_food_cost_percent,
                       mri.menu_section,
                       mri.menu_descriptor,
                       r.name
                FROM menu_rollout_items mri
                JOIN recipes r ON r.id = mri.recipe_id
                WHERE mri.rollout_id = %s
                ORDER BY r.name
            """, (rollout_id,))
            saved_items = cur.fetchall()

            recipe_ids = [row['recipe_id'] for row in saved_items]
            batch_values = [row['batches'] for row in saved_items]
            menu_prices = [row.get('menu_price') for row in saved_items]
            target_values = [row.get('target_food_cost_percent') for row in saved_items]
            section_values = [row.get('menu_section') for row in saved_items]
            descriptor_values = [row.get('menu_descriptor') for row in saved_items]
            if recipe_ids:
                menu_items, total_cost, _ = parse_menu_items(
                    cur,
                    unit_system,
                    recipe_ids,
                    batch_values,
                    menu_prices,
                    target_values,
                    None,
                    default_target_percent,
                    section_values,
                    descriptor_values
                )
                q_factor_percent, q_amount, grand_total = compute_q_factor(total_cost, q_factor_percent)
                pricing_summary = apply_menu_pricing(menu_items, default_target_percent)
                ingredient_master, ingredient_total_cost, batch_recipes = build_rollout_breakdown(
                    cur,
                    unit_system,
                    menu_items
                )
                menu_groups = group_menu_items(menu_items)


        return render_template(
            'menu_rollout_form.html',
            mode='edit',
            rollout=rollout,
            recipes=recipes_list,
            menu_items=menu_items,
            menu_groups=menu_groups,
            total_cost=total_cost,
            q_factor_percent=q_factor_percent,
            q_amount=q_amount,
            grand_total=grand_total,
            pricing_summary=pricing_summary,
            default_target_percent=to_float(default_target_percent) or 20,
            ingredient_master=ingredient_master,
            ingredient_total_cost=ingredient_total_cost,
            batch_recipes=batch_recipes,
            section_options=MENU_SECTION_OPTIONS,
            venue_options=venue_options
        )

@bp.route('/menu-rollouts/<rollout_id>/order-guide')
@login_required
def menu_rollout_order_guide(rollout_id):
    conn = get_db()
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("SELECT recipe_id, batches FROM menu_rollout_items WHERE rollout_id = %s", (rollout_id,))
        items = cur.fetchall()
        if not items:
            flash('Add recipes to this rollout before exporting an order guide.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        recipe_ids = [row['recipe_id'] for row in items]
        batch_values = [row['batches'] for row in items]
        menu_items, _, errors = parse_menu_items(cur, unit_system, recipe_ids, batch_values)
        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        ingredient_master, _, _ = build_rollout_breakdown(cur, unit_system, menu_items)
        if not ingredient_master:
            flash('No ingredients found for this rollout.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        grouped = {}
        for ing in ingredient_master:
            vendor = ing.get('vendor') or 'Unassigned Vendor'
            category = ing.get('category') or 'Uncategorized'
            grouped.setdefault(vendor, {}).setdefault(category, []).append(ing)

        wb = Workbook()
        existing_titles = set()
        header_fill = PatternFill('solid', fgColor='E2E8F0')
        category_fill = PatternFill('solid', fgColor='DCFCE7')
        stripe_fill = PatternFill('solid', fgColor='F8FAFC')
        header_font = Font(bold=True, color='1F2937')
        category_font = Font(bold=True, color='14532D')
        align = Alignment(vertical='center')
        thin = Side(border_style='thin', color='E5E7EB')
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        headers = ["Ingredient", "Unit", "Vendor Code", "G-Code", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        column_widths = [36, 10, 16, 16, 10, 10, 10, 10, 10, 10]

        first_sheet = True
        for vendor, categories in sorted(grouped.items(), key=lambda item: item[0].lower()):
            ws = wb.active if first_sheet else wb.create_sheet()
            first_sheet = False
            ws.title = sanitize_sheet_title(vendor, existing_titles)

            ws.append(headers)
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = column_widths[col_idx - 1]

            ws.freeze_panes = "A2"

            row_idx = 2
            for category, items_list in sorted(categories.items(), key=lambda item: item[0].lower()):
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
                cell = ws.cell(row=row_idx, column=1, value=category)
                cell.fill = category_fill
                cell.font = category_font
                cell.alignment = align
                cell.border = border
                row_idx += 1

                zebra = False
                for ing in sorted(items_list, key=lambda item: (item.get('name') or '').lower()):
                    row_values = [
                        ing.get('name') or '',
                        ing.get('unit') or ing.get('display_unit') or '',
                        ing.get('vendor_code') or '',
                        ing.get('g_code') or '',
                        '', '', '', '', '', ''
                    ]
                    for col_idx, value in enumerate(row_values, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = align
                        cell.border = border
                        if zebra:
                            cell.fill = stripe_fill
                    zebra = not zebra
                    row_idx += 1

                row_idx += 1


        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{make_safe_filename(rollout.get('name') or 'order_guide')}_order_guide.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )


@bp.route('/menu-rollouts/<rollout_id>/ops-workbook')
@login_required
def menu_rollout_ops_export(rollout_id):
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("""
            SELECT mri.recipe_id,
                   mri.batches,
                   mri.menu_price,
                   mri.target_food_cost_percent,
                   mri.menu_section,
                   mri.menu_descriptor,
                   r.name
            FROM menu_rollout_items mri
            JOIN recipes r ON r.id = mri.recipe_id
            WHERE mri.rollout_id = %s
            ORDER BY r.name
        """, (rollout_id,))
        items = cur.fetchall()
        if not items:
            flash('Add recipes to this rollout before exporting an ops workbook.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        recipe_ids = [row['recipe_id'] for row in items]
        batch_values = [row['batches'] for row in items]
        menu_prices = [row.get('menu_price') for row in items]
        target_values = [row.get('target_food_cost_percent') for row in items]
        section_values = [row.get('menu_section') for row in items]
        descriptor_values = [row.get('menu_descriptor') for row in items]
        default_target = rollout.get('target_food_cost_percent') or 20

        menu_items, _, errors = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target,
            section_values,
            descriptor_values
        )
        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        apply_menu_pricing(menu_items, default_target)
        ops_dataset = build_rollout_ops_dataset(cur, unit_system, menu_items)
        if not ops_dataset.get('station_groups') and not ops_dataset.get('menu_cards'):
            flash('No rollout operations data was generated.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

    wb = Workbook()
    header_fill = PatternFill('solid', fgColor='E2E8F0')
    header_font = Font(bold=True, color='1F2937')
    section_fill = PatternFill('solid', fgColor='FFF7ED')
    section_font = Font(bold=True, color='9A3412')
    note_fill = PatternFill('solid', fgColor='F8FAFC')
    thin = Side(border_style='thin', color='E5E7EB')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    align = Alignment(vertical='top', wrap_text=True)

    def style_headers(sheet, row=1):
        for cell in sheet[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = align

    def style_sheet(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = align
        sheet.freeze_panes = 'A2'

    ws_station = wb.active
    ws_station.title = '01 Station Map'
    ws_station.append([
        'Station',
        'Section',
        'Menu Item',
        'Descriptor',
        'Equipment',
        'Direct RB Pulls',
        'Direct Ingredient Pulls',
        'Menu Price',
        'Food Cost %',
    ])
    style_headers(ws_station)
    for group in ops_dataset.get('station_groups', []):
        for card in group.get('menu_cards', []):
            subrecipe_label = ', '.join(
                f"{row.get('recipe_name')} ({row.get('display_required_qty')} {row.get('display_required_unit')})"
                for row in card.get('direct_subrecipes', [])
            )
            ingredient_label = ', '.join(
                f"{row.get('ingredient_name')} ({row.get('display_total_qty')} {row.get('display_total_unit')})"
                for row in card.get('direct_ingredients', [])
            )
            ws_station.append([
                group.get('station'),
                card.get('menu_section'),
                card.get('menu_name'),
                card.get('menu_descriptor'),
                card.get('equipment'),
                subrecipe_label,
                ingredient_label,
                round(to_float(card.get('price')), 2) if card.get('price') is not None else None,
                round(to_float(card.get('food_cost_percent')), 1) if card.get('food_cost_percent') is not None else None,
            ])
    for width, col in [(18, 'A'), (18, 'B'), (28, 'C'), (32, 'D'), (18, 'E'), (36, 'F'), (36, 'G'), (10, 'H'), (12, 'I')]:
        ws_station.column_dimensions[col].width = width
    style_sheet(ws_station)

    ws_pulls = wb.create_sheet('02 Station Pulls')
    ws_pulls.append([
        'Station',
        'Pull Type',
        'Item',
        'Qty',
        'Unit',
        'Approx Batches',
        'Vendor',
        'Vendor SKU',
        'G-Code',
        'Source Menu Items',
    ])
    style_headers(ws_pulls)
    for group in ops_dataset.get('station_groups', []):
        ws_pulls.append([group.get('station'), 'Summary', '', '', '', '', '', '', '', ''])
        for cell in ws_pulls[ws_pulls.max_row]:
            cell.fill = section_fill
            cell.font = section_font
        for row in group.get('subrecipe_rows', []):
            ws_pulls.append([
                group.get('station'),
                'RB',
                row.get('recipe_name'),
                row.get('display_required_qty'),
                row.get('display_required_unit'),
                row.get('display_required_batches'),
                '',
                '',
                '',
                ', '.join(row.get('source_menu_items') or []),
            ])
        for row in group.get('ingredient_rows', []):
            ws_pulls.append([
                group.get('station'),
                'Ingredient',
                row.get('ingredient_name'),
                row.get('display_total_qty'),
                row.get('display_total_unit'),
                '',
                row.get('vendor'),
                row.get('vendor_code'),
                row.get('g_code'),
                ', '.join(row.get('source_menu_items') or []),
            ])
        ws_pulls.append([''] * 10)
    for width, col in [(18, 'A'), (12, 'B'), (28, 'C'), (10, 'D'), (10, 'E'), (14, 'F'), (20, 'G'), (16, 'H'), (14, 'I'), (32, 'J')]:
        ws_pulls.column_dimensions[col].width = width
    style_sheet(ws_pulls)

    ws_batches = wb.create_sheet('03 Batch Builds')
    ws_batches.append([
        'Batch Recipe',
        'Used In Menu Items',
        'Required Qty',
        'Required Unit',
        'Approx Batches',
        'Yield',
        'Station',
        'Equipment',
        'Row Type',
        'Component',
        'Qty',
        'Unit',
        'Notes',
    ])
    style_headers(ws_batches)
    for card in ops_dataset.get('batch_cards', []):
        ws_batches.append([
            card.get('recipe_name'),
            ', '.join(card.get('source_menu_items') or []),
            card.get('display_required_qty'),
            card.get('display_required_unit'),
            card.get('display_required_batches'),
            f"{card.get('yield_qty') or ''} {card.get('yield_unit') or ''}".strip(),
            card.get('station'),
            card.get('equipment'),
            'Batch',
            card.get('recipe_name'),
            '',
            '',
            '',
        ])
        for row in card.get('flat_components', []):
            ws_batches.append([
                card.get('recipe_name'),
                '',
                '',
                '',
                '',
                '',
                '',
                '',
                row.get('notes') or row.get('name'),
                row.get('name'),
                row.get('qty'),
                row.get('unit'),
                row.get('notes'),
            ])
        ws_batches.append([''] * 13)
    for width, col in [(26, 'A'), (28, 'B'), (10, 'C'), (10, 'D'), (14, 'E'), (14, 'F'), (18, 'G'), (18, 'H'), (14, 'I'), (30, 'J'), (10, 'K'), (10, 'L'), (28, 'M')]:
        ws_batches.column_dimensions[col].width = width
    style_sheet(ws_batches)

    ws_menu = wb.create_sheet('04 Menu Builds')
    ws_menu.append([
        'Station',
        'Section',
        'Menu Item',
        'Descriptor',
        'Equipment',
        'Row Type',
        'Component',
        'Qty',
        'Unit',
        'Notes',
    ])
    style_headers(ws_menu)
    for card in ops_dataset.get('menu_cards', []):
        ws_menu.append([
            card.get('station'),
            card.get('menu_section'),
            card.get('menu_name'),
            card.get('menu_descriptor'),
            card.get('equipment'),
            'Menu Item',
            card.get('menu_name'),
            '',
            '',
            '',
        ])
        for row in card.get('flat_components', []):
            ws_menu.append([
                card.get('station'),
                card.get('menu_section'),
                card.get('menu_name'),
                '',
                '',
                row.get('notes') or 'Component',
                row.get('name'),
                row.get('qty'),
                row.get('unit'),
                row.get('notes'),
            ])
        ws_menu.append([''] * 10)
    for width, col in [(18, 'A'), (18, 'B'), (28, 'C'), (30, 'D'), (18, 'E'), (14, 'F'), (34, 'G'), (10, 'H'), (10, 'I'), (28, 'J')]:
        ws_menu.column_dimensions[col].width = width
    style_sheet(ws_menu)

    ws_order = wb.create_sheet('05 Order Guide')
    ws_order.append([
        'Vendor',
        'Category',
        'Ingredient',
        'Unit',
        'Vendor Code',
        'G-Code',
        'Used In Stations',
        'Used In Menu Items',
        'Quantity Basis',
    ])
    style_headers(ws_order)
    for ingredient in ops_dataset.get('ingredient_master', []):
        ws_order.append([
            ingredient.get('vendor') or 'Unassigned Vendor',
            ingredient.get('category') or '',
            ingredient.get('name'),
            ingredient.get('unit') or ingredient.get('display_unit') or '',
            ingredient.get('vendor_code') or '',
            ingredient.get('g_code') or '',
            ', '.join(ingredient.get('used_in_stations') or []),
            ', '.join(ingredient.get('used_in_menu_items') or []),
            ops_dataset.get('quantity_basis_label'),
        ])
    for width, col in [(20, 'A'), (18, 'B'), (28, 'C'), (10, 'D'), (16, 'E'), (14, 'F'), (24, 'G'), (36, 'H'), (22, 'I')]:
        ws_order.column_dimensions[col].width = width
    style_sheet(ws_order)

    ws_notes = wb.create_sheet('06 Notes')
    ws_notes.append(['Rollout', rollout.get('name') or 'Menu Rollout'])
    ws_notes.append(['Venue', rollout.get('venue') or ''])
    ws_notes.append(['Quarter', f"{rollout.get('quarter') or ''} {rollout.get('year') or ''}".strip()])
    ws_notes.append(['Quantity Basis', ops_dataset.get('quantity_basis_label')])
    ws_notes.append([
        'Operator Note',
        'Workbook quantities are driven by the saved rollout setup. If every menu line is saved at 1, this workbook reflects sold-as build quantities rather than forecasted production volumes.',
    ])
    for row in ws_notes.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = align
    for row_idx in range(1, 6):
        ws_notes.cell(row=row_idx, column=1).fill = note_fill
        ws_notes.cell(row=row_idx, column=1).font = header_font
    ws_notes.column_dimensions['A'].width = 18
    ws_notes.column_dimensions['B'].width = 110

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{make_safe_filename(rollout.get('name') or 'menu_rollout')}_ops_workbook.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/menu-rollouts/<rollout_id>/ops-print')
@login_required
def menu_rollout_ops_print(rollout_id):
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("""
            SELECT mri.recipe_id,
                   mri.batches,
                   mri.menu_price,
                   mri.target_food_cost_percent,
                   mri.menu_section,
                   mri.menu_descriptor,
                   r.name
            FROM menu_rollout_items mri
            JOIN recipes r ON r.id = mri.recipe_id
            WHERE mri.rollout_id = %s
            ORDER BY r.name
        """, (rollout_id,))
        items = cur.fetchall()
        if not items:
            flash('Add recipes to this rollout before printing an ops packet.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        recipe_ids = [row['recipe_id'] for row in items]
        batch_values = [row['batches'] for row in items]
        menu_prices = [row.get('menu_price') for row in items]
        target_values = [row.get('target_food_cost_percent') for row in items]
        section_values = [row.get('menu_section') for row in items]
        descriptor_values = [row.get('menu_descriptor') for row in items]
        default_target = rollout.get('target_food_cost_percent') or 20

        menu_items, _, errors = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target,
            section_values,
            descriptor_values
        )
        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        apply_menu_pricing(menu_items, default_target)
        ops_dataset = build_rollout_ops_dataset(cur, unit_system, menu_items)

    return render_template(
        'menu_rollout_ops_print.html',
        rollout=rollout,
        ops_dataset=ops_dataset,
        generated_at=datetime.now(datetime.UTC),
    )

@bp.route('/menu-rollouts/<rollout_id>/pricing-export')
@login_required
def menu_rollout_pricing_export(rollout_id):
    conn = get_db()
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("""
            SELECT mri.recipe_id,
                   mri.batches,
                   mri.menu_price,
                   mri.target_food_cost_percent,
                   mri.menu_section,
                   mri.menu_descriptor,
                   r.name
            FROM menu_rollout_items mri
            JOIN recipes r ON r.id = mri.recipe_id
            WHERE mri.rollout_id = %s
            ORDER BY r.name
        """, (rollout_id,))
        items = cur.fetchall()
        if not items:
            flash('Add recipes to this rollout before exporting.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        recipe_ids = [row['recipe_id'] for row in items]
        batch_values = [row['batches'] for row in items]
        menu_prices = [row.get('menu_price') for row in items]
        target_values = [row.get('target_food_cost_percent') for row in items]
        section_values = [row.get('menu_section') for row in items]
        descriptor_values = [row.get('menu_descriptor') for row in items]
        default_target = rollout.get('target_food_cost_percent') or 20

        menu_items, _, errors = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target,
            section_values,
            descriptor_values
        )
        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        apply_menu_pricing(menu_items, default_target)
        grouped = group_menu_items(menu_items)

        wb = Workbook()
        ws = wb.active
        ws.title = "Menu Pricing"

        header_fill = PatternFill('solid', fgColor='E2E8F0')
        header_font = Font(bold=True, color='1F2937')
        section_fill = PatternFill('solid', fgColor='FFF7ED')
        section_font = Font(bold=True, color='9A3412')
        align = Alignment(vertical='center')
        thin = Side(border_style='thin', color='E5E7EB')
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        headers = [
            "Section",
            "Recipe",
            "Menu Descriptor",
            "Batches",
            "Cost / Batch",
            "Target FC%",
            "Target Food Cost Price",
            "Menu Price",
            "Food Cost %"
        ]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
            cell.border = border

        row_idx = 2
        for group in grouped:
            section = group['section']
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            section_cell = ws.cell(row=row_idx, column=1, value=section)
            section_cell.fill = section_fill
            section_cell.font = section_font
            section_cell.alignment = align
            section_cell.border = border
            row_idx += 1

            for item in group['items']:
                ws.cell(row=row_idx, column=1, value=section).border = border
                ws.cell(row=row_idx, column=2, value=item['recipe']['name']).border = border
                ws.cell(row=row_idx, column=3, value=item.get('menu_descriptor') or '').border = border
                ws.cell(row=row_idx, column=4, value=item['batches']).border = border
                ws.cell(row=row_idx, column=5, value=round(to_float(item['base_cost']), 2)).border = border
                ws.cell(row=row_idx, column=6, value=round(to_float(item.get('target_food_cost_percent')), 1)).border = border
                ws.cell(row=row_idx, column=7, value=round(to_float(item.get('suggested_price')), 2)).border = border
                ws.cell(row=row_idx, column=8, value=item.get('menu_price')).border = border
                fc_val = item.get('food_cost_percent')
                ws.cell(row=row_idx, column=9, value=round(fc_val, 1) if fc_val else None).border = border
                row_idx += 1

            row_idx += 1

        column_widths = [18, 24, 34, 10, 14, 18, 12, 12, 12]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{make_safe_filename(rollout.get('name') or 'menu_rollout')}_menu_pricing.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

@bp.route('/menu-rollouts/<rollout_id>/packet')
@login_required
def menu_rollout_packet_export(rollout_id):
    conn = get_db()
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("""
            SELECT mri.recipe_id,
                   mri.batches,
                   mri.menu_price,
                   mri.target_food_cost_percent,
                   mri.menu_section,
                   mri.menu_descriptor,
                   r.name
            FROM menu_rollout_items mri
            JOIN recipes r ON r.id = mri.recipe_id
            WHERE mri.rollout_id = %s
            ORDER BY r.name
        """, (rollout_id,))
        items = cur.fetchall()
        if not items:
            flash('Add recipes to this rollout before exporting a packet.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        recipe_ids = [row['recipe_id'] for row in items]
        batch_values = [row['batches'] for row in items]
        menu_prices = [row.get('menu_price') for row in items]
        target_values = [row.get('target_food_cost_percent') for row in items]
        section_values = [row.get('menu_section') for row in items]
        descriptor_values = [row.get('menu_descriptor') for row in items]
        default_target = rollout.get('target_food_cost_percent') or 20

        menu_items, _, errors = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target,
            section_values,
            descriptor_values
        )
        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        apply_menu_pricing(menu_items, default_target)
        menu_groups = group_menu_items(menu_items)

        cur.execute("""
            SELECT id, name, category, unit, cost_per_unit, vendor, vendor_code, g_code
            FROM ingredients
        """)
        ingredient_map = {row['id']: row for row in cur.fetchall()}

        ingredient_usage = {}
        rm_component_sets = []
        subrecipe_ids = set()
        for item in menu_items:
            components, total_cost, _ = build_component_tree(cur, item['recipe_id'], item.get('batches') or 1, 0, set(), unit_system)
            recipe_name = item['recipe']['name']
            menu_descriptor = item.get('menu_descriptor') or ''
            usage_label = recipe_name
            rm_component_sets.append({
                'menu_name': recipe_name,
                'menu_descriptor': menu_descriptor,
                'recipe_name': recipe_name,
                'components': components,
                'total_cost': total_cost
            })
            collect_subrecipes_from_components(components, subrecipe_ids)
            collect_ingredient_usage_from_components(
                components,
                usage_label,
                ingredient_usage
            )

        rb_recipes = []
        if subrecipe_ids:
            cur.execute("""
                SELECT id, name, category, yield_qty, yield_unit, recipe_type
                FROM recipes
                WHERE id = ANY(%s)
                ORDER BY name
            """, (list(subrecipe_ids),))
            for recipe in cur.fetchall():
                if normalize_recipe_type(recipe.get('recipe_type')) != 'batch':
                    continue
                components, total_cost, _ = build_component_tree(cur, recipe['id'], 1, 0, set(), unit_system)
                yield_qty = to_float(recipe.get('yield_qty'))
                yield_pricing = summarize_yield_pricing(
                    total_cost,
                    recipe.get('yield_qty'),
                    recipe.get('yield_unit'),
                    unit_system
                )
                rb_recipes.append({
                    'recipe': recipe,
                    'components': components,
                    'total_cost': total_cost,
                    'yield_qty': yield_qty,
                    'yield_unit': recipe.get('yield_unit') or '',
                    'display_yield_qty': yield_pricing['display_yield_qty'],
                    'display_yield_qty_value': yield_pricing['display_yield_qty_value'],
                    'display_yield_unit': yield_pricing['display_yield_unit'],
                    'cost_per_yield': yield_pricing['cost_per_yield'],
                    'cost_per_yield_unit': yield_pricing['cost_per_yield_unit']
                })

        ingredient_master, _, _ = build_rollout_breakdown(cur, unit_system, menu_items)

        wb = Workbook()
        header_fill = PatternFill('solid', fgColor='E2E8F0')
        header_font = Font(bold=True, color='1F2937')
        section_fill = PatternFill('solid', fgColor='FFF7ED')
        section_font = Font(bold=True, color='9A3412')
        thin = Side(border_style='thin', color='E5E7EB')
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        def style_headers(sheet, row=1):
            for cell in sheet[row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

        # 1) Menu lines
        ws_menu = wb.active
        ws_menu.title = "01 Menu Lines"
        ws_menu.append([
            "Section", "Recipe", "Menu Descriptor", "Cost/Serving", "Target FC%", "Target Food Cost Price", "Menu Price", "Food Cost %"
        ])
        style_headers(ws_menu)
        for group in menu_groups:
            ws_menu.append([group['section'], '', '', '', '', '', '', ''])
            for cell in ws_menu[ws_menu.max_row]:
                cell.fill = section_fill
                cell.font = section_font
            for item in group['items']:
                ws_menu.append([
                    group['section'],
                    item['recipe']['name'],
                    item.get('menu_descriptor') or '',
                    round(to_float(item.get('base_cost')), 4),
                    round(to_float(item.get('target_food_cost_percent')), 1),
                    round(to_float(item.get('suggested_price')), 2),
                    item.get('menu_price'),
                    round(to_float(item.get('food_cost_percent')), 1) if item.get('food_cost_percent') else None
                ])
        for width, col in [(20, 'A'), (24, 'B'), (36, 'C'), (14, 'D'), (12, 'E'), (18, 'F'), (12, 'G'), (12, 'H')]:
            ws_menu.column_dimensions[col].width = width

        # 2) RM builds
        ws_rm = wb.create_sheet("02 RM Builds")
        ws_rm.append([
            "Menu Name", "Menu Descriptor", "Type", "Component", "Qty", "Unit", "Ext Cost", "G-Code", "Vendor", "Vendor SKU", "Notes"
        ])
        style_headers(ws_rm)
        for rm in rm_component_sets:
            ws_rm.append([rm['menu_name'], rm.get('menu_descriptor') or '', "RM", rm['recipe_name'], '', '', round(to_float(rm['total_cost']), 4), '', '', '', 'Total plated cost'])
            for row in flatten_components_for_packet(rm['components'], ingredient_map, explode_subrecipes=False, preserve_base_units=True):
                ws_rm.append([
                    rm['menu_name'],
                    rm.get('menu_descriptor') or '',
                    row.get('type'),
                    row.get('name'),
                    row.get('quantity'),
                    row.get('unit'),
                    round(to_float(row.get('ext_cost')), 4),
                    row.get('g_code'),
                    row.get('vendor'),
                    row.get('vendor_code'),
                    row.get('notes')
                ])
            ws_rm.append([''] * 11)
        for width, col in [(28, 'A'), (24, 'B'), (14, 'C'), (40, 'D'), (10, 'E'), (10, 'F'), (12, 'G'), (14, 'H'), (18, 'I'), (16, 'J'), (30, 'K')]:
            ws_rm.column_dimensions[col].width = width

        # 3) RB batches
        ws_rb = wb.create_sheet("03 RB Batches")
        ws_rb.append([
            "RB Recipe", "Yield Qty", "Yield Unit", "Cost/Batch", "Cost/Yield Unit", "Type", "Component", "Qty", "Unit", "Ext Cost", "G-Code", "Vendor", "Vendor SKU"
        ])
        style_headers(ws_rb)
        for rb in rb_recipes:
            recipe = rb['recipe']
            ws_rb.append([
                recipe['name'],
                rb.get('display_yield_qty_value') or '',
                rb.get('display_yield_unit') or '',
                round(to_float(rb['total_cost']), 4),
                round(to_float(rb['cost_per_yield']), 4) if rb.get('cost_per_yield') else None,
                "RB",
                recipe['name'],
                '',
                '',
                '',
                '',
                '',
                ''
            ])
            for row in flatten_components_for_packet(rb['components'], ingredient_map, preserve_base_units=True):
                ws_rb.append([
                    recipe['name'],
                    '',
                    '',
                    '',
                    '',
                    row.get('type'),
                    row.get('name'),
                    row.get('quantity'),
                    row.get('unit'),
                    round(to_float(row.get('ext_cost')), 4),
                    row.get('g_code'),
                    row.get('vendor'),
                    row.get('vendor_code')
                ])
            ws_rb.append([''] * 13)
        for width, col in [(28, 'A'), (10, 'B'), (12, 'C'), (12, 'D'), (14, 'E'), (12, 'F'), (40, 'G'), (10, 'H'), (10, 'I'), (12, 'J'), (14, 'K'), (18, 'L'), (16, 'M')]:
            ws_rb.column_dimensions[col].width = width

        # 4) Ingredient crosswalk
        ws_ing = wb.create_sheet("04 Ingredient Xwalk")
        ws_ing.append(["Ingredient", "Category", "Unit", "Cost/Unit", "G-Code", "Vendor", "Vendor SKU", "Used In RM Items"])
        style_headers(ws_ing)
        for ing in ingredient_master:
            used_in = sorted(ingredient_usage.get(ing['id'], set()))
            ws_ing.append([
                ing.get('name'),
                ing.get('category'),
                ing.get('unit'),
                round(to_float(ing.get('cost_per_unit')), 5) if ing.get('cost_per_unit') is not None else None,
                ing.get('g_code'),
                ing.get('vendor'),
                ing.get('vendor_code'),
                ', '.join(used_in)
            ])
        for width, col in [(32, 'A'), (18, 'B'), (10, 'C'), (12, 'D'), (14, 'E'), (18, 'F'), (16, 'G'), (48, 'H')]:
            ws_ing.column_dimensions[col].width = width

        # 5) Entry order guide
        ws_steps = wb.create_sheet("05 Entry Order")
        ws_steps.append(["Step", "Action"])
        style_headers(ws_steps)
        ws_steps.append([1, "Update ingredient records in Acumatica first (cost/uom/vendor/G-code)."])
        ws_steps.append([2, f"Enter RB batch recipes ({len(rb_recipes)} total), including yields and components."])
        ws_steps.append([3, f"Enter RM plated recipes ({len(menu_items)} total), linking RB components where used."])
        ws_steps.append([4, "Apply menu pricing and verify target food cost % against target food cost price."])
        ws_steps.append([5, "Final review: run where-used checks for every ingredient with G-code mapping."])
        ws_steps.column_dimensions['A'].width = 8
        ws_steps.column_dimensions['B'].width = 110

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"{make_safe_filename(rollout.get('name') or 'menu_rollout')}_acumatica_packet.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

@bp.route('/menu-rollouts/<rollout_id>/packet-print')
@login_required
def menu_rollout_packet_print(rollout_id):
    conn = get_db()
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("""
            SELECT mri.recipe_id,
                   mri.batches,
                   mri.menu_price,
                   mri.target_food_cost_percent,
                   mri.menu_section,
                   mri.menu_descriptor,
                   r.name
            FROM menu_rollout_items mri
            JOIN recipes r ON r.id = mri.recipe_id
            WHERE mri.rollout_id = %s
            ORDER BY r.name
        """, (rollout_id,))
        items = cur.fetchall()
        if not items:
            flash('Add recipes to this rollout before printing a packet.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        recipe_ids = [row['recipe_id'] for row in items]
        batch_values = [row['batches'] for row in items]
        menu_prices = [row.get('menu_price') for row in items]
        target_values = [row.get('target_food_cost_percent') for row in items]
        section_values = [row.get('menu_section') for row in items]
        descriptor_values = [row.get('menu_descriptor') for row in items]
        default_target = rollout.get('target_food_cost_percent') or 20

        menu_items, _, errors = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target,
            section_values,
            descriptor_values
        )
        if errors:
            flash(' '.join(sorted(set(errors))), 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        apply_menu_pricing(menu_items, default_target)
        menu_groups = group_menu_items(menu_items)

        cur.execute("""
            SELECT id, name, category, unit, cost_per_unit, vendor, vendor_code, g_code
            FROM ingredients
        """)
        ingredient_map = {row['id']: row for row in cur.fetchall()}

        ingredient_usage = {}
        rm_component_sets = []
        subrecipe_ids = set()
        for item in menu_items:
            components, total_cost, _ = build_component_tree(cur, item['recipe_id'], item.get('batches') or 1, 0, set(), unit_system)
            recipe_name = item['recipe']['name']
            menu_descriptor = item.get('menu_descriptor') or ''
            usage_label = recipe_name
            rm_component_sets.append({
                'menu_name': recipe_name,
                'menu_descriptor': menu_descriptor,
                'recipe_name': recipe_name,
                'total_cost': total_cost,
                'rows': flatten_components_for_packet(components, ingredient_map, explode_subrecipes=False, preserve_base_units=True)
            })
            collect_subrecipes_from_components(components, subrecipe_ids)
            collect_ingredient_usage_from_components(components, usage_label, ingredient_usage)

        rb_recipes = []
        if subrecipe_ids:
            cur.execute("""
                SELECT id, name, category, yield_qty, yield_unit, recipe_type
                FROM recipes
                WHERE id = ANY(%s)
                ORDER BY name
            """, (list(subrecipe_ids),))
            for recipe in cur.fetchall():
                if normalize_recipe_type(recipe.get('recipe_type')) != 'batch':
                    continue
                components, total_cost, _ = build_component_tree(cur, recipe['id'], 1, 0, set(), unit_system)
                yield_pricing = summarize_yield_pricing(
                    total_cost,
                    recipe.get('yield_qty'),
                    recipe.get('yield_unit'),
                    unit_system
                )
                rb_recipes.append({
                    'recipe': recipe,
                    'total_cost': total_cost,
                    'display_yield_qty': yield_pricing['display_yield_qty'],
                    'display_yield_unit': yield_pricing['display_yield_unit'],
                    'cost_per_yield': yield_pricing['cost_per_yield'],
                    'cost_per_yield_unit': yield_pricing['cost_per_yield_unit'],
                    'rows': flatten_components_for_packet(components, ingredient_map, preserve_base_units=True)
                })

        ingredient_master, _, _ = build_rollout_breakdown(cur, unit_system, menu_items)
        for ing in ingredient_master:
            ing['used_in'] = ', '.join(sorted(ingredient_usage.get(ing['id'], set())))


        return render_template(
            'menu_rollout_packet_print.html',
            rollout=rollout,
            menu_groups=menu_groups,
            rm_component_sets=rm_component_sets,
            rb_recipes=rb_recipes,
            ingredient_master=ingredient_master,
            generated_at=datetime.now(datetime.UTC),
            default_target_percent=default_target
        )

@bp.route('/menu-rollouts/<rollout_id>/print')
@login_required
def menu_rollout_print(rollout_id):
    conn = get_db()
    with get_cursor() as cur:
        unit_system = get_unit_system()

        cur.execute("SELECT * FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        cur.execute("""
            SELECT mri.recipe_id,
                   mri.batches,
                   mri.menu_price,
                   mri.target_food_cost_percent,
                   mri.menu_section,
                   mri.menu_descriptor,
                   r.name
            FROM menu_rollout_items mri
            JOIN recipes r ON r.id = mri.recipe_id
            WHERE mri.rollout_id = %s
            ORDER BY r.name
        """, (rollout_id,))
        items = cur.fetchall()
        if not items:
            flash('Add recipes to this rollout before printing.', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))

        recipe_ids = [row['recipe_id'] for row in items]
        batch_values = [row['batches'] for row in items]
        menu_prices = [row.get('menu_price') for row in items]
        target_values = [row.get('target_food_cost_percent') for row in items]
        section_values = [row.get('menu_section') for row in items]
        descriptor_values = [row.get('menu_descriptor') for row in items]
        default_target = rollout.get('target_food_cost_percent') or 20

        menu_items, _, _ = parse_menu_items(
            cur,
            unit_system,
            recipe_ids,
            batch_values,
            menu_prices,
            target_values,
            None,
            default_target,
            section_values,
            descriptor_values
        )
        apply_menu_pricing(menu_items, default_target)
        menu_groups = group_menu_items(menu_items)


        return render_template(
            'menu_rollout_print.html',
            rollout=rollout,
            menu_groups=menu_groups,
            default_target_percent=default_target,
            generated_at=datetime.now(datetime.UTC)
        )

@bp.route('/menu-rollouts/<rollout_id>/delete', methods=['POST'])
@login_required
def menu_rollout_delete(rollout_id):
    conn = get_db()
    with get_cursor() as cur:

        cur.execute("SELECT id, name FROM menu_rollouts WHERE id = %s", (rollout_id,))
        rollout = cur.fetchone()
        if not rollout:
            flash('Menu rollout not found', 'error')
            return redirect(url_for('menu_rollouts'))

        try:
            cur.execute("DELETE FROM menu_rollout_items WHERE rollout_id = %s", (rollout_id,))
            cur.execute("DELETE FROM menu_rollouts WHERE id = %s", (rollout_id,))
            conn.commit()
            flash('Menu rollout deleted', 'success')
            return redirect(url_for('menu_rollouts'))
        except Exception:
            conn.rollback()
            flash('Error deleting menu rollout', 'error')
            return redirect(url_for('menu_rollout_edit', rollout_id=rollout_id))
