import os
import shutil
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

from db import get_cursor, get_db
from flask import Blueprint, Response, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from xml.sax.saxutils import escape
from helpers.banquet import (
    BANQUET_ACTIVE_STATUSES,
    BANQUET_BEO_UPLOAD_ROOT,
    BANQUET_STATUS_CHOICES,
    BANQUET_VENUE_ID,
    auto_complete_past_banquet_events,
    banquet_tables_ready,
    build_banquet_datasets,
    build_banquet_prep_groups,
    clone_banquet_menu_item,
    ensure_banquet_event_beo_files_table,
    ensure_banquet_menu_item_event_only_column,
    ensure_banquet_shopping_checks_table,
    event_beo_file_path,
    get_banquet_date_window,
    get_banquet_event,
    get_banquet_event_beo_file,
    get_banquet_event_lines,
    get_banquet_menu_item,
    get_banquet_menu_item_component_summaries,
    get_banquet_menu_item_ingredients,
    get_banquet_menu_item_recipe_components,
    list_banquet_event_beo_files,
    list_banquet_event_beo_files_by_events,
    list_banquet_menu_items,
    load_banquet_shopping_checklist,
    parse_event_recipe_lines,
    parse_menu_item_ingredient_lines,
    parse_menu_item_recipe_lines,
    resolve_banquet_venue,
    resolve_or_create_banquet_menu_item,
    store_banquet_event_beo_file,
    upsert_menu_item_recipe_link,
)
from helpers.db_helpers import db_column_exists, ensure_banquet_menu_item_base_yield_columns
from helpers.formatting import make_safe_filename, split_instruction_steps
from helpers.menu import auto_match_menu_recipe_id, clean_menu_text
from helpers.recipes import menu_line_base_multiplier
from helpers.shared import (
    BANQUET_MENU_SECTION_OPTIONS,
    generate_id,
    handle_route_error,
    normalize_return_path,
    parse_catering_pdf_to_template_items,
    parse_float_field,
    to_float,
)
from helpers.units import format_number, get_unit_system, normalize_count_unit
from helpers.venues import get_active_venues, get_component_recipes_for_venue, get_plated_recipes_for_venue

bp = Blueprint('banquet', __name__)

@bp.errorhandler(Exception)
def handle_banquet_error(error):
    return handle_route_error(error, 'banquet')

def banquet_event_line_choice_column_ready(cur):
    return db_column_exists(cur, 'public.banquet_event_menu_items', 'choice_selections')

def insert_banquet_event_menu_line(
    cur,
    event_id,
    menu_item_id,
    menu_item_name,
    recipe_id,
    quantity,
    quantity_unit,
    menu_section,
    menu_descriptor,
    notes,
    sort_order,
    choice_selections=None,
    has_choice_selections=False
):
    if has_choice_selections:
        cur.execute("""
            INSERT INTO banquet_event_menu_items (
                event_id, menu_item_id, menu_item_name, recipe_id, quantity, quantity_unit,
                menu_section, menu_descriptor, notes, choice_selections, sort_order
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            event_id,
            menu_item_id,
            menu_item_name,
            recipe_id,
            quantity,
            quantity_unit,
            menu_section,
            menu_descriptor,
            notes,
            choice_selections or None,
            sort_order
        ))
        return

    cur.execute("""
        INSERT INTO banquet_event_menu_items (
            event_id, menu_item_id, menu_item_name, recipe_id, quantity, quantity_unit,
            menu_section, menu_descriptor, notes, sort_order
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        event_id,
        menu_item_id,
        menu_item_name,
        recipe_id,
        quantity,
        quantity_unit,
        menu_section,
        menu_descriptor,
        notes,
        sort_order
    ))

def get_banquet_event_recipe_options(cur, venue_id):
    recipe_options = get_component_recipes_for_venue(cur, venue_id)
    seen_labels = set()
    recipe_label_map = {}
    for recipe in recipe_options:
        recipe_type = (recipe.get('recipe_type') or '').upper()
        if recipe.get('option_source') == 'menu_item' and recipe.get('menu_item_name'):
            base_label = f"{recipe.get('menu_item_name')} [Menu Item] -> {recipe.get('name')}{f' ({recipe_type})' if recipe_type else ''}"
        else:
            base_label = f"{recipe.get('name')}{f' ({recipe_type})' if recipe_type else ''}"
        candidate = base_label
        suffix = 2
        while candidate.lower() in seen_labels:
            candidate = f"{base_label} [{suffix}]"
            suffix += 1
        seen_labels.add(candidate.lower())
        recipe['display_label'] = candidate
        recipe_id = recipe.get('id')
        if recipe_id and (recipe_id not in recipe_label_map or recipe.get('option_source') == 'recipe'):
            recipe_label_map[recipe_id] = candidate
    return recipe_options, recipe_label_map

@bp.route('/banquet-planner')
@login_required
def banquet_planner():
    conn = get_db()
    with get_cursor() as cur:
        today = date.today()
        ten_day_end = today + timedelta(days=9)

        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        selected_venue = banquet_venue.get('id') or ''
        banquet_venue_name = banquet_venue.get('name') or 'Banquets'

        start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return render_template(
                'banquet_planner.html',
                venues=[banquet_venue] if selected_venue else venues,
                banquet_venue_name=banquet_venue_name,
                selected_venue=selected_venue,
                start_date=start_date.isoformat(),
                end_date=end_date.isoformat(),
                upcoming_events=[],
                event_count=0,
                menu_line_count=0,
                shopping_count=0,
                weekly_prep_count=0,
                menu_item_count=0,
                banquet_metrics={
                    'total_events': 0,
                    'completed_events': 0,
                    'active_next_10_days': 0,
                    'avg_guests_90d': 0
                },
                past_events=[]
            )

        auto_complete_past_banquet_events(cur, selected_venue)
        datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())

        cur.execute("""
            SELECT e.id,
                   e.name,
                   e.event_date,
                   e.guest_count AS guests,
                   e.status,
                   COALESCE(v.name, e.building, 'Banquet') AS venue_name,
                   COUNT(emi.id) AS recipe_lines
            FROM banquet_events e
            LEFT JOIN venues v ON v.id = e.venue_id
            LEFT JOIN banquet_event_menu_items emi ON emi.event_id = e.id
            WHERE e.event_date BETWEEN %s AND %s
              AND (%s = '' OR e.venue_id = %s)
            GROUP BY e.id, v.name
            ORDER BY e.event_date, e.name
            LIMIT 60
        """, (start_date, end_date, selected_venue, selected_venue))
        upcoming_events = cur.fetchall()

        if selected_venue:
            cur.execute("""
                SELECT COUNT(*) AS count
                FROM banquet_menu_items
                WHERE venue_id = %s OR venue_id IS NULL
            """, (selected_venue,))
        else:
            cur.execute("SELECT COUNT(*) AS count FROM banquet_menu_items")
        menu_item_count = (cur.fetchone() or {}).get('count', 0)

        cur.execute("""
            SELECT
                COUNT(*) AS total_events,
                COUNT(*) FILTER (WHERE status = 'completed') AS completed_events,
                COUNT(*) FILTER (WHERE event_date BETWEEN %s AND %s AND status = ANY(%s)) AS active_next_10_days
            FROM banquet_events e
            WHERE (%s = '' OR e.venue_id = %s)
        """, (today, ten_day_end, list(BANQUET_ACTIVE_STATUSES), selected_venue, selected_venue))
        metrics_row = cur.fetchone() or {}

        cur.execute("""
            SELECT ROUND(AVG(guest_count)::numeric, 1) AS avg_guests_90d
            FROM banquet_events e
            WHERE guest_count IS NOT NULL
              AND guest_count > 0
              AND e.event_date >= %s
              AND (%s = '' OR e.venue_id = %s)
        """, (today - timedelta(days=90), selected_venue, selected_venue))
        avg_row = cur.fetchone() or {}
        avg_guests_90d = avg_row.get('avg_guests_90d')

        banquet_metrics = {
            'total_events': int(metrics_row.get('total_events') or 0),
            'completed_events': int(metrics_row.get('completed_events') or 0),
            'active_next_10_days': int(metrics_row.get('active_next_10_days') or 0),
            'avg_guests_90d': float(avg_guests_90d) if avg_guests_90d is not None else 0
        }

        cur.execute("""
            SELECT e.id,
                   e.name,
                   e.event_date,
                   e.guest_count AS guests,
                   e.status,
                   COALESCE(v.name, e.building, 'Banquet') AS venue_name,
                   COUNT(emi.id) AS recipe_lines
            FROM banquet_events e
            LEFT JOIN venues v ON v.id = e.venue_id
            LEFT JOIN banquet_event_menu_items emi ON emi.event_id = e.id
            WHERE e.event_date < %s
              AND (%s = '' OR e.venue_id = %s)
            GROUP BY e.id, v.name
            ORDER BY e.event_date DESC, e.name
            LIMIT 20
        """, (today, selected_venue, selected_venue))
        past_events = cur.fetchall()


        return render_template(
            'banquet_planner.html',
            venues=[banquet_venue] if selected_venue else venues,
            banquet_venue_name=banquet_venue_name,
            selected_venue=selected_venue,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            upcoming_events=upcoming_events,
            event_count=datasets['event_count'],
            menu_line_count=datasets['menu_line_count'],
            shopping_count=len(datasets['shopping_ingredients']),
            weekly_prep_count=len(datasets['weekly_prep']),
            menu_item_count=menu_item_count,
            banquet_metrics=banquet_metrics,
            past_events=past_events
        )

@bp.route('/banquet-planner/events/new', methods=['GET', 'POST'])
@login_required
def banquet_event_new():
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))
        had_event_only_column = db_column_exists(cur, 'public.banquet_menu_items', 'is_event_only')
        ensure_banquet_menu_item_event_only_column(cur)
        if not had_event_only_column:
            conn.commit()

        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        default_venue_id = banquet_venue.get('id') or ''
        banquet_venue_name = banquet_venue.get('name') or 'Banquets'
        has_line_choice_selections = banquet_event_line_choice_column_ready(cur)
        recipe_options, recipe_label_map = get_banquet_event_recipe_options(cur, default_venue_id)

        event = {
            'name': '',
            'event_date': '',
            'guests': '',
            'venue_id': default_venue_id,
            'building': '',
            'room': '',
            'service_timing': '',
            'dietary_notes': '',
            'notes': '',
            'status': 'planning'
        }

        if request.method == 'POST':
            errors = []
            event_name = clean_menu_text(request.form.get('name'))
            event_date_raw = (request.form.get('event_date') or '').strip()
            venue_id = default_venue_id

            status = (request.form.get('status') or 'planning').strip().lower()
            if status not in {choice[0] for choice in BANQUET_STATUS_CHOICES}:
                status = 'planning'

            guests = parse_float_field(request.form.get('guests'), 'Guests', errors, required=False, min_value=1)
            event_date = None
            try:
                event_date = datetime.strptime(event_date_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append('Event date is required.')

            lines, line_errors = parse_event_recipe_lines(request)
            errors.extend(line_errors)
            if not lines:
                errors.append('Add at least one menu line for the event.')
            if not event_name:
                errors.append('Event name is required.')

            event = {
                'name': event_name,
                'event_date': event_date_raw,
                'guests': request.form.get('guests') or '',
                'venue_id': venue_id,
                'building': clean_menu_text(request.form.get('building')),
                'room': clean_menu_text(request.form.get('room')),
                'service_timing': clean_menu_text(request.form.get('service_timing')),
                'dietary_notes': clean_menu_text(request.form.get('dietary_notes')),
                'notes': clean_menu_text(request.form.get('notes')),
                'status': status
            }

            if errors:
                flash(' '.join(sorted(set(errors))), 'error')
            else:
                event_id = generate_id('bev_')
                try:
                    cur.execute("""
                        INSERT INTO banquet_events (
                            id, name, event_date, guest_count, venue_id, building, room, service_timing, dietary_notes, notes, status
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        event_id,
                        event_name,
                        event_date,
                        int(guests) if guests is not None else None,
                        venue_id or None,
                        event['building'] or None,
                        event['room'] or None,
                        event['service_timing'] or None,
                        event['dietary_notes'] or None,
                        event['notes'] or None,
                        status
                    ))

                    for idx, line in enumerate(lines):
                        menu_item_id = line.get('menu_item_id')
                        line_name = line.get('menu_item_name') or ''
                        line_section = line.get('menu_section')
                        line_descriptor = line.get('menu_descriptor')
                        recipe_id = line.get('recipe_id')

                        if menu_item_id:
                            cur.execute("""
                                SELECT id, name, menu_section, menu_descriptor
                                FROM banquet_menu_items
                                WHERE id = %s
                            """, (menu_item_id,))
                            menu_item = cur.fetchone()
                            if not menu_item:
                                continue
                            line_name = line_name or menu_item.get('name') or line_name
                            line_section = line_section or menu_item.get('menu_section')
                            line_descriptor = line_descriptor or menu_item.get('menu_descriptor')
                            if not recipe_id:
                                cur.execute("""
                                    SELECT recipe_id
                                    FROM banquet_menu_item_recipes
                                    WHERE menu_item_id = %s
                                    ORDER BY id
                                    LIMIT 1
                                """, (menu_item_id,))
                                match = cur.fetchone()
                                recipe_id = match.get('recipe_id') if match else None
                        else:
                            menu_item_id = resolve_or_create_banquet_menu_item(
                                cur,
                                line,
                                venue_id,
                                is_event_only=line.get('is_custom_item', False)
                            )
                            if not menu_item_id:
                                continue
                            upsert_menu_item_recipe_link(cur, menu_item_id, recipe_id)

                        insert_banquet_event_menu_line(
                            cur,
                            event_id,
                            menu_item_id,
                            line_name,
                            recipe_id,
                            line['quantity'],
                            line['quantity_unit'] or 'each',
                            line_section or None,
                            line_descriptor or None,
                            line['notes'] or None,
                            idx,
                            choice_selections=line.get('choice_selections'),
                            has_choice_selections=has_line_choice_selections
                        )

                    conn.commit()
                    flash('Banquet event created.', 'success')
                    return redirect(url_for('banquet_event_edit', event_id=event_id))
                except Exception:
                    conn.rollback()
                    flash('Error creating event.', 'error')

            lines = lines or []
        else:
            lines = []

        for line in lines:
            if 'notes' not in line:
                line['notes'] = line.get('line_notes')
            if 'choice_selections' not in line:
                line['choice_selections'] = line.get('line_choice_selections')
            line['is_custom_item'] = bool(line.get('is_custom_item') or line.get('is_event_only'))
            line['recipe_display_label'] = recipe_label_map.get(line.get('recipe_id'), line.get('recipe_name') or '')

        menu_items = list_banquet_menu_items(cur, event.get('venue_id') or '')
        menu_item_ids = [item.get('id') for item in menu_items if item.get('id')]
        menu_item_choice_options = {}
        for menu_item_id, summary in get_banquet_menu_item_component_summaries(cur, menu_item_ids).items():
            options = []
            for recipe in summary.get('recipes') or []:
                choice_group = clean_menu_text(recipe.get('choice_group') or '')
                recipe_id = recipe.get('id')
                if not choice_group or not recipe_id:
                    continue
                options.append({
                    'recipe_id': recipe_id,
                    'recipe_name': recipe.get('name') or 'Choice',
                    'choice_group': choice_group
                })
            if options:
                menu_item_choice_options[menu_item_id] = options
        return render_template(
            'banquet_event_form.html',
            page_title='New Banquet Event',
            event=event,
            event_beo_files=[],
            guest_count_history=[],
            lines=lines,
            menu_items=menu_items,
            recipe_options=recipe_options,
            menu_item_choice_options=menu_item_choice_options,
            venues=venues,
            banquet_venue_name=banquet_venue_name,
            section_options=BANQUET_MENU_SECTION_OPTIONS,
            status_options=BANQUET_STATUS_CHOICES
        )

@bp.route('/banquet-planner/events/<event_id>/edit', methods=['GET', 'POST'])
@login_required
def banquet_event_edit(event_id):
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))
        had_event_only_column = db_column_exists(cur, 'public.banquet_menu_items', 'is_event_only')
        ensure_banquet_menu_item_event_only_column(cur)
        if not had_event_only_column:
            conn.commit()

        event = get_banquet_event(cur, event_id)
        if not event:
            flash('Event not found.', 'error')
            return redirect(url_for('banquet_planner'))
        cur.execute("SELECT to_regclass('public.banquet_event_guest_log') AS table_ref")
        guest_log_table_ready = bool((cur.fetchone() or {}).get('table_ref'))
        has_line_choice_selections = banquet_event_line_choice_column_ready(cur)

        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        banquet_venue_id = banquet_venue.get('id') or ''
        banquet_venue_name = banquet_venue.get('name') or 'Banquets'
        recipe_options, recipe_label_map = get_banquet_event_recipe_options(cur, banquet_venue_id)
        if banquet_venue_id:
            event['venue_id'] = banquet_venue_id

        if request.method == 'GET':
            auto_complete_past_banquet_events(cur, banquet_venue_id)
            event = get_banquet_event(cur, event_id)

        if request.method == 'POST':
            form_action = (request.form.get('form_action') or 'save').strip().lower()
            if form_action == 'upload_beo':
                upload = request.files.get('beo_pdf')
                try:
                    saved_file, upload_error = store_banquet_event_beo_file(cur, event_id, upload)
                    if upload_error:
                        flash(upload_error, 'error')
                        conn.rollback()
                    else:
                        conn.commit()
                        flash(f"Attached BEO PDF: {saved_file.get('original_filename')}", 'success')
                except Exception:
                    conn.rollback()
                    flash('Error attaching BEO PDF.', 'error')
                return redirect(url_for('banquet_event_edit', event_id=event_id))

            if form_action == 'delete_beo':
                file_id = clean_menu_text(request.form.get('beo_file_id'))
                if not file_id:
                    flash('Missing BEO file reference.', 'error')
                    return redirect(url_for('banquet_event_edit', event_id=event_id))
                file_row = get_banquet_event_beo_file(cur, event_id, file_id)
                if not file_row:
                    flash('BEO file not found.', 'error')
                    return redirect(url_for('banquet_event_edit', event_id=event_id))
                file_path = event_beo_file_path(event_id, file_row.get('stored_filename') or '')
                try:
                    ensure_banquet_event_beo_files_table(cur)
                    cur.execute("""
                        DELETE FROM banquet_event_beo_files
                        WHERE id = %s
                          AND event_id = %s
                    """, (file_id, event_id))
                    conn.commit()
                    if file_path and os.path.exists(file_path):
                        os.remove(file_path)
                    flash(f"Removed BEO PDF: {file_row.get('original_filename')}", 'success')
                except Exception:
                    conn.rollback()
                    flash('Error removing BEO PDF.', 'error')
                return redirect(url_for('banquet_event_edit', event_id=event_id))

            errors = []
            event_name = clean_menu_text(request.form.get('name'))
            event_date_raw = (request.form.get('event_date') or '').strip()
            venue_id = banquet_venue_id

            status = (request.form.get('status') or 'planning').strip().lower()
            if status not in {choice[0] for choice in BANQUET_STATUS_CHOICES}:
                status = 'planning'

            guests = parse_float_field(request.form.get('guests'), 'Guests', errors, required=False, min_value=1)
            event_date = None
            try:
                event_date = datetime.strptime(event_date_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append('Event date is required.')

            lines, line_errors = parse_event_recipe_lines(request)
            errors.extend(line_errors)
            if not lines:
                errors.append('Add at least one menu line for the event.')
            if not event_name:
                errors.append('Event name is required.')

            if errors:
                flash(' '.join(sorted(set(errors))), 'error')
                event = {
                    **event,
                    'name': event_name,
                    'event_date': event_date_raw,
                    'guests': request.form.get('guests') or '',
                    'venue_id': venue_id,
                    'building': clean_menu_text(request.form.get('building')),
                    'room': clean_menu_text(request.form.get('room')),
                    'service_timing': clean_menu_text(request.form.get('service_timing')),
                    'dietary_notes': clean_menu_text(request.form.get('dietary_notes')),
                    'notes': clean_menu_text(request.form.get('notes')),
                    'status': status
                }
            else:
                try:
                    previous_guest_count = event.get('guest_count')
                    new_guest_count = int(guests) if guests is not None else None
                    cur.execute("""
                        UPDATE banquet_events
                        SET name = %s,
                            event_date = %s,
                            guest_count = %s,
                            venue_id = %s,
                            building = %s,
                            room = %s,
                            service_timing = %s,
                            dietary_notes = %s,
                            notes = %s,
                            status = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (
                        event_name,
                        event_date,
                        new_guest_count,
                        venue_id or None,
                        clean_menu_text(request.form.get('building')) or None,
                        clean_menu_text(request.form.get('room')) or None,
                        clean_menu_text(request.form.get('service_timing')) or None,
                        clean_menu_text(request.form.get('dietary_notes')) or None,
                        clean_menu_text(request.form.get('notes')) or None,
                        status,
                        event_id
                    ))

                    cur.execute("""
                        SELECT id,
                               menu_item_id,
                               recipe_id,
                               menu_item_name,
                               menu_section,
                               menu_descriptor,
                               {choice_selections_sql}
                        FROM banquet_event_menu_items
                        WHERE event_id = %s
                    """.format(
                        choice_selections_sql=("choice_selections" if has_line_choice_selections else "NULL::text AS choice_selections")
                    ), (event_id,))
                    existing_line_map = {str(row.get('id')): row for row in cur.fetchall() if row.get('id') is not None}
                    cur.execute("DELETE FROM banquet_event_menu_items WHERE event_id = %s", (event_id,))
                    for idx, line in enumerate(lines):
                        line_id = str(line.get('line_id') or '').strip()
                        menu_item_id = line.get('menu_item_id')
                        line_name = line.get('menu_item_name') or ''
                        line_section = line.get('menu_section')
                        line_descriptor = line.get('menu_descriptor')
                        recipe_id = line.get('recipe_id')
                        line_choice_selections = line.get('choice_selections')
                        if not menu_item_id and line_id and line_id in existing_line_map:
                            previous = existing_line_map[line_id]
                            menu_item_id = previous.get('menu_item_id')
                            if not recipe_id:
                                recipe_id = previous.get('recipe_id')
                            line_name = line_name or previous.get('menu_item_name') or ''
                            line_section = line_section or previous.get('menu_section')
                            line_descriptor = line_descriptor or previous.get('menu_descriptor')
                            line_choice_selections = line_choice_selections or previous.get('choice_selections')

                        if menu_item_id:
                            cur.execute("""
                                SELECT id, name, menu_section, menu_descriptor
                                FROM banquet_menu_items
                                WHERE id = %s
                            """, (menu_item_id,))
                            menu_item = cur.fetchone()
                            if not menu_item:
                                continue
                            line_name = line_name or menu_item.get('name') or line_name
                            line_section = line_section or menu_item.get('menu_section')
                            line_descriptor = line_descriptor or menu_item.get('menu_descriptor')
                            if not recipe_id:
                                cur.execute("""
                                    SELECT recipe_id
                                    FROM banquet_menu_item_recipes
                                    WHERE menu_item_id = %s
                                    ORDER BY id
                                    LIMIT 1
                                """, (menu_item_id,))
                                match = cur.fetchone()
                                recipe_id = match.get('recipe_id') if match else None
                        else:
                            menu_item_id = resolve_or_create_banquet_menu_item(
                                cur,
                                line,
                                venue_id,
                                is_event_only=line.get('is_custom_item', False)
                            )
                            if not menu_item_id:
                                continue
                            upsert_menu_item_recipe_link(cur, menu_item_id, recipe_id)

                        insert_banquet_event_menu_line(
                            cur,
                            event_id,
                            menu_item_id,
                            line_name,
                            recipe_id,
                            line['quantity'],
                            line['quantity_unit'] or 'each',
                            line_section or None,
                            line_descriptor or None,
                            line['notes'] or None,
                            idx,
                            choice_selections=line_choice_selections,
                            has_choice_selections=has_line_choice_selections
                        )
                    if guest_log_table_ready and previous_guest_count != new_guest_count:
                        cur.execute("""
                            INSERT INTO banquet_event_guest_log (event_id, old_count, new_count)
                            VALUES (%s, %s, %s)
                        """, (event_id, previous_guest_count, new_guest_count))
                    conn.commit()
                    flash('Banquet event updated.', 'success')
                    return redirect(url_for('banquet_event_edit', event_id=event_id))
                except Exception:
                    conn.rollback()
                    flash('Error updating event.', 'error')

        if request.method == 'GET':
            lines = get_banquet_event_lines(cur, event_id)
        else:
            lines = parse_event_recipe_lines(request)[0]

        for line in lines:
            if 'notes' not in line:
                line['notes'] = line.get('line_notes')
            if 'choice_selections' not in line:
                line['choice_selections'] = line.get('line_choice_selections')
            line['is_custom_item'] = bool(line.get('is_custom_item') or line.get('is_event_only'))
            line['recipe_display_label'] = recipe_label_map.get(line.get('recipe_id'), line.get('recipe_name') or '')

        menu_items = list_banquet_menu_items(cur, event.get('venue_id') or '')
        menu_item_ids = [item.get('id') for item in menu_items if item.get('id')]
        menu_item_choice_options = {}
        for menu_item_id, summary in get_banquet_menu_item_component_summaries(cur, menu_item_ids).items():
            options = []
            for recipe in summary.get('recipes') or []:
                choice_group = clean_menu_text(recipe.get('choice_group') or '')
                recipe_id = recipe.get('id')
                if not choice_group or not recipe_id:
                    continue
                options.append({
                    'recipe_id': recipe_id,
                    'recipe_name': recipe.get('name') or 'Choice',
                    'choice_group': choice_group
                })
            if options:
                menu_item_choice_options[menu_item_id] = options
        event_beo_files = list_banquet_event_beo_files(cur, event_id)
        guest_count_history = []
        if guest_log_table_ready:
            cur.execute("""
                SELECT old_count, new_count, changed_at
                FROM banquet_event_guest_log
                WHERE event_id = %s
                ORDER BY changed_at DESC
                LIMIT 20
            """, (event_id,))
            guest_count_history = cur.fetchall()
        return render_template(
            'banquet_event_form.html',
            page_title='Edit Banquet Event',
            event=event,
            event_beo_files=event_beo_files,
            guest_count_history=guest_count_history,
            lines=lines,
            menu_items=menu_items,
            recipe_options=recipe_options,
            menu_item_choice_options=menu_item_choice_options,
            venues=venues,
            banquet_venue_name=banquet_venue_name,
            section_options=BANQUET_MENU_SECTION_OPTIONS,
            status_options=BANQUET_STATUS_CHOICES
        )

@bp.route('/banquet-planner/events/<event_id>/beo/<file_id>/download')
@login_required
def banquet_event_beo_download(event_id, file_id):
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            return ("Banquet tables are missing.", 404)

        event = get_banquet_event(cur, event_id)
        if not event:
            return ("Event not found.", 404)

        file_row = get_banquet_event_beo_file(cur, event_id, file_id)
        if not file_row:
            return ("BEO file not found.", 404)

        file_path = event_beo_file_path(event_id, file_row.get('stored_filename') or '')
        if not os.path.exists(file_path):
            return ("Stored BEO file is missing.", 404)

        view_raw = (request.args.get('view') or '').strip().lower()
        as_attachment = view_raw not in ('1', 'true', 'yes', 'on')
        download_name = file_row.get('original_filename') or f"{event_id}_beo.pdf"
        return send_file(
            file_path,
            mimetype='application/pdf',
            as_attachment=as_attachment,
            download_name=download_name
        )

@bp.route('/banquet-planner/events/<event_id>/delete', methods=['POST'])
@login_required
def banquet_event_delete(event_id):
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))

        cur.execute("SELECT id, name FROM banquet_events WHERE id = %s", (event_id,))
        event = cur.fetchone()
        if not event:
            flash('Event not found.', 'error')
            return redirect(url_for('banquet_planner'))
        event_upload_folder = os.path.join(BANQUET_BEO_UPLOAD_ROOT, event_id)
        try:
            cur.execute("DELETE FROM banquet_events WHERE id = %s", (event_id,))
            conn.commit()
            shutil.rmtree(event_upload_folder, ignore_errors=True)
            flash(f"Deleted event: {event['name']}", 'success')
        except Exception:
            conn.rollback()
            flash('Error deleting event.', 'error')
        return redirect(url_for('banquet_planner'))

@bp.route('/banquet/events/<event_id>/duplicate', methods=['POST'])
@login_required
def banquet_event_duplicate(event_id):
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))

        cur.execute("""
            SELECT id, name, event_date, guest_count, venue_id, building, room, service_timing, dietary_notes, notes
            FROM banquet_events
            WHERE id = %s
        """, (event_id,))
        source_event = cur.fetchone()
        if not source_event:
            flash('Event not found.', 'error')
            return redirect(url_for('banquet_planner'))

        duplicate_name = clean_menu_text(request.form.get('duplicate_name'))
        if not duplicate_name:
            duplicate_name = f"{source_event.get('name') or 'Event'} (Copy)"

        duplicate_date_raw = (request.form.get('duplicate_event_date') or '').strip()
        try:
            duplicate_date = datetime.strptime(duplicate_date_raw, '%Y-%m-%d').date()
        except ValueError:
            flash('Duplicate date is required in YYYY-MM-DD format.', 'error')
            return redirect(url_for('banquet_event_edit', event_id=event_id))

        new_event_id = generate_id('bev_')
        has_line_choice_selections = banquet_event_line_choice_column_ready(cur)
        try:
            cur.execute("""
                INSERT INTO banquet_events (
                    id, name, event_date, guest_count, venue_id, building, room, service_timing, dietary_notes, notes, status
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                new_event_id,
                duplicate_name,
                duplicate_date,
                source_event.get('guest_count'),
                source_event.get('venue_id'),
                source_event.get('building'),
                source_event.get('room'),
                source_event.get('service_timing'),
                source_event.get('dietary_notes'),
                source_event.get('notes'),
                'planning'
            ))

            cur.execute("""
                SELECT
                    menu_item_id,
                    menu_item_name,
                    recipe_id,
                    quantity,
                    quantity_unit,
                    menu_section,
                    menu_descriptor,
                    notes,
                    sort_order,
                    {choice_selections_sql}
                FROM banquet_event_menu_items
                WHERE event_id = %s
                ORDER BY sort_order, id
            """.format(
                choice_selections_sql=("choice_selections" if has_line_choice_selections else "NULL::text AS choice_selections")
            ), (event_id,))
            source_lines = cur.fetchall()

            for line in source_lines:
                insert_banquet_event_menu_line(
                    cur,
                    new_event_id,
                    line.get('menu_item_id'),
                    line.get('menu_item_name'),
                    line.get('recipe_id'),
                    line.get('quantity'),
                    line.get('quantity_unit') or 'each',
                    line.get('menu_section'),
                    line.get('menu_descriptor'),
                    line.get('notes'),
                    line.get('sort_order'),
                    choice_selections=line.get('choice_selections'),
                    has_choice_selections=has_line_choice_selections
                )

            conn.commit()
            flash(f"Duplicated event as: {duplicate_name}", 'success')
        except Exception:
            conn.rollback()
            flash('Error duplicating event.', 'error')
            return redirect(url_for('banquet_event_edit', event_id=event_id))

        return redirect(url_for('banquet_event_edit', event_id=new_event_id))

@bp.route('/banquet-planner/events/<event_id>/apply-template', methods=['POST'])
@login_required
def banquet_event_apply_template(event_id):
    # Backwards-compatible endpoint: now applies the full menu catalog for the event venue.
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_event_edit', event_id=event_id))

        cur.execute("SELECT id, venue_id, guest_count FROM banquet_events WHERE id = %s", (event_id,))
        event = cur.fetchone()
        if not event:
            flash('Event not found.', 'error')
            return redirect(url_for('banquet_planner'))

        rows = list_banquet_menu_items(cur, event.get('venue_id') or '')
        if not rows:
            flash('No menu catalog items found for this venue.', 'error')
            return redirect(url_for('banquet_event_edit', event_id=event_id))

        guests = int(event.get('guest_count') or 0)
        has_line_choice_selections = banquet_event_line_choice_column_ready(cur)
        try:
            existing_lines = get_banquet_event_lines(cur, event_id)
            sort_base = len(existing_lines)
            for idx, item in enumerate(rows):
                qty = 1
                if guests > 0 and qty == 1:
                    qty = guests
                line = {
                    'menu_item_name': item.get('name') or '',
                    'menu_section': item.get('menu_section'),
                    'menu_descriptor': item.get('menu_descriptor'),
                    'recipe_id': item.get('linked_recipe_id')
                }
                menu_item_id = resolve_or_create_banquet_menu_item(cur, line, event.get('venue_id') or '')
                upsert_menu_item_recipe_link(cur, menu_item_id, line.get('recipe_id'))
                insert_banquet_event_menu_line(
                    cur,
                    event_id,
                    menu_item_id,
                    line.get('menu_item_name'),
                    line.get('recipe_id'),
                    qty,
                    'each',
                    line.get('menu_section'),
                    line.get('menu_descriptor'),
                    'Loaded from menu catalog',
                    sort_base + idx,
                    choice_selections=None,
                    has_choice_selections=has_line_choice_selections
                )
            conn.commit()
            flash('Loaded menu catalog lines to event.', 'success')
        except Exception:
            conn.rollback()
            flash('Error loading menu catalog.', 'error')
        return redirect(url_for('banquet_event_edit', event_id=event_id))

@bp.route('/banquet-planner/events/<event_id>/lines/<int:line_id>/clone-menu-item', methods=['POST'])
@login_required
def banquet_event_line_clone_menu_item(event_id, line_id):
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))

        cur.execute("""
            SELECT e.id,
                   e.name,
                   e.venue_id,
                   emi.id AS line_id,
                   emi.menu_item_id
            FROM banquet_events e
            JOIN banquet_event_menu_items emi ON emi.event_id = e.id
            WHERE e.id = %s AND emi.id = %s
        """, (event_id, line_id))
        row = cur.fetchone()
        if not row or not row.get('menu_item_id'):
            flash('Could not find the selected menu line to clone.', 'error')
            return redirect(url_for('banquet_event_edit', event_id=event_id))

        try:
            clone_label = clean_menu_text(request.form.get('clone_label')) or f"Custom {row.get('name') or 'Event'}"
            clone_item = clone_banquet_menu_item(
                cur,
                row.get('menu_item_id'),
                row.get('venue_id'),
                clone_label=clone_label
            )
            if not clone_item:
                conn.rollback()
                flash('Unable to clone menu item.', 'error')
                return redirect(url_for('banquet_event_edit', event_id=event_id))

            cur.execute("""
                SELECT recipe_id
                FROM banquet_menu_item_recipes
                WHERE menu_item_id = %s
                ORDER BY id
                LIMIT 1
            """, (clone_item['id'],))
            primary = cur.fetchone()
            recipe_id = primary.get('recipe_id') if primary else None

            cur.execute("""
                UPDATE banquet_event_menu_items
                SET menu_item_id = %s,
                    menu_item_name = %s,
                    menu_section = %s,
                    menu_descriptor = %s,
                    recipe_id = %s
                WHERE id = %s AND event_id = %s
            """, (
                clone_item['id'],
                clone_item.get('name'),
                clone_item.get('menu_section'),
                clone_item.get('menu_descriptor'),
                recipe_id,
                line_id,
                event_id
            ))
            conn.commit()

            return_to = url_for('banquet_event_edit', event_id=event_id)
            flash('Created custom copy. Update components, then return to event.', 'success')
            return redirect(url_for('banquet_menu_item_edit', menu_item_id=clone_item['id'], return_to=return_to))
        except Exception:
            conn.rollback()
            flash('Error creating custom menu copy.', 'error')
            return redirect(url_for('banquet_event_edit', event_id=event_id))

@bp.route('/banquet-planner/templates/import', methods=['GET', 'POST'])
@login_required
def banquet_template_import():
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))
        had_event_only_column = db_column_exists(cur, 'public.banquet_menu_items', 'is_event_only')
        ensure_banquet_menu_item_event_only_column(cur)
        if not had_event_only_column:
            conn.commit()
        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        banquet_venue_id = banquet_venue.get('id') or ''
        banquet_venue_name = banquet_venue.get('name') or 'Banquets'

        if request.method == 'POST':
            venue_id = banquet_venue_id
            source_pdf = request.files.get('menu_pdf')
            if not source_pdf or not source_pdf.filename.lower().endswith('.pdf'):
                flash('Upload a PDF menu file.', 'error')
            else:
                try:
                    parsed_items = parse_catering_pdf_to_template_items(source_pdf.stream)
                except Exception:
                    parsed_items = []
                if not parsed_items:
                    flash('Could not parse menu items from the PDF.', 'error')
                else:
                    recipes = get_plated_recipes_for_venue(cur, venue_id)
                    try:
                        created_count = 0
                        updated_count = 0
                        for item in parsed_items:
                            menu_name = clean_menu_text(item.get('item_name'))
                            if not menu_name:
                                continue
                            cur.execute("""
                                SELECT id
                                FROM banquet_menu_items
                                WHERE LOWER(name) = LOWER(%s)
                                  AND (venue_id = %s OR venue_id IS NULL)
                                ORDER BY CASE WHEN venue_id = %s THEN 0 ELSE 1 END, created_at
                                LIMIT 1
                            """, (menu_name, venue_id or None, venue_id or None))
                            existing = cur.fetchone()
                            matched_recipe_id = auto_match_menu_recipe_id(menu_name, recipes)
                            if existing:
                                cur.execute("""
                                    UPDATE banquet_menu_items
                                    SET menu_section = COALESCE(NULLIF(%s, ''), menu_section),
                                        menu_descriptor = COALESCE(NULLIF(%s, ''), menu_descriptor),
                                        venue_id = COALESCE(venue_id, %s),
                                        updated_at = CURRENT_TIMESTAMP
                                    WHERE id = %s
                                """, (
                                    item.get('menu_section') or '',
                                    item.get('menu_descriptor') or '',
                                    venue_id or None,
                                    existing['id']
                                ))
                                menu_item_id = existing['id']
                                updated_count += 1
                            else:
                                menu_item_id = generate_id('bmi_')
                                cur.execute("""
                                    INSERT INTO banquet_menu_items (id, name, venue_id, menu_section, menu_descriptor, notes)
                                    VALUES (%s, %s, %s, %s, %s, %s)
                                """, (
                                    menu_item_id,
                                    menu_name,
                                    venue_id or None,
                                    item.get('menu_section') or None,
                                    item.get('menu_descriptor') or None,
                                    f"Imported from {source_pdf.filename}"
                                ))
                                created_count += 1

                            if matched_recipe_id:
                                upsert_menu_item_recipe_link(cur, menu_item_id, matched_recipe_id)

                        conn.commit()
                        flash(f'Imported menu catalog: {created_count} created, {updated_count} updated.', 'success')
                        return redirect(url_for('banquet_planner'))
                    except Exception:
                        conn.rollback()
                        flash('Error saving imported menu catalog.', 'error')

        menu_items = list_banquet_menu_items(cur)
        menu_item_ids = [row.get('id') for row in menu_items if row.get('id')]
        summaries = get_banquet_menu_item_component_summaries(cur, menu_item_ids)
        for item in menu_items:
            summary = summaries.get(item.get('id'), {'recipes': [], 'ingredients': []})
            item['recipe_components'] = summary.get('recipes', [])
            item['ingredient_components'] = summary.get('ingredients', [])
            item['yield_display'] = f"{format_number(to_float(item.get('base_yield_qty')) or 1)} {item.get('base_yield_unit') or 'each'}"
        return render_template(
            'banquet_template_import.html',
            venues=venues,
            menu_items=menu_items,
            banquet_venue_name=banquet_venue_name
        )

def save_banquet_menu_item(cur, menu_item_id, form_state, ingredient_rows):
    ensure_banquet_menu_item_base_yield_columns(cur)
    has_base_yield = db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_qty') and db_column_exists(cur, 'public.banquet_menu_items', 'base_yield_unit')
    has_choice_columns = db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_group') and db_column_exists(cur, 'public.banquet_menu_item_recipes', 'choice_weight_percent')
    if menu_item_id:
        if has_base_yield:
            cur.execute("""
                UPDATE banquet_menu_items
                SET name = %s,
                    venue_id = %s,
                    menu_section = %s,
                    menu_descriptor = %s,
                    base_yield_qty = %s,
                    base_yield_unit = %s,
                    notes = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['base_yield_qty'],
                form_state['base_yield_unit'],
                form_state['notes'],
                menu_item_id
            ))
        else:
            cur.execute("""
                UPDATE banquet_menu_items
                SET name = %s,
                    venue_id = %s,
                    menu_section = %s,
                    menu_descriptor = %s,
                    notes = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['notes'],
                menu_item_id
            ))
    else:
        menu_item_id = generate_id('bmi_')
        if has_base_yield:
            cur.execute("""
                INSERT INTO banquet_menu_items (
                    id, name, venue_id, menu_section, menu_descriptor, base_yield_qty, base_yield_unit, notes
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                menu_item_id,
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['base_yield_qty'],
                form_state['base_yield_unit'],
                form_state['notes']
            ))
        else:
            cur.execute("""
                INSERT INTO banquet_menu_items (
                    id, name, venue_id, menu_section, menu_descriptor, notes
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                menu_item_id,
                form_state['name'],
                form_state['venue_id'],
                form_state['menu_section'],
                form_state['menu_descriptor'],
                form_state['notes']
            ))

    cur.execute("DELETE FROM banquet_menu_item_recipes WHERE menu_item_id = %s", (menu_item_id,))
    cur.execute("DELETE FROM banquet_menu_item_ingredients WHERE menu_item_id = %s", (menu_item_id,))

    resolved_recipe_components = []
    resolved_ingredient_rows = list(ingredient_rows or [])

    for component in form_state.get('recipe_components', []):
        source_menu_item_id = component.get('source_menu_item_id')
        if source_menu_item_id:
            source_item = get_banquet_menu_item(cur, source_menu_item_id)
            if source_item:
                source_multiplier = menu_line_base_multiplier(
                    component.get('quantity'),
                    component.get('unit'),
                    source_item.get('base_yield_qty') or 1,
                    source_item.get('base_yield_unit') or 'each'
                )
                if source_multiplier > 0:
                    source_recipe_components = get_banquet_menu_item_recipe_components(cur, source_menu_item_id)
                    apply_parent_choice = bool(component.get('choice_group')) and len(source_recipe_components) == 1
                    for source_component in source_recipe_components:
                        source_qty = to_float(source_component.get('quantity'))
                        expanded_qty = source_qty * source_multiplier
                        if expanded_qty <= 0:
                            continue
                        source_choice_group = source_component.get('choice_group')
                        parent_choice_group = component.get('choice_group') if apply_parent_choice else None
                        choice_group = source_choice_group or parent_choice_group
                        source_choice_weight = source_component.get('choice_weight_percent')
                        parent_choice_weight = component.get('choice_weight_percent') if apply_parent_choice else None
                        choice_weight = source_choice_weight if source_choice_group else parent_choice_weight
                        resolved_recipe_components.append({
                            'recipe_id': source_component.get('recipe_id'),
                            'quantity': expanded_qty,
                            'unit': source_component.get('unit') or source_component.get('yield_unit') or component.get('unit') or 'each',
                            'choice_group': choice_group or None,
                            'choice_weight_percent': choice_weight if choice_group else None
                        })

                    for source_ingredient in get_banquet_menu_item_ingredients(cur, source_menu_item_id):
                        source_ing_qty = to_float(source_ingredient.get('quantity'))
                        expanded_ing_qty = source_ing_qty * source_multiplier
                        if expanded_ing_qty <= 0:
                            continue
                        resolved_ingredient_rows.append({
                            'ingredient_id': source_ingredient.get('ingredient_id'),
                            'quantity': expanded_ing_qty,
                            'unit': source_ingredient.get('unit') or None
                        })
                    continue

        resolved_recipe_components.append({
            'recipe_id': component.get('recipe_id'),
            'quantity': to_float(component.get('quantity')),
            'unit': component.get('unit'),
            'choice_group': component.get('choice_group'),
            'choice_weight_percent': component.get('choice_weight_percent')
        })

    recipe_component_totals = {}
    for component in resolved_recipe_components:
        recipe_id = component.get('recipe_id')
        qty = to_float(component.get('quantity'))
        if not recipe_id or qty <= 0:
            continue
        unit = component.get('unit') or ''
        choice_group = component.get('choice_group') or ''
        choice_weight = component.get('choice_weight_percent')
        key = (
            recipe_id,
            unit,
            choice_group,
            round(to_float(choice_weight), 6) if choice_weight is not None else None
        )
        if key not in recipe_component_totals:
            recipe_component_totals[key] = dict(component)
            recipe_component_totals[key]['quantity'] = qty
        else:
            recipe_component_totals[key]['quantity'] = to_float(recipe_component_totals[key].get('quantity')) + qty

    final_recipe_components = list(recipe_component_totals.values())
    ingredient_totals = {}
    for row in resolved_ingredient_rows:
        ingredient_id = row.get('ingredient_id')
        qty = to_float(row.get('quantity'))
        if not ingredient_id or qty <= 0:
            continue
        unit = row.get('unit') or ''
        key = (ingredient_id, unit)
        ingredient_totals[key] = ingredient_totals.get(key, 0) + qty

    final_ingredient_rows = [
        {
            'ingredient_id': ingredient_id,
            'quantity': qty,
            'unit': unit or None
        }
        for (ingredient_id, unit), qty in ingredient_totals.items()
    ]

    for component in final_recipe_components:
        if has_choice_columns:
            cur.execute("""
                INSERT INTO banquet_menu_item_recipes (menu_item_id, recipe_id, quantity, unit, choice_group, choice_weight_percent)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                menu_item_id,
                component['recipe_id'],
                component['quantity'],
                component['unit'],
                component.get('choice_group'),
                component.get('choice_weight_percent')
            ))
        else:
            cur.execute("""
                INSERT INTO banquet_menu_item_recipes (menu_item_id, recipe_id, quantity, unit)
                VALUES (%s, %s, %s, %s)
            """, (
                menu_item_id,
                component['recipe_id'],
                component['quantity'],
                component['unit']
            ))

    for row in final_ingredient_rows:
        cur.execute("""
            INSERT INTO banquet_menu_item_ingredients (menu_item_id, ingredient_id, quantity, unit)
            VALUES (%s, %s, %s, %s)
        """, (
            menu_item_id,
            row['ingredient_id'],
            row['quantity'],
            row['unit'] or None
        ))
    return menu_item_id

def render_banquet_menu_item_form(cur, mode='new', menu_item_id=None, form_state=None, ingredient_rows=None, errors=None, return_to=None):
    venues = get_active_venues(cur)
    banquet_venue = resolve_banquet_venue(venues)
    banquet_venue_id = banquet_venue.get('id') or ''
    banquet_venue_name = banquet_venue.get('name') or 'Banquets'

    if form_state is None:
        form_state = {
            'name': '',
            'venue_id': banquet_venue_id or None,
            'menu_section': '',
            'menu_descriptor': '',
            'base_yield_qty': 1,
            'base_yield_unit': 'each',
            'notes': '',
            'recipe_components': []
        }
        if menu_item_id:
            existing = get_banquet_menu_item(cur, menu_item_id)
            if existing:
                form_state.update({
                    'name': existing.get('name') or '',
                    'venue_id': existing.get('venue_id') or banquet_venue_id or None,
                    'menu_section': existing.get('menu_section') or '',
                    'menu_descriptor': existing.get('menu_descriptor') or '',
                    'base_yield_qty': to_float(existing.get('base_yield_qty')) or 1,
                    'base_yield_unit': existing.get('base_yield_unit') or 'each',
                    'notes': existing.get('notes') or '',
                    'recipe_components': get_banquet_menu_item_recipe_components(cur, menu_item_id)
                })
    if ingredient_rows is None:
        ingredient_rows = get_banquet_menu_item_ingredients(cur, menu_item_id) if menu_item_id else []

    component_recipes = get_component_recipes_for_venue(cur, banquet_venue_id)
    seen_component_labels = set()
    for recipe in component_recipes:
        recipe_type = (recipe.get('recipe_type') or '').upper()
        if recipe.get('option_source') == 'menu_item' and recipe.get('menu_item_name'):
            base_label = f"{recipe.get('menu_item_name')} [Menu Item] -> {recipe.get('name')}{f' ({recipe_type})' if recipe_type else ''}"
        else:
            base_label = f"{recipe.get('name')}{f' ({recipe_type})' if recipe_type else ''}"
        candidate = base_label
        suffix = 2
        while candidate.lower() in seen_component_labels:
            candidate = f"{base_label} [{suffix}]"
            suffix += 1
        seen_component_labels.add(candidate.lower())
        recipe['display_label'] = candidate
    cur.execute("SELECT id, name, unit, category FROM ingredients ORDER BY name")
    ingredients = cur.fetchall()
    for ingredient in ingredients:
        unit_label = ingredient.get('unit')
        ingredient['display_label'] = f"{ingredient.get('name')}{f' ({unit_label})' if unit_label else ''}"

    recipe_label_map = {}
    for row in component_recipes:
        recipe_id = row.get('id')
        if not recipe_id:
            continue
        row_label = row.get('display_label') or row.get('name') or ''
        if recipe_id not in recipe_label_map or row.get('option_source') == 'recipe':
            recipe_label_map[recipe_id] = row_label
    ingredient_label_map = {row['id']: row.get('display_label') or row.get('name') for row in ingredients}

    for row in form_state.get('recipe_components', []):
        if isinstance(row, dict):
            row['display_label'] = recipe_label_map.get(row.get('recipe_id'), row.get('recipe_name') or '')
    for row in ingredient_rows:
        if isinstance(row, dict):
            row['display_label'] = ingredient_label_map.get(row.get('ingredient_id'), row.get('ingredient_name') or '')

    return render_template(
        'banquet_menu_item_form.html',
        mode=mode,
        menu_item_id=menu_item_id,
        form_state=form_state,
        ingredient_rows=ingredient_rows,
        component_recipes=component_recipes,
        ingredients=ingredients,
        section_options=BANQUET_MENU_SECTION_OPTIONS,
        banquet_venue_name=banquet_venue_name,
        errors=errors or [],
        return_to=return_to
    )

@bp.route('/banquet-planner/menu-items/new', methods=['GET', 'POST'])
@login_required
def banquet_menu_item_new():
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))
        return_to = normalize_return_path(request.values.get('return_to'))

        venues = get_active_venues(cur)
        banquet_venue_id = (resolve_banquet_venue(venues) or {}).get('id') or ''
        component_recipes = get_component_recipes_for_venue(cur, banquet_venue_id)
        valid_recipe_ids = {row['id'] for row in component_recipes}
        valid_source_menu_item_ids = {row.get('source_menu_item_id') for row in component_recipes if row.get('source_menu_item_id')}
        cur.execute("SELECT id FROM ingredients")
        valid_ingredient_ids = {row['id'] for row in cur.fetchall()}

        if request.method == 'POST':
            errors = []
            base_yield_qty_raw = (request.form.get('base_yield_qty') or '').strip()
            base_yield_qty = parse_float_field(
                base_yield_qty_raw,
                'Menu yield quantity',
                errors,
                required=True,
                min_value=0.0001
            )
            base_yield_unit_raw = clean_menu_text(request.form.get('base_yield_unit'))
            base_yield_unit = normalize_count_unit(base_yield_unit_raw)
            if base_yield_unit_raw and not base_yield_unit:
                errors.append('Sold As Unit must be "each" or "dozen".')
            base_yield_unit = base_yield_unit or 'each'
            form_state = {
                'name': clean_menu_text(request.form.get('name')),
                'venue_id': banquet_venue_id or None,
                'menu_section': clean_menu_text(request.form.get('menu_section')) or None,
                'menu_descriptor': clean_menu_text(request.form.get('menu_descriptor')) or None,
                'base_yield_qty': base_yield_qty if base_yield_qty is not None else (base_yield_qty_raw or 1),
                'base_yield_unit': base_yield_unit if base_yield_unit else (base_yield_unit_raw or 'each'),
                'notes': clean_menu_text(request.form.get('notes')) or None,
                'recipe_components': []
            }
            if not form_state['name']:
                errors.append('Menu item name is required.')

            recipe_components, component_errors = parse_menu_item_recipe_lines(
                request,
                valid_recipe_ids,
                valid_source_menu_item_ids
            )
            form_state['recipe_components'] = recipe_components
            errors.extend(component_errors)

            ingredient_rows, ingredient_errors = parse_menu_item_ingredient_lines(request, valid_ingredient_ids)
            errors.extend(ingredient_errors)

            if errors:
                response = render_banquet_menu_item_form(
                    cur,
                    mode='new',
                    menu_item_id=None,
                    form_state=form_state,
                    ingredient_rows=ingredient_rows,
                    errors=sorted(set(errors)),
                    return_to=return_to
                )
                return response

            try:
                save_banquet_menu_item(cur, None, form_state, ingredient_rows)
                conn.commit()
                flash('Banquet menu item created.', 'success')
                return redirect(return_to or url_for('banquet_template_import'))
            except Exception:
                conn.rollback()
                response = render_banquet_menu_item_form(
                    cur,
                    mode='new',
                    menu_item_id=None,
                    form_state=form_state,
                    ingredient_rows=ingredient_rows,
                    errors=['Error saving menu item.'],
                    return_to=return_to
                )
                return response

        response = render_banquet_menu_item_form(cur, mode='new', menu_item_id=None, return_to=return_to)
        return response

@bp.route('/banquet-planner/menu-items/<menu_item_id>/edit', methods=['GET', 'POST'])
@login_required
def banquet_menu_item_edit(menu_item_id):
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))
        return_to = normalize_return_path(request.values.get('return_to'))

        existing = get_banquet_menu_item(cur, menu_item_id)
        if not existing:
            flash('Menu item not found.', 'error')
            return redirect(url_for('banquet_template_import'))

        venues = get_active_venues(cur)
        banquet_venue_id = (resolve_banquet_venue(venues) or {}).get('id') or existing.get('venue_id') or ''
        component_recipes = get_component_recipes_for_venue(cur, banquet_venue_id)
        valid_recipe_ids = {row['id'] for row in component_recipes}
        valid_source_menu_item_ids = {row.get('source_menu_item_id') for row in component_recipes if row.get('source_menu_item_id')}
        cur.execute("SELECT id FROM ingredients")
        valid_ingredient_ids = {row['id'] for row in cur.fetchall()}

        if request.method == 'POST':
            errors = []
            base_yield_qty_raw = (request.form.get('base_yield_qty') or '').strip()
            base_yield_qty = parse_float_field(
                base_yield_qty_raw,
                'Menu yield quantity',
                errors,
                required=True,
                min_value=0.0001
            )
            base_yield_unit_raw = clean_menu_text(request.form.get('base_yield_unit'))
            base_yield_unit = normalize_count_unit(base_yield_unit_raw)
            if base_yield_unit_raw and not base_yield_unit:
                errors.append('Sold As Unit must be "each" or "dozen".')
            base_yield_unit = base_yield_unit or 'each'
            form_state = {
                'name': clean_menu_text(request.form.get('name')),
                'venue_id': banquet_venue_id or None,
                'menu_section': clean_menu_text(request.form.get('menu_section')) or None,
                'menu_descriptor': clean_menu_text(request.form.get('menu_descriptor')) or None,
                'base_yield_qty': base_yield_qty if base_yield_qty is not None else (base_yield_qty_raw or 1),
                'base_yield_unit': base_yield_unit if base_yield_unit else (base_yield_unit_raw or 'each'),
                'notes': clean_menu_text(request.form.get('notes')) or None,
                'recipe_components': []
            }
            if not form_state['name']:
                errors.append('Menu item name is required.')

            recipe_components, component_errors = parse_menu_item_recipe_lines(
                request,
                valid_recipe_ids,
                valid_source_menu_item_ids
            )
            form_state['recipe_components'] = recipe_components
            errors.extend(component_errors)

            ingredient_rows, ingredient_errors = parse_menu_item_ingredient_lines(request, valid_ingredient_ids)
            errors.extend(ingredient_errors)

            if errors:
                response = render_banquet_menu_item_form(
                    cur,
                    mode='edit',
                    menu_item_id=menu_item_id,
                    form_state=form_state,
                    ingredient_rows=ingredient_rows,
                    errors=sorted(set(errors)),
                    return_to=return_to
                )
                return response

            try:
                save_banquet_menu_item(cur, menu_item_id, form_state, ingredient_rows)
                conn.commit()
                flash('Banquet menu item updated.', 'success')
                return redirect(return_to or url_for('banquet_template_import'))
            except Exception:
                conn.rollback()
                response = render_banquet_menu_item_form(
                    cur,
                    mode='edit',
                    menu_item_id=menu_item_id,
                    form_state=form_state,
                    ingredient_rows=ingredient_rows,
                    errors=['Error saving menu item.'],
                    return_to=return_to
                )
                return response

        response = render_banquet_menu_item_form(cur, mode='edit', menu_item_id=menu_item_id, return_to=return_to)
        return response

@bp.route('/banquet-planner/menu-items/<menu_item_id>/delete', methods=['POST'])
@login_required
def banquet_menu_item_delete(menu_item_id):
    conn = get_db()
    with get_cursor() as cur:
        if not banquet_tables_ready(cur):
            flash('Banquet tables are missing. Run migrations first.', 'error')
            return redirect(url_for('banquet_planner'))

        cur.execute("SELECT id, name FROM banquet_menu_items WHERE id = %s", (menu_item_id,))
        menu_item = cur.fetchone()
        if not menu_item:
            flash('Menu item not found.', 'error')
            return redirect(url_for('banquet_template_import'))
        try:
            cur.execute("DELETE FROM banquet_menu_items WHERE id = %s", (menu_item_id,))
            conn.commit()
            flash(f"Deleted menu item: {menu_item['name']}", 'success')
        except Exception:
            conn.rollback()
            flash('Error deleting menu item.', 'error')
        return redirect(url_for('banquet_template_import'))

@bp.route('/banquet-planner/templates/<template_id>/delete', methods=['POST'])
@login_required
def banquet_template_delete(template_id):
    # Backwards-compatible alias while old UI links still exist.
    return banquet_menu_item_delete(template_id)

@bp.route('/banquet-planner/shopping')
@login_required
def banquet_shopping():
    conn = get_db()
    with get_cursor() as cur:
        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        selected_venue = banquet_venue.get('id') or ''
        start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
        auto_complete_past_banquet_events(cur, selected_venue)
        datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
        checklist_state = load_banquet_shopping_checklist(cur, selected_venue, start_date, end_date)
        grouped = defaultdict(lambda: defaultdict(list))
        for row in datasets['shopping_ingredients']:
            vendor = row.get('vendor') or 'Unassigned Vendor'
            category = row.get('category') or 'Uncategorized'
            grouped[vendor][category].append(row)
        return render_template(
            'banquet_shopping.html',
            venues=[banquet_venue] if selected_venue else venues,
            selected_venue=selected_venue,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            grouped=grouped,
            checklist_state=checklist_state,
            ingredient_total=len(datasets['shopping_ingredients']),
            total_cost=datasets['shopping_total_cost']
        )

@bp.route('/banquet-planner/shopping/checklist', methods=['GET', 'POST'])
@login_required
def banquet_shopping_checklist_upsert():
    if request.method == 'GET':
        venue_id = (request.args.get('venue_id') or BANQUET_VENUE_ID).strip() or BANQUET_VENUE_ID
        try:
            start_date = datetime.strptime((request.args.get('start_date') or '').strip(), '%Y-%m-%d').date()
            end_date = datetime.strptime((request.args.get('end_date') or '').strip(), '%Y-%m-%d').date()
        except (ValueError, AttributeError):
            return jsonify({'ok': False, 'error': 'Invalid date window'}), 400

        conn = get_db()
        with get_cursor() as cur:
            ensure_banquet_shopping_checks_table(cur)

            state = load_banquet_shopping_checklist(cur, venue_id, start_date, end_date)
            return jsonify({'ok': True, 'state': state})

        payload = request.get_json(silent=True) or {}
        venue_id = (payload.get('venue_id') or BANQUET_VENUE_ID).strip() or BANQUET_VENUE_ID
        item_key = clean_menu_text(payload.get('item_key'))
        if not item_key:
            return jsonify({'ok': False, 'error': 'Missing item key'}), 400

        try:
            start_date = datetime.strptime((payload.get('start_date') or '').strip(), '%Y-%m-%d').date()
            end_date = datetime.strptime((payload.get('end_date') or '').strip(), '%Y-%m-%d').date()
        except (ValueError, AttributeError):
            return jsonify({'ok': False, 'error': 'Invalid date window'}), 400

        checked = bool(payload.get('checked'))
        note = clean_menu_text(payload.get('note'))
        ingredient_id = clean_menu_text(payload.get('ingredient_id')) or None
        unit = clean_menu_text(payload.get('unit')) or None
        vendor = clean_menu_text(payload.get('vendor')) or None

        conn = get_db()
        with get_cursor() as cur:
            ensure_banquet_shopping_checks_table(cur)

            try:
                if not checked and not note:
                    cur.execute("""
                        DELETE FROM banquet_shopping_checks
                        WHERE venue_id = %s
                          AND start_date = %s
                          AND end_date = %s
                          AND item_key = %s
                    """, (venue_id, start_date, end_date, item_key))
                else:
                    cur.execute("""
                        INSERT INTO banquet_shopping_checks (
                            venue_id, start_date, end_date, item_key,
                            ingredient_id, unit, vendor, checked, note, updated_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        ON CONFLICT (venue_id, start_date, end_date, item_key)
                        DO UPDATE SET
                            ingredient_id = EXCLUDED.ingredient_id,
                            unit = EXCLUDED.unit,
                            vendor = EXCLUDED.vendor,
                            checked = EXCLUDED.checked,
                            note = EXCLUDED.note,
                            updated_at = NOW()
                    """, (venue_id, start_date, end_date, item_key, ingredient_id, unit, vendor, checked, note))
                conn.commit()
                return jsonify({'ok': True})
            except Exception:
                conn.rollback()
                return jsonify({'ok': False, 'error': 'Failed to save checklist row'}), 500

@bp.route('/banquet-planner/shopping/checklist/reset', methods=['POST'])
@login_required
def banquet_shopping_checklist_reset():
    payload = request.get_json(silent=True) or {}
    venue_id = (payload.get('venue_id') or BANQUET_VENUE_ID).strip() or BANQUET_VENUE_ID
    try:
        start_date = datetime.strptime((payload.get('start_date') or '').strip(), '%Y-%m-%d').date()
        end_date = datetime.strptime((payload.get('end_date') or '').strip(), '%Y-%m-%d').date()
    except (ValueError, AttributeError):
        return jsonify({'ok': False, 'error': 'Invalid date window'}), 400

    conn = get_db()
    with get_cursor() as cur:
        ensure_banquet_shopping_checks_table(cur)

        try:
            cur.execute("""
                DELETE FROM banquet_shopping_checks
                WHERE venue_id = %s
                  AND start_date = %s
                  AND end_date = %s
            """, (venue_id, start_date, end_date))
            conn.commit()
            return jsonify({'ok': True})
        except Exception:
            conn.rollback()
            return jsonify({'ok': False, 'error': 'Failed to reset checklist'}), 500

@bp.route('/banquet-planner/prep-weekly')
@login_required
def banquet_prep_weekly():
    conn = get_db()
    with get_cursor() as cur:
        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        selected_venue = banquet_venue.get('id') or ''
        start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
        auto_complete_past_banquet_events(cur, selected_venue)
        datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
        event_names = {event['id']: event['name'] for event in datasets['events']}
        for row in datasets['weekly_prep']:
            row['used_in_event_names'] = [event_names[event_id] for event_id in row['used_in_events'] if event_id in event_names]
        for row in datasets.get('weekly_menu_pulls', []):
            row['used_in_event_names'] = [event_names[event_id] for event_id in row.get('used_in_events', []) if event_id in event_names]
        return render_template(
            'banquet_prep_weekly.html',
            venues=[banquet_venue] if selected_venue else venues,
            selected_venue=selected_venue,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            weekly_prep=datasets['weekly_prep'],
            weekly_menu_pulls=datasets.get('weekly_menu_pulls', [])
        )

@bp.route('/banquet-planner/prep-daily')
@login_required
def banquet_prep_daily():
    conn = get_db()
    with get_cursor() as cur:
        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        selected_venue = banquet_venue.get('id') or ''
        start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
        auto_complete_past_banquet_events(cur, selected_venue)
        datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
        return render_template(
            'banquet_prep_daily.html',
            venues=[banquet_venue] if selected_venue else venues,
            selected_venue=selected_venue,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            daily_groups=datasets['daily_groups']
        )

@bp.route('/banquet-planner/export/excel')
@login_required
def banquet_export_excel():
    conn = get_db()
    with get_cursor() as cur:
        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        selected_venue = banquet_venue.get('id') or ''
        start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
        auto_complete_past_banquet_events(cur, selected_venue)
        datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Event Summary'
        headers = ['Date', 'Event', 'Venue', 'Guests', 'Menu Item', 'Section', 'Recipe', 'Qty', 'Unit', 'Descriptor', 'Line Mods', 'Estimated Cost']
        ws_summary.append(headers)
        for col in ws_summary[1]:
            col.font = Font(bold=True, color='FFFFFF')
            col.fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        for event in datasets['events']:
            for line in event['lines']:
                ws_summary.append([
                    event.get('event_date').isoformat() if event.get('event_date') else '',
                    event.get('name'),
                    event.get('venue_name'),
                    event.get('guests'),
                    line.get('menu_item_name'),
                    line.get('menu_section'),
                    ', '.join(r.get('name') for r in line.get('recipes', []) if r.get('name')) or (line.get('recipe') or {}).get('name'),
                    line.get('quantity'),
                    line.get('quantity_unit'),
                    line.get('menu_descriptor'),
                    line.get('notes'),
                    line.get('estimated_cost_total')
                ])

        ws_shopping = wb.create_sheet('Shopping List')
        ws_shopping.append(['Vendor', 'Category', 'Ingredient', 'Qty', 'Unit', 'Price/Unit', 'Ext Cost', 'Vendor Code', 'G-Code', 'Used In'])
        for col in ws_shopping[1]:
            col.font = Font(bold=True, color='FFFFFF')
            col.fill = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
        for ing in datasets['shopping_ingredients']:
            ws_shopping.append([
                ing.get('vendor') or 'Unassigned',
                ing.get('category'),
                ing.get('name'),
                ing.get('display_quantity'),
                ing.get('display_unit'),
                ing.get('cost_per_unit'),
                ing.get('ext_cost'),
                ing.get('vendor_code'),
                ing.get('g_code'),
                ', '.join(ing.get('used_in') or [])
            ])

        ws_weekly = wb.create_sheet('Weekly Prep')
        ws_weekly.append(['Prep Recipe', 'Type', 'Required Output', 'Yield', 'Required Batches', 'Estimated Cost', 'Used In Events', 'Used In Menu Items'])
        for col in ws_weekly[1]:
            col.font = Font(bold=True, color='FFFFFF')
            col.fill = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
        event_name_map = {event['id']: event['name'] for event in datasets['events']}
        for row in datasets['weekly_prep']:
            ws_weekly.append([
                row.get('recipe_name'),
                (row.get('recipe_type') or 'other'),
                f"{row['display_required']['quantity']} {row['display_required']['unit']}",
                f"{format_number(row.get('yield_qty'))} {row.get('yield_unit') or ''}",
                round(row.get('required_batches') or 0, 4),
                row.get('estimated_cost'),
                ', '.join(event_name_map.get(event_id, event_id) for event_id in row.get('used_in_events', [])),
                ', '.join(row.get('used_in_menu_items', []))
            ])

        ws_daily = wb.create_sheet('Daily Prep')
        ws_daily.append(['Date', 'Event', 'Menu Item', 'Descriptor', 'Qty', 'Unit', 'Recipe', 'Line Mods'])
        for col in ws_daily[1]:
            col.font = Font(bold=True, color='FFFFFF')
            col.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        for day in datasets['daily_groups']:
            for event in day['events']:
                for line in event.get('lines', []):
                    ws_daily.append([
                        day['date'].isoformat() if day.get('date') else '',
                        event.get('name'),
                        line.get('menu_item_name'),
                        line.get('menu_descriptor'),
                        line.get('quantity'),
                        line.get('quantity_unit'),
                        ', '.join(r.get('name') for r in line.get('recipes', []) if r.get('name')) or (line.get('recipe') or {}).get('name'),
                        line.get('notes')
                    ])

        for ws in [ws_summary, ws_shopping, ws_weekly, ws_daily]:
            for column_cells in ws.columns:
                max_length = 0
                letter = column_cells[0].column_letter
                for cell in column_cells:
                    value = '' if cell.value is None else str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 56)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        venue_name = banquet_venue.get('name') if banquet_venue else 'banquets'
        filename = f"{make_safe_filename(venue_name)}_{start_date.isoformat()}_{end_date.isoformat()}_banquet_planner.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

@bp.route('/banquet-planner/packet/print')
@login_required
def banquet_packet_print():
    conn = get_db()
    with get_cursor() as cur:
        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        selected_venue = banquet_venue.get('id') or ''
        start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
        include_shopping = (request.args.get('include_shopping') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        include_beo_raw = request.args.get('include_beo')
        if include_beo_raw is None:
            include_beo = True
        else:
            include_beo = (include_beo_raw or '').strip().lower() in ('1', 'true', 'yes', 'on')
        print_lite = (request.args.get('print_lite') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        auto_complete_past_banquet_events(cur, selected_venue)
        datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
        beo_files_by_event = list_banquet_event_beo_files_by_events(cur, [event.get('id') for event in datasets.get('events', [])])
        event_name_map = {event['id']: event['name'] for event in datasets.get('events', [])}
        for prep in datasets.get('weekly_prep', []):
            prep['instruction_steps'] = split_instruction_steps(prep.get('instructions'))
        prep_groups = build_banquet_prep_groups(cur, datasets, get_unit_system())
        for pull in datasets.get('weekly_menu_pulls', []):
            pull['used_in_event_names'] = [event_name_map[event_id] for event_id in pull.get('used_in_events', []) if event_id in event_name_map]
        venue_name = banquet_venue.get('name') if banquet_venue else 'Banquets'
        return render_template(
            'banquet_packet_print.html',
            selected_venue=selected_venue,
            venue_name=venue_name,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            include_shopping=include_shopping,
            include_beo=include_beo,
            print_lite=print_lite,
            beo_files_by_event=beo_files_by_event,
            generated_at=datetime.now().strftime('%b %d, %Y %I:%M %p'),
            prep_groups=prep_groups,
            datasets=datasets
        )


@bp.route('/banquet-planner/packet/pdf')
@login_required
def banquet_packet_pdf():
    conn = get_db()
    with get_cursor() as cur:
        venues = get_active_venues(cur)
        banquet_venue = resolve_banquet_venue(venues)
        selected_venue = banquet_venue.get('id') or ''
        start_date, end_date = get_banquet_date_window(request.args.get('start_date'), request.args.get('end_date'))
        include_shopping = (request.args.get('include_shopping') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        include_beo_raw = request.args.get('include_beo')
        if include_beo_raw is None:
            include_beo = True
        else:
            include_beo = (include_beo_raw or '').strip().lower() in ('1', 'true', 'yes', 'on')
        print_lite = (request.args.get('print_lite') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        simple_raw = (request.args.get('simple') or '').strip().lower()
        simplified_view = simple_raw in ('1', 'true', 'yes', 'on')

        auto_complete_past_banquet_events(cur, selected_venue)
        datasets = build_banquet_datasets(cur, start_date, end_date, selected_venue, get_unit_system())
        event_name_map = {event['id']: event['name'] for event in datasets.get('events', [])}
        for prep in datasets.get('weekly_prep', []):
            prep['instruction_steps'] = split_instruction_steps(prep.get('instructions'))
        prep_groups = build_banquet_prep_groups(cur, datasets, get_unit_system())
        for pull in datasets.get('weekly_menu_pulls', []):
            pull['used_in_event_names'] = [event_name_map[event_id] for event_id in pull.get('used_in_events', []) if event_id in event_name_map]
        venue_name = banquet_venue.get('name') if banquet_venue else 'Banquets'

    header_navy = colors.HexColor('#1B2A4A')
    light_gray = colors.HexColor('#F5F5F5')
    border_gray = colors.HexColor('#DDDDDD')
    muted_text = colors.HexColor('#4F5B6E')
    white = colors.white
    margin_x = 0.5 * inch
    margin_y = 0.4 * inch
    generated_at = datetime.now().strftime('%b %d, %Y %I:%M %p')

    def to_text(value, default='-'):
        text = str(value or '').strip()
        return text or default

    def html_text(value, default='-'):
        return escape(to_text(value, default)).replace('\n', '<br/>')

    def p(value, style, default='-'):
        return Paragraph(html_text(value, default), style)

    base_style = ParagraphStyle(
        'banquet-pdf-base',
        fontName='Helvetica',
        fontSize=8,
        leading=9.4,
        textColor=colors.black,
    )
    tiny_style = ParagraphStyle(
        'banquet-pdf-tiny',
        parent=base_style,
        fontSize=7.2,
        leading=8.3,
    )
    table_header_style = ParagraphStyle(
        'banquet-pdf-table-header',
        parent=base_style,
        fontName='Helvetica-Bold',
        textColor=white,
        fontSize=7.2,
        leading=8.3,
    )
    section_title_style = ParagraphStyle(
        'banquet-pdf-section-title',
        parent=base_style,
        fontName='Helvetica-Bold',
        fontSize=10.5,
        leading=12,
        textColor=header_navy,
    )
    subheading_style = ParagraphStyle(
        'banquet-pdf-subheading',
        parent=base_style,
        fontName='Helvetica-Bold',
        fontSize=8.6,
        leading=10.2,
        textColor=header_navy,
    )
    muted_style = ParagraphStyle(
        'banquet-pdf-muted',
        parent=base_style,
        fontSize=7.3,
        leading=8.4,
        textColor=muted_text,
    )

    def draw_page_header(pdf_canvas, doc):
        page_width, page_height = letter
        top_y = page_height - 22
        pdf_canvas.saveState()
        pdf_canvas.setFillColor(header_navy)
        pdf_canvas.setFont('Helvetica-Bold', 11)
        pdf_canvas.drawString(doc.leftMargin, top_y, 'Banquet Planning Packet')
        pdf_canvas.setFillColor(muted_text)
        pdf_canvas.setFont('Helvetica', 7.3)
        meta = f'{to_text(venue_name)} | {start_date.isoformat()} - {end_date.isoformat()}'
        if print_lite:
            meta = f'{meta} | Print Lite'
        elif simplified_view:
            meta = f'{meta} | Simplified View'
        pdf_canvas.drawString(doc.leftMargin, top_y - 10, meta)
        pdf_canvas.drawRightString(page_width - doc.rightMargin, top_y - 10, f'Generated {generated_at}')
        pdf_canvas.setStrokeColor(border_gray)
        pdf_canvas.setLineWidth(0.6)
        pdf_canvas.line(doc.leftMargin, top_y - 14.5, page_width - doc.rightMargin, top_y - 14.5)
        pdf_canvas.restoreState()

    def draw_footer(pdf_canvas, page_num, total_pages):
        page_width, _ = letter
        footer_y = 16
        pdf_canvas.setStrokeColor(border_gray)
        pdf_canvas.setLineWidth(0.6)
        pdf_canvas.line(margin_x, footer_y + 10, page_width - margin_x, footer_y + 10)
        pdf_canvas.setFont('Helvetica', 7.5)
        pdf_canvas.setFillColor(muted_text)
        pdf_canvas.drawString(margin_x, footer_y + 2, 'Foxtown HQ')
        pdf_canvas.drawRightString(page_width - margin_x, footer_y + 2, f'Page {page_num} of {total_pages}')

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total_pages = len(self._saved_page_states)
            for page_num, state in enumerate(self._saved_page_states, start=1):
                self.__dict__.update(state)
                draw_footer(self, page_num, total_pages)
                super().showPage()
            super().save()

    def draw_first_page(pdf_canvas, doc):
        draw_page_header(pdf_canvas, doc)

    def draw_later_pages(pdf_canvas, doc):
        draw_page_header(pdf_canvas, doc)

    def style_table(table, align_right_cols=None):
        commands = [
            ('BACKGROUND', (0, 0), (-1, 0), header_navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, light_gray]),
            ('GRID', (0, 0), (-1, -1), 0.45, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        if align_right_cols:
            for idx in align_right_cols:
                commands.append(('ALIGN', (idx, 1), (idx, -1), 'RIGHT'))
        table.setStyle(TableStyle(commands))
        return table

    def step_lines(card):
        steps = card.get('instruction_steps') or []
        if steps:
            return steps
        return split_instruction_steps(card.get('instructions'))

    def render_sub_card(story, card, depth=1):
        story.append(Paragraph(f"Sub-Recipe (L{depth}): {to_text(card.get('recipe_name'), 'Sub-Recipe')}", subheading_style))
        ingredient_rows = card.get('ingredient_rows') or []
        if ingredient_rows:
            sub_ing_data = [
                [
                    Paragraph('✓', table_header_style),
                    Paragraph('Ingredient', table_header_style),
                    Paragraph('Qty', table_header_style),
                    Paragraph('Unit', table_header_style),
                ]
            ]
            for ing in ingredient_rows:
                sub_ing_data.append(
                    [
                        p('[ ]', tiny_style),
                        p(ing.get('name') or 'Ingredient', base_style),
                        p(ing.get('display_quantity') if ing.get('display_quantity') not in (None, '') else '—', base_style),
                        p(ing.get('display_unit') or '—', base_style),
                    ]
                )
            sub_ing_table = Table(sub_ing_data, colWidths=[0.3 * inch, 4.0 * inch, 0.8 * inch, 0.75 * inch], repeatRows=1)
            style_table(sub_ing_table, align_right_cols=[2])
            story.append(sub_ing_table)
        for idx, step in enumerate(step_lines(card), start=1):
            story.append(Paragraph(f'{idx}. {html_text(step, "")}', tiny_style))
        for child in card.get('child_cards') or []:
            story.append(Spacer(1, 4))
            render_sub_card(story, child, depth + 1)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=margin_x,
        rightMargin=margin_x,
        topMargin=0.9 * inch,
        bottomMargin=0.5 * inch,
        title='Banquet Planning Packet',
    )
    story = [Spacer(1, 2)]

    if include_beo:
        story.append(Paragraph('BEO Service Brief', section_title_style))
        story.append(Paragraph(f'{start_date.isoformat()} - {end_date.isoformat()}', muted_style))
        story.append(Spacer(1, 4))
        events = datasets.get('events') or []
        if events:
            for event in events:
                story.append(
                    Paragraph(
                        f"<b>{escape(to_text(event.get('name'), 'Event'))}</b> | "
                        f"{to_text(event.get('event_date'))} | {to_text(event.get('venue_name'))} | "
                        f"{to_text(event.get('guests'), '—')} guests",
                        base_style,
                    )
                )
                if event.get('service_timing'):
                    story.append(Paragraph(f"Service Timing: {html_text(event.get('service_timing'), '')}", muted_style))
                if event.get('dietary_notes'):
                    story.append(Paragraph(f"Dietary / Allergy Notes: {html_text(event.get('dietary_notes'), '')}", muted_style))
                if event.get('notes'):
                    story.append(Paragraph(f"General Notes: {html_text(event.get('notes'), '')}", muted_style))

                lines_data = [
                    [
                        Paragraph('✓', table_header_style),
                        Paragraph('Menu Item', table_header_style),
                        Paragraph('Qty', table_header_style),
                        Paragraph('Notes / Mods', table_header_style),
                    ]
                ]
                current_section = None
                for line in event.get('lines', []) or []:
                    section_name = line.get('menu_section') or 'Other'
                    if section_name != current_section:
                        current_section = section_name
                        lines_data.append(
                            [
                                p('', base_style, default=''),
                                Paragraph(f"<b>{escape(section_name)}</b>", base_style),
                                p('', base_style, default=''),
                                p('', base_style, default=''),
                            ]
                        )
                    name_cell = line.get('menu_item_name') or 'Menu Item'
                    if line.get('menu_descriptor') and not simplified_view:
                        name_cell = f"{name_cell}<br/><font color=\"#4F5B6E\">{escape(line.get('menu_descriptor'))}</font>"
                    lines_data.append(
                        [
                            p('[ ]', tiny_style),
                            Paragraph(name_cell, base_style),
                            p(f"{format_number(line.get('quantity'))} {to_text(line.get('quantity_unit'), '')}".strip(), tiny_style),
                            p(line.get('notes') or '—', tiny_style),
                        ]
                    )
                if len(lines_data) == 1:
                    lines_data.append([p('', base_style, ''), p('No menu lines', base_style), p('', base_style, ''), p('', base_style, '')])
                lines_table = Table(lines_data, colWidths=[0.26 * inch, 2.9 * inch, 1.0 * inch, 1.54 * inch], repeatRows=1)
                style_table(lines_table, align_right_cols=[2])
                story.append(lines_table)
                story.append(Spacer(1, 6))
        else:
            story.append(Paragraph('No events found in this date range.', muted_style))

    story.append(PageBreak())
    story.append(Paragraph('Event Summary', section_title_style))
    events = datasets.get('events') or []
    if events:
        summary_data = [
            [
                Paragraph('Date', table_header_style),
                Paragraph('Event', table_header_style),
                Paragraph('Service Timing', table_header_style),
                Paragraph('Guests', table_header_style),
            ]
        ]
        for event in events:
            summary_data.append(
                [
                    p(event.get('event_date') or '—', base_style),
                    p(event.get('name') or 'Event', base_style),
                    p(event.get('service_timing') or '—', base_style),
                    p(event.get('guests') or '—', base_style),
                ]
            )
        summary_table = Table(summary_data, colWidths=[1.0 * inch, 2.3 * inch, 2.0 * inch, 0.4 * inch], repeatRows=1)
        style_table(summary_table)
        story.append(summary_table)
    else:
        story.append(Paragraph('No events found in this date range.', muted_style))

    if not print_lite:
        if include_shopping:
            story.append(PageBreak())
            story.append(Paragraph('Shopping List', section_title_style))
            vendor_groups = defaultdict(list)
            for item in datasets.get('shopping_ingredients', []) or []:
                vendor_groups[item.get('vendor') or 'Unassigned Vendor'].append(item)
            if vendor_groups:
                for vendor, items in vendor_groups.items():
                    story.append(Paragraph(vendor, subheading_style))
                    shop_data = [
                        [
                            Paragraph('✓', table_header_style),
                            Paragraph('Ingredient', table_header_style),
                            Paragraph('Qty', table_header_style),
                            Paragraph('Unit', table_header_style),
                            Paragraph('Category', table_header_style),
                        ]
                    ]
                    for item in items:
                        category = item.get('category') or 'Uncategorized'
                        if simplified_view:
                            category = '—'
                        shop_data.append(
                            [
                                p('[ ]', tiny_style),
                                p(item.get('name') or 'Ingredient', base_style),
                                p(item.get('display_quantity') if item.get('display_quantity') not in (None, '') else '—', base_style),
                                p(item.get('display_unit') or '—', base_style),
                                p(category, tiny_style),
                            ]
                        )
                    shop_table = Table(shop_data, colWidths=[0.24 * inch, 3.35 * inch, 0.8 * inch, 0.7 * inch, 1.21 * inch], repeatRows=1)
                    style_table(shop_table, align_right_cols=[2])
                    story.append(shop_table)
                    story.append(Spacer(1, 6))
            else:
                story.append(Paragraph('No shopping items for this date range.', muted_style))

        story.append(PageBreak())
        story.append(Paragraph('Weekly Prep Cards', section_title_style))
        if prep_groups:
            for group in prep_groups:
                story.append(Paragraph(f"Main Item: {to_text(group.get('main_label'), 'Menu Item')}", subheading_style))
                root = group.get('root') or {}
                root_title = f"{to_text(root.get('recipe_name'), 'Recipe')} | {to_text((root.get('display_required') or {}).get('quantity'), '—')} {to_text((root.get('display_required') or {}).get('unit'), '')}"
                story.append(Paragraph(root_title, base_style))
                story.append(
                    Paragraph(
                        f"Used in: {' · '.join(root.get('used_in_menu_items') or []) or '—'}",
                        muted_style,
                    )
                )
                story.append(Spacer(1, 3))

                ing_data = [
                    [
                        Paragraph('✓', table_header_style),
                        Paragraph('Ingredient', table_header_style),
                        Paragraph('Qty', table_header_style),
                        Paragraph('Unit', table_header_style),
                    ]
                ]
                for ing in root.get('ingredient_rows') or []:
                    ing_data.append(
                        [
                            p('[ ]', tiny_style),
                            p(ing.get('name') or 'Ingredient', base_style),
                            p(ing.get('display_quantity') if ing.get('display_quantity') not in (None, '') else '—', base_style),
                            p(ing.get('display_unit') or '—', base_style),
                        ]
                    )
                if len(ing_data) > 1:
                    ing_table = Table(ing_data, colWidths=[0.3 * inch, 3.9 * inch, 0.9 * inch, 0.75 * inch], repeatRows=1)
                    style_table(ing_table, align_right_cols=[2])
                    story.append(ing_table)
                else:
                    story.append(Paragraph('No direct raw ingredients - built from sub-recipes below.', muted_style))

                sub_rows = root.get('subrecipe_rows') or []
                if sub_rows:
                    story.append(Spacer(1, 4))
                    sub_data = [
                        [
                            Paragraph('✓', table_header_style),
                            Paragraph('Sub-Recipe', table_header_style),
                            Paragraph('Qty', table_header_style),
                            Paragraph('Unit', table_header_style),
                        ]
                    ]
                    for sub in sub_rows:
                        sub_name = sub.get('recipe_name') or 'Sub-Recipe'
                        if sub.get('covered_by_parent_recipe'):
                            covered = sub.get('covered_by_recipe_name') or 'another parent recipe'
                            sub_name = f"{sub_name}<br/><font color=\"#4F5B6E\">Prepared under {escape(covered)}</font>"
                        sub_data.append(
                            [
                                p('[ ]', tiny_style),
                                Paragraph(sub_name, base_style),
                                p((sub.get('display_required') or {}).get('quantity') or '—', base_style),
                                p((sub.get('display_required') or {}).get('unit') or '—', base_style),
                            ]
                        )
                    sub_table = Table(sub_data, colWidths=[0.3 * inch, 4.0 * inch, 0.75 * inch, 0.8 * inch], repeatRows=1)
                    style_table(sub_table, align_right_cols=[2])
                    story.append(sub_table)

                steps = step_lines(root)
                if steps:
                    story.append(Spacer(1, 4))
                    story.append(Paragraph('Method', subheading_style))
                    for idx, step in enumerate(steps, start=1):
                        story.append(Paragraph(f'{idx}. {html_text(step, "")}', base_style))

                for sub_card in group.get('sub_cards') or []:
                    story.append(Spacer(1, 5))
                    render_sub_card(story, sub_card, 1)

                story.append(Spacer(1, 8))
        else:
            story.append(Paragraph('No scaled recipe prep cards in this date window.', muted_style))

        weekly_menu_pulls = datasets.get('weekly_menu_pulls') or []
        if weekly_menu_pulls:
            story.append(PageBreak())
            story.append(Paragraph('Direct Menu Pulls', section_title_style))
            for pull in weekly_menu_pulls:
                story.append(Paragraph(to_text(pull.get('menu_item_name'), 'Menu Item'), subheading_style))
                if pull.get('used_in_event_names'):
                    story.append(Paragraph(', '.join(pull.get('used_in_event_names') or []), muted_style))
                pull_data = [
                    [
                        Paragraph('✓', table_header_style),
                        Paragraph('Ingredient', table_header_style),
                        Paragraph('Qty', table_header_style),
                        Paragraph('Unit', table_header_style),
                    ]
                ]
                for ing in pull.get('ingredient_rows') or []:
                    pull_data.append(
                        [
                            p('[ ]', tiny_style),
                            p(ing.get('name') or 'Ingredient', base_style),
                            p(ing.get('display_quantity') if ing.get('display_quantity') not in (None, '') else '—', base_style),
                            p(ing.get('display_unit') or '—', base_style),
                        ]
                    )
                if len(pull_data) == 1:
                    pull_data.append([p('', base_style, ''), p('No direct raw ingredients on this menu item.', base_style), p('', base_style, ''), p('', base_style, '')])
                pull_table = Table(pull_data, colWidths=[0.3 * inch, 4.05 * inch, 0.8 * inch, 0.7 * inch], repeatRows=1)
                style_table(pull_table, align_right_cols=[2])
                story.append(pull_table)
                story.append(Spacer(1, 6))

        story.append(PageBreak())
        story.append(Paragraph('Daily Prep', section_title_style))
        daily_groups = datasets.get('daily_groups') or []
        if daily_groups:
            for day in daily_groups:
                story.append(Paragraph(to_text(day.get('date')), subheading_style))
                for event in day.get('events') or []:
                    header = f"{to_text(event.get('name'), 'Event')} | {to_text(event.get('guests'), '—')} guests | {to_text(event.get('venue_name'), '—')}"
                    story.append(Paragraph(header, base_style))
                    line_data = [
                        [
                            Paragraph('✓', table_header_style),
                            Paragraph('Menu Item', table_header_style),
                            Paragraph('Qty', table_header_style),
                            Paragraph('Notes', table_header_style),
                        ]
                    ]
                    for line in event.get('lines') or []:
                        line_data.append(
                            [
                                p('[ ]', tiny_style),
                                p(line.get('menu_item_name') or 'Menu Item', base_style),
                                p(f"{format_number(line.get('quantity'))} {to_text(line.get('quantity_unit'), '')}".strip(), tiny_style),
                                p(line.get('notes') or '—', tiny_style),
                            ]
                        )
                    if len(line_data) > 1:
                        line_table = Table(line_data, colWidths=[0.3 * inch, 3.45 * inch, 0.95 * inch, 1.15 * inch], repeatRows=1)
                        style_table(line_table, align_right_cols=[2])
                        story.append(line_table)
                    story.append(Spacer(1, 4))
        else:
            story.append(Paragraph('No daily prep lines in this date range.', muted_style))

    doc.build(story, onFirstPage=draw_first_page, onLaterPages=draw_later_pages, canvasmaker=NumberedCanvas)
    payload = buffer.getvalue()
    buffer.close()

    filename = f"{make_safe_filename(venue_name)}_{start_date.isoformat()}_{end_date.isoformat()}_banquet_packet.pdf"
    return Response(
        payload,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )
